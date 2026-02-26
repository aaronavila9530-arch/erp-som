import os
import tempfile
import subprocess
from datetime import datetime, date
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from psycopg2.extras import RealDictCursor


class DraftSurveyExcelPdfService:
    """
    Genera XLSX desde template + convierte a PDF (solo hojas seleccionadas).
    Fuente de datos: draft_survey + draft_survey_ballast + general_draft_survey
    """

    # =========================================================
    # TEMPLATE PATH (BACKEND)
    # =========================================================
    TEMPLATE_PATH = os.path.abspath(
        os.path.join(
            os.path.dirname(__file__),
            "..",
            "templates",
            "draft_survey_template.xlsx"
        )
    )

    # =========================================================
    # SOLO HOJAS A INCLUIR (EN ORDEN)
    # =========================================================
    KEEP_SHEETS = [
        "General",
        "Draft",
        "ECE DRAUGHT SURVEY CODE",
        "Draught survey report ECE DRAUGHT SURVEY CODE D2",
    ]

    # =========================================================
    # MASTER MAPPING (REUSA TU LÓGICA)
    # OJO: Aquí puedes importar/pegar tu EXCEL_MAPPING actual.
    # =========================================================
    from services.draft_survey_excel_service import DraftSurveyExcelGenerator as _BaseGen  # reutilizamos mapping

    EXCEL_MAPPING = _BaseGen.EXCEL_MAPPING

    # =========================================================
    # DB: FETCH (3 TABLAS) → PAYLOAD
    # =========================================================
    def _fetch_payload_by_report_number(self, conn, draft_report_number: str) -> dict:

        draft_report_number = str(draft_report_number or "").strip()
        if not draft_report_number:
            raise ValueError("draft_report_number is required")

        cur = conn.cursor(cursor_factory=RealDictCursor)

        try:
            # -----------------------------
            # 1) draft_survey
            # -----------------------------
            cur.execute("""
                SELECT *
                FROM draft_survey
                WHERE draft_report_number = %s
                LIMIT 1
            """, (draft_report_number,))
            draft_row = cur.fetchone() or {}

            # -----------------------------
            # 2) draft_survey_ballast
            # -----------------------------
            cur.execute("""
                SELECT *
                FROM draft_survey_ballast
                WHERE draft_report_number = %s
                LIMIT 1
            """, (draft_report_number,))
            ballast_row = cur.fetchone() or {}

            # -----------------------------
            # 3) general_draft_survey
            # -----------------------------
            cur.execute("""
                SELECT *
                FROM general_draft_survey
                WHERE draft_report_number = %s
                LIMIT 1
            """, (draft_report_number,))
            general_row = cur.fetchone() or {}

        finally:
            cur.close()

        if not draft_row and not general_row and not ballast_row:
            return {}

        payload = {}
        payload.update(general_row or {})
        payload.update(draft_row or {})
        payload.update(ballast_row or {})

        # Normalizaciones mínimas
        payload["draft_report_number"] = draft_report_number

        # Alias para Draft sheet (tu mapping usa cargo/port_from/port_to)
        if "cargo" not in payload:
            payload["cargo"] = payload.get("init_cargo") or payload.get("final_cargo")
        if "port_from" not in payload:
            payload["port_from"] = payload.get("init_port_from") or payload.get("port_from")
        if "port_to" not in payload:
            payload["port_to"] = payload.get("init_port_to") or payload.get("port_to")

        # Alias para Deductions (ballast) porque tu mapping usa nombres tipo:
        # init_FPT_sounding, init_WBT 1P_sounding, init_FW P_height, etc.
        payload.update(self._build_ballast_aliases(ballast_row or {}))

        return payload

    # =========================================================
    # BALLEST: MAP KEYS DB -> TEMPLATE KEYS (TU MAPPING)
    # =========================================================
    def _build_ballast_aliases(self, row: dict) -> dict:
        """
        DB trae keys tipo:
          init_fpt_sounding, init_wbt_1p_sounding, init_fw_p_height, ...
        Template mapping usa keys tipo:
          init_FPT_sounding, init_WBT 1P_sounding, init_FW P_height, ...
        """
        out = {}

        def _tank_code_to_template(code: str) -> str:
            # fpt -> FPT, apt -> APT, wbt -> WBT
            return code.upper()

        def _side_to_template(side: str) -> str:
            return side.upper()

        for k, v in (row or {}).items():
            if v is None:
                continue

            key = str(k)

            # ---------- FPT / APT ----------
            # init_fpt_sounding -> init_FPT_sounding
            if key.startswith(("init_fpt_", "final_fpt_", "init_apt_", "final_apt_")):
                parts = key.split("_")
                prefix = parts[0]           # init/final
                tank = parts[1]             # fpt/apt
                rest = "_".join(parts[2:])  # sounding/volume/density
                out[f"{prefix}_{_tank_code_to_template(tank)}_{rest}"] = v
                continue

            # ---------- SLOP TANK ----------
            # init_slop_tank_sounding -> init_SLOP TANK_sounding
            if key.startswith(("init_slop_tank_", "final_slop_tank_")):
                parts = key.split("_")
                prefix = parts[0]
                rest = "_".join(parts[3:])  # sounding/volume/density
                out[f"{prefix}_SLOP TANK_{rest}"] = v
                continue

            # ---------- FW WASH ----------
            # init_fw_wash_sounding -> init_FW WASH_sounding
            if key.startswith(("init_fw_wash_", "final_fw_wash_")):
                parts = key.split("_")
                prefix = parts[0]
                rest = "_".join(parts[3:])
                out[f"{prefix}_FW WASH_{rest}"] = v
                continue

            # ---------- FRESH WATER HEIGHT/VOLUME ----------
            # init_fw_p_height -> init_FW P_height
            if key.startswith(("init_fw_", "final_fw_")):
                parts = key.split("_")
                prefix = parts[0]          # init/final
                fw = parts[1]              # fw
                tank = parts[2]            # p/s/dist
                rest = "_".join(parts[3:]) # height/volume
                out[f"{prefix}_FW {tank.upper()}_{rest}"] = v
                continue

            # ---------- WBT 1P..20S ----------
            # init_wbt_1p_sounding -> init_WBT 1P_sounding
            if key.startswith(("init_wbt_", "final_wbt_")):
                # init_wbt_1p_sounding
                parts = key.split("_")
                prefix = parts[0]   # init/final
                tank = parts[1]     # wbt
                numside = parts[2]  # 1p, 10s, 20p...
                rest = "_".join(parts[3:])  # sounding/volume/density

                # separar número y lado
                num = "".join(ch for ch in numside if ch.isdigit())
                side = "".join(ch for ch in numside if ch.isalpha())

                if num and side:
                    out[f"{prefix}_{_tank_code_to_template(tank)} {num}{_side_to_template(side)}_{rest}"] = v
                continue

        return out

    # =========================================================
    # SAFE SETTERS (MERGE SAFE + NUM SAFE)
    # =========================================================
    def _safe_set(self, ws: Worksheet, cell: str, value):

        if value in (None, ""):
            return

        # bool -> YES/NO
        if isinstance(value, bool):
            value = "YES" if value else "NO"

        # num strings -> float/int (best effort)
        try:
            if isinstance(value, str):
                vv = value.strip().replace(",", ".")
                if vv.replace(".", "", 1).isdigit():
                    value = float(vv) if "." in vv else int(vv)
        except Exception:
            pass

        # merged cells
        for merged in ws.merged_cells.ranges:
            if cell in merged:
                ws.cell(row=merged.min_row, column=merged.min_col).value = value
                return

        ws[cell].value = value

    def _safe_set_date(self, ws: Worksheet, cell: str, value):

        if not value:
            return

        parsed = None

        try:
            if isinstance(value, (datetime, date)):
                parsed = value

            elif isinstance(value, str):
                v = value.strip()

                # si viene "02-25-2026 12:00" => cortar fecha
                date_part = v.split(" ")[0]

                date_formats = [
                    "%m-%d-%Y",
                    "%d-%m-%Y",
                    "%Y-%m-%d",
                    "%m/%d/%Y",
                    "%d/%m/%Y",
                    "%Y/%m/%d",
                ]

                for fmt in date_formats:
                    try:
                        parsed = datetime.strptime(date_part, fmt)
                        break
                    except Exception:
                        continue

        except Exception:
            parsed = None

        if not parsed:
            return

        for merged in ws.merged_cells.ranges:
            if cell in merged:
                c = ws.cell(row=merged.min_row, column=merged.min_col)
                c.value = parsed
                c.number_format = "DD-MM-YYYY"
                return

        ws[cell].value = parsed
        ws[cell].number_format = "DD-MM-YYYY"

    # =========================================================
    # GENERATE XLSX (SOLO HOJAS KEEP_SHEETS)
    # =========================================================
    def generate_excel_by_report_number(self, conn, draft_report_number: str) -> str:

        if not os.path.exists(self.TEMPLATE_PATH):
            raise FileNotFoundError(f"Draft Survey template not found: {self.TEMPLATE_PATH}")

        payload = self._fetch_payload_by_report_number(conn, draft_report_number)

        if not payload:
            raise RuntimeError("Draft Survey record not found for that report number")

        wb = load_workbook(self.TEMPLATE_PATH)

        # 1) Llenar según mapping
        for sheet_name, config in (self.EXCEL_MAPPING or {}).items():

            if sheet_name not in wb.sheetnames:
                continue

            ws = wb[sheet_name]

            fields = (config or {}).get("fields", {}) or {}
            date_fields = set((config or {}).get("date_fields", []) or [])

            for key, cell in fields.items():

                value = (payload or {}).get(key)

                if key in date_fields:
                    self._safe_set_date(ws, cell, value)
                else:
                    self._safe_set(ws, cell, value)

        # 2) Eliminar hojas NO requeridas (esto “mergea” a un solo PDF final)
        for s in list(wb.sheetnames):
            if s not in self.KEEP_SHEETS:
                wb.remove(wb[s])

        # 3) Reordenar exactamente como KEEP_SHEETS
        for i, name in enumerate(self.KEEP_SHEETS):
            if name in wb.sheetnames:
                wb._sheets.insert(i, wb._sheets.pop(wb.sheetnames.index(name)))

        # 4) Guardar temp xlsx
        tmp_dir = tempfile.mkdtemp(prefix="draft_excel_")
        out_xlsx = os.path.join(tmp_dir, f"draft_survey_{draft_report_number}.xlsx")
        wb.save(out_xlsx)

        if not os.path.exists(out_xlsx) or os.path.getsize(out_xlsx) == 0:
            raise RuntimeError("Excel was not generated")

        return out_xlsx

    # =========================================================
    # XLSX -> PDF (LIBREOFFICE)
    # =========================================================
    def convert_excel_to_pdf(self, excel_path: str) -> str:

        if not excel_path or not os.path.exists(excel_path):
            raise RuntimeError("Excel path invalid or missing")

        soffice_path = os.getenv("LIBREOFFICE_PATH", "soffice")

        output_dir = tempfile.mkdtemp(prefix="draft_excel_pdf_")
        libre_profile = tempfile.mkdtemp(prefix="lo_profile_")

        cmd = [
            soffice_path,
            "--headless",
            "--nologo",
            "--nodefault",
            "--nolockcheck",
            "--nofirststartwizard",
            "--norestore",
            f"-env:UserInstallation=file://{libre_profile}",
            "--convert-to",
            "pdf",
            "--outdir",
            output_dir,
            excel_path
        ]

        result = subprocess.run(
            cmd,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True
        )

        if result.returncode != 0:
            raise RuntimeError(f"LibreOffice PDF conversion failed:\n{result.stderr}")

        pdf_name = Path(excel_path).with_suffix(".pdf").name
        pdf_path = os.path.join(output_dir, pdf_name)

        if not os.path.exists(pdf_path):
            raise RuntimeError("PDF was not created")

        if os.path.getsize(pdf_path) == 0:
            raise RuntimeError("PDF was generated but is empty")

        return pdf_path

    # =========================================================
    # PUBLIC: GENERATE PDF BY REPORT NUMBER
    # =========================================================
    def generate_pdf_by_report_number(self, conn, draft_report_number: str) -> str:
        xlsx_path = self.generate_excel_by_report_number(conn, draft_report_number)
        return self.convert_excel_to_pdf(xlsx_path)
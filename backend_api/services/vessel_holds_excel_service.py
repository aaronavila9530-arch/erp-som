import os
import tempfile
import subprocess
from pathlib import Path
from openpyxl import load_workbook


class VesselHoldsInspectionExcelService:

    # =========================================================
    # TEMPLATE PATH
    # =========================================================
    def _get_template_path(self):

        base_dir = Path(__file__).resolve().parent.parent

        template_path = base_dir / "templates" / "holds_inspection_certificate.xlsx"

        if not template_path.exists():
            raise FileNotFoundError(
                f"Excel template not found: {template_path}"
            )

        return template_path


    # =========================================================
    # BUILD EXCEL (INSERT DATA FROM DB)
    # =========================================================
    def _build_excel(self, data: dict) -> str:

        template_path = self._get_template_path()

        wb = load_workbook(template_path)

        try:

            if "data" not in wb.sheetnames:
                raise Exception("Sheet 'data' not found in template")

            ws = wb["data"]

            ordered_values = [

                data.get("id"),
                data.get("report_number"),
                data.get("port"),
                data.get("country"),
                data.get("vessel"),
                data.get("voyage"),
                data.get("load_port"),
                data.get("place"),
                data.get("installation"),
                data.get("product"),
                data.get("date"),
                data.get("inspection_time"),
                data.get("vessel_holds"),
                data.get("vessel_holds_status"),
                data.get("cargo_holds"),
                data.get("accepted_time"),
                data.get("place_location"),
                data.get("place_date"),
                data.get("hose_test_start"),
                data.get("hose_test_end"),
                data.get("remarks"),
                data.get("created_at"),
                data.get("updated_at"),
                data.get("status"),
                data.get("master_chief_officer")

            ]

            for idx, value in enumerate(ordered_values, start=1):

                ws.cell(row=idx, column=2, value=value)

            tmp_dir = tempfile.mkdtemp(prefix="holds_excel_")

            excel_path = os.path.join(
                tmp_dir,
                f"holds_inspection_certificate_{data.get('id')}.xlsx"
            )

            wb.save(excel_path)

            return excel_path

        finally:
            wb.close()


    # =========================================================
    # CONVERT EXCEL → PDF (LibreOffice)
    # =========================================================
    def _convert_excel_to_pdf(self, excel_path: str) -> str:

        if not excel_path or not os.path.exists(excel_path):
            raise FileNotFoundError("Excel file not found for PDF conversion")

        output_dir = tempfile.mkdtemp(prefix="holds_pdf_")

        command = [
            "soffice",
            "--headless",
            "--nologo",
            "--nodefault",
            "--nolockcheck",
            "--nofirststartwizard",
            "--norestore",
            "--convert-to",
            "pdf",
            excel_path,
            "--outdir",
            output_dir
        ]

        result = subprocess.run(
            command,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True
        )

        if result.returncode != 0:

            raise RuntimeError(
                f"LibreOffice conversion failed:\n{result.stderr or result.stdout}"
            )

        pdf_files = list(Path(output_dir).glob("*.pdf"))

        if not pdf_files:
            raise RuntimeError("PDF file was not generated")

        return str(pdf_files[0])


    # =========================================================
    # PUBLIC — GENERATE PDF
    # =========================================================
    def generate_pdf(self, data: dict) -> str:

        if not isinstance(data, dict):
            raise ValueError("Invalid payload for PDF generation")

        excel_path = self._build_excel(data)

        # 🔴 PREPARE SHEET BEFORE EXPORT
        self._prepare_print(excel_path)

        pdf_path = self._convert_excel_to_pdf(excel_path)

        if not pdf_path or not os.path.exists(pdf_path):
            raise RuntimeError("PDF generation failed")

        return pdf_path

    # =========================================================
    # PUBLIC — GENERATE EXCEL
    # =========================================================
    def generate_excel(self, data: dict) -> str:

        if not isinstance(data, dict):
            raise ValueError("Invalid payload for Excel generation")

        excel_path = self._build_excel(data)

        if not excel_path or not os.path.exists(excel_path):
            raise RuntimeError("Excel generation failed")

        return excel_path


    # =========================================================
    # PREPARE PRINT (ONLY CERTIFICATE SHEET)
    # =========================================================
    def _prepare_print(self, excel_path: str):

        wb = load_workbook(excel_path)

        try:

            if "VESSEL HOLDS INSPECTION CERTIFI" not in wb.sheetnames:
                raise Exception("Certificate sheet not found")

            for ws in wb.worksheets:

                if ws.title == "VESSEL HOLDS INSPECTION CERTIFI":
                    ws.sheet_state = "visible"
                    wb.active = wb.index(ws)

                    # fuerza area de impresion
                    ws.print_area = "A1:K60"

                else:
                    ws.sheet_state = "hidden"

            wb.calculation.fullCalcOnLoad = True

            wb.save(excel_path)

        finally:
            wb.close()
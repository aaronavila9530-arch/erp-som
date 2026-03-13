import os
import tempfile
import subprocess
from pathlib import Path
from openpyxl import load_workbook


class SamplingCertificateExcelService:

    # =========================================================
    # TEMPLATE PATH
    # =========================================================
    def _get_template_path(self):

        base_dir = Path(__file__).resolve().parent.parent

        template_path = base_dir / "templates" / "sampling_certificate_template.xlsx"

        if not template_path.exists():
            raise FileNotFoundError(
                f"Excel template not found: {template_path}"
            )

        return template_path


    # =========================================================
    # BUILD EXCEL (INSERT DATA FROM DB)
    # =========================================================
    def _build_excel(self, data: dict) -> str:

        template_path = self._get_template_path()

        wb = load_workbook(template_path)

        try:

            if "data" not in wb.sheetnames:
                raise Exception("Sheet 'data' not found in template")

            ws = wb["data"]

            ordered_values = [

                data.get("id"),
                data.get("report_no"),
                data.get("port"),
                data.get("country"),
                data.get("customer"),
                data.get("certificate_no"),
                data.get("vessel"),
                data.get("date"),
                data.get("place"),
                data.get("cargo"),
                data.get("holds_inspected"),
                data.get("hold_1_seal"),
                data.get("hold_2_seal"),
                data.get("hold_3_seal"),
                data.get("hold_4_seal"),
                data.get("hold_5_seal"),
                data.get("hold_6_seal"),
                data.get("hold_7_seal"),
                data.get("hold_8_seal"),
                data.get("hold_9_seal"),
                data.get("hold_10_seal"),
                data.get("observations"),
                data.get("closing_date"),
                data.get("closing_time"),
                data.get("master"),
                data.get("created_at"),
                data.get("updated_at"),
                data.get("status")

            ]

            # B1 hacia abajo
            for idx, value in enumerate(ordered_values, start=1):

                ws.cell(row=idx, column=2, value=value)

            tmp_dir = tempfile.mkdtemp(prefix="sampling_excel_")

            excel_path = os.path.join(
                tmp_dir,
                f"sampling_certificate_{data.get('id')}.xlsx"
            )

            wb.save(excel_path)

            return excel_path

        finally:
            wb.close()


    # =========================================================
    # CONVERT EXCEL → PDF (LibreOffice)
    # =========================================================
    def _convert_excel_to_pdf(self, excel_path: str) -> str:

        if not excel_path or not os.path.exists(excel_path):
            raise FileNotFoundError("Excel file not found for PDF conversion")

        output_dir = tempfile.mkdtemp(prefix="sampling_pdf_")

        command = [
            "soffice",
            "--headless",
            "--nologo",
            "--nodefault",
            "--nolockcheck",
            "--nofirststartwizard",
            "--norestore",
            "--convert-to",
            "pdf",
            excel_path,
            "--outdir",
            output_dir
        ]

        result = subprocess.run(
            command,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True
        )

        if result.returncode != 0:

            raise RuntimeError(
                f"LibreOffice conversion failed:\n{result.stderr or result.stdout}"
            )

        pdf_files = list(Path(output_dir).glob("*.pdf"))

        if not pdf_files:
            raise RuntimeError("PDF file was not generated")

        return str(pdf_files[0])


    # =========================================================
    # PUBLIC — GENERATE EXCEL
    # =========================================================
    def generate_excel(self, data: dict) -> str:

        if not isinstance(data, dict):
            raise ValueError("Invalid payload for Excel generation")

        excel_path = self._build_excel(data)

        if not excel_path or not os.path.exists(excel_path):
            raise RuntimeError("Excel generation failed")

        return excel_path


    # =========================================================
    # PUBLIC — GENERATE PDF
    # =========================================================
    def generate_pdf(self, data: dict) -> str:

        if not isinstance(data, dict):
            raise ValueError("Invalid payload for PDF generation")

        excel_path = self._build_excel(data)

        # preparar impresión
        self._prepare_print(excel_path)

        pdf_path = self._convert_excel_to_pdf(excel_path)

        if not pdf_path or not os.path.exists(pdf_path):
            raise RuntimeError("PDF generation failed")

        return pdf_path


    # =========================================================
    # PREPARE PRINT (ONLY CERTIFICATE SAMPLING SHEET)
    # =========================================================
    def _prepare_print(self, excel_path: str):

        wb = load_workbook(excel_path)

        try:

            if "CERTIFICATE SAMPLING" not in wb.sheetnames:
                raise Exception("Sheet 'CERTIFICATE SAMPLING' not found")

            for ws in wb.worksheets:

                if ws.title == "CERTIFICATE SAMPLING":

                    ws.sheet_state = "visible"
                    wb.active = wb.index(ws)

                    ws.print_area = "A1:K60"

                else:
                    ws.sheet_state = "hidden"

            wb.calculation.fullCalcOnLoad = True

            wb.save(excel_path)

        finally:
            wb.close()
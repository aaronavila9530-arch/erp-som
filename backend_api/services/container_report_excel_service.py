import os
import tempfile
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from datetime import datetime
from openpyxl.styles import Alignment

TEMPLATE_PATH = os.path.abspath(
    os.path.join(
        os.path.dirname(__file__),
        "..",
        "templates",
        "container_report_template.xlsx"
    )
)


class ContainerReportExcelGenerator:

    def _safe_set(self, ws: Worksheet, cell: str, value):
        for merged in ws.merged_cells.ranges:
            if cell in merged:
                ws.cell(
                    row=merged.min_row,
                    column=merged.min_col
                ).value = value
                return
        ws[cell].value = value

    def _safe_set_date(self, ws: Worksheet, cell: str, value):
        if not value:
            return

        try:
            if isinstance(value, str):
                value = datetime.fromisoformat(value)
        except Exception:
            pass

        for merged in ws.merged_cells.ranges:
            if cell in merged:
                c = ws.cell(
                    row=merged.min_row,
                    column=merged.min_col
                )
                c.value = value
                c.number_format = 'DD-MM-YYYY"  "HH:MM'
                return

        ws[cell].value = value
        ws[cell].number_format = 'DD-MM-YYYY"  "HH:MM'


    def _safe_hyperlink(
        self,
        ws: Worksheet,
        cell: str,
        url: str
    ):
        for merged in ws.merged_cells.ranges:
            if cell in merged:
                c = ws.cell(
                    row=merged.min_row,
                    column=merged.min_col
                )
                c.value = url
                c.hyperlink = url
                c.alignment = Alignment(
                    horizontal="left",
                    vertical="center"
                )
                return

        ws[cell].value = url
        ws[cell].hyperlink = url
        ws[cell].alignment = Alignment(
            horizontal="left",
            vertical="center"
        )

    def _check(self, ws: Worksheet, cell: str, flag: bool):
        self._safe_set(ws, cell, "✔" if flag else "")


# =====================================================
# MAIN GENERATOR
# =====================================================
def generate_container_report_excel(report: dict) -> str:

    generator = ContainerReportExcelGenerator()

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active

    # =====================================================
    # REPORT LINK (AGREGADO)
    # =====================================================
    generator._safe_set(ws, "AD3", report.get("linked_report_number"))
    generator._safe_set(ws, "Q3", report.get("container_type_text"))

    generator._safe_set(ws, "C5", report.get("report_no"))
    generator._safe_set(ws, "C6", report.get("bl"))
    generator._safe_set(ws, "C7", report.get("seals"))
    generator._safe_set(ws, "C8", report.get("appointment"))
    generator._safe_set(ws, "D9", report.get("shippers"))

    generator._safe_set(ws, "P5", report.get("inspection_place"))
    generator._safe_set(ws, "P6", report.get("contact_person"))
    generator._safe_set(ws, "X8", report.get("on_behalf_of"))
    generator._safe_set(ws, "X9", report.get("consignee_notify"))

    generator._safe_set(ws, "AA5", report.get("vessel"))

    # =====================================================
    # DATES — FORMAT DD-MM-YYYY (SAFE)
    # =====================================================
    generator._safe_set_date(ws, "AC6", report.get("contact_datetime"))
    generator._safe_set_date(ws, "P7", report.get("init_inspection_datetime"))
    generator._safe_set_date(ws, "V7", report.get("init_to"))
    generator._safe_set_date(ws, "AC7", report.get("final_inspection_datetime"))
    generator._safe_set_date(ws, "AI7", report.get("final_to"))

    # =====================================================
    # CONTAINER DESCRIPTION
    # =====================================================
    generator._check(ws, "A12", report.get("container_size_20"))
    generator._check(ws, "A13", report.get("container_size_40"))

    generator._check(ws, "E12", report.get("container_type_dry"))
    generator._check(ws, "E13", report.get("container_type_reefer"))
    generator._check(ws, "I12", report.get("container_type_iso"))
    generator._check(ws, "I13", report.get("container_type_flat_rack"))

    generator._check(ws, "N12", report.get("container_load_fcl"))
    generator._check(ws, "N13", report.get("container_load_lcl"))

    # =====================================================
    # CAUSE OF INSPECTION
    # =====================================================
    generator._check(ws, "Q12", report.get("cause_seals_bl"))
    generator._check(ws, "Q13", report.get("cause_change_seals"))
    generator._check(ws, "W12", report.get("cause_customs"))
    generator._check(ws, "W13", report.get("cause_transfer"))
    generator._check(ws, "AB12", report.get("cause_leaking"))
    generator._check(ws, "AB13", report.get("cause_damage"))
    generator._check(ws, "AG12", report.get("cause_stuff_condition"))
    generator._check(ws, "AG13", report.get("cause_stuff_condition"))

    generator._safe_set(ws, "I14", report.get("cause_detail"))

    # =====================================================
    # GOODS & PACKAGES
    # =====================================================
    generator._safe_set(ws, "B17", report.get("goods_description"))

    generator._check(ws, "U17", report.get("package_carton"))
    generator._check(ws, "U18", report.get("package_bags"))
    generator._check(ws, "U19", report.get("package_boxes"))
    generator._check(ws, "Y17", report.get("package_drums"))
    generator._check(ws, "Y18", report.get("package_pallets"))
    generator._check(ws, "Y19", report.get("package_bulk"))
    generator._check(ws, "AB17", report.get("package_bales"))
    generator._check(ws, "AB18", report.get("package_crates"))
    generator._check(ws, "AB19", report.get("package_other"))

    generator._safe_set(ws, "AF17", report.get("qty_1_left"))
    generator._safe_set(ws, "AI17", report.get("qty_1_right"))
    generator._safe_set(ws, "AF18", report.get("qty_2_left"))
    generator._safe_set(ws, "AI18", report.get("qty_2_right"))
    generator._safe_set(ws, "AF19", report.get("qty_3_left"))
    generator._safe_set(ws, "AI19", report.get("qty_3_right"))

    generator._safe_set(ws, "B22", report.get("package_marking"))
    generator._safe_set(ws, "B25", report.get("goods_condition"))

    # =====================================================
    # NARRATIVES
    # =====================================================
    generator._safe_set(ws, "B27", report.get("damage_details"))
    generator._safe_set(ws, "B31", report.get("remarks"))
    generator._safe_set(ws, "B37", report.get("conclusion"))

    picture_link = report.get("picture_link")
    if picture_link:
        generator._safe_hyperlink(ws, "B42", picture_link)

    # =====================================================
    # DOCUMENTS
    # =====================================================
    generator._check(ws, "A42", report.get("doc_bl"))
    generator._check(ws, "A43", report.get("doc_packing_list"))
    generator._check(ws, "A44", report.get("doc_shipping_invoice"))
    generator._check(ws, "A45", report.get("doc_cargo_manifest"))
    generator._check(ws, "A46", report.get("doc_commercial_invoice"))
    generator._check(ws, "A47", report.get("doc_delivery_record"))
    generator._check(ws, "A48", report.get("doc_notice_loss"))
    generator._check(ws, "A49", report.get("doc_insurance_policy"))
    generator._check(ws, "A50", report.get("doc_other"))

    # =====================================================
    # QUALITY
    # =====================================================
    generator._check(ws, "J42", report.get("quality_packing_exam"))
    generator._check(ws, "J43", report.get("quality_un_witness"))
    generator._check(ws, "J44", report.get("quality_visual_exam"))
    generator._check(ws, "J45", report.get("quality_product_exam"))
    generator._check(ws, "J46", report.get("quality_documents"))
    generator._check(ws, "J47", report.get("quality_sanitary_cert"))
    generator._check(ws, "J48", report.get("quality_phytosanitary_cert"))
    generator._check(ws, "J49", report.get("quality_factory_cert"))
    generator._check(ws, "J50", report.get("quality_origin_cert"))

    # =====================================================
    # INSPECTED CONTAINER
    # =====================================================
    generator._safe_set(ws, "W45", report.get("ic_manuf"))
    generator._safe_set(ws, "W46", report.get("ic_csc"))
    generator._safe_set(ws, "X47", report.get("ic_max_gw"))
    generator._safe_set(ws, "X48", report.get("ic_tare"))

    # =====================================================
    # GENERAL DETAILS
    # =====================================================
    generator._check(ws, "P55", report.get("new_commodity"))
    generator._check(ws, "V55", report.get("used_commodity"))
    generator._safe_set(ws, "W52", report.get("net_weight"))
    generator._safe_set(ws, "W53", report.get("gross_weight"))
    generator._safe_set(ws, "W54", report.get("volume"))

    # =====================================================
    # TRANSFER TO CONTAINER
    # =====================================================
    generator._safe_set(ws, "AF45", report.get("tr_number"))
    generator._safe_set(ws, "AF46", report.get("tr_manuf"))
    generator._safe_set(ws, "AF47", report.get("tr_csc"))
    generator._safe_set(ws, "AF48", report.get("tr_seal"))
    generator._safe_set(ws, "AG49", report.get("tr_max_gw"))
    generator._safe_set(ws, "AG50", report.get("tr_tare"))

    # =====================================================
    # SCOPE OF INSPECTION
    # =====================================================
    generator._check(ws, "AB52", report.get("scope_100"))
    generator._check(ws, "AB53", report.get("scope_random"))
    generator._safe_set(ws, "AB54", report.get("scope_items"))

    # =====================================================
    # PERSONS PRESENT
    # =====================================================
    generator._safe_set(ws, "B56", report.get("person_1_name"))
    generator._safe_set(ws, "N56", report.get("person_1_position"))

    generator._safe_set(ws, "B57", report.get("person_2_name"))
    generator._safe_set(ws, "N57", report.get("person_2_position"))

    generator._safe_set(ws, "B58", report.get("person_3_name"))
    generator._safe_set(ws, "N58", report.get("person_3_position"))

    # =====================================================
    # SAVE TEMP FILE
    # =====================================================
    fd, output_path = tempfile.mkstemp(
        suffix=".xlsx",
        prefix=f"container_report_{report.get('id', 'x')}_"
    )
    os.close(fd)

    wb.save(output_path)
    return output_path

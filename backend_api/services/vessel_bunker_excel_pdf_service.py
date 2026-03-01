import os
import tempfile
import subprocess
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins


class VesselBunkerExcelPdfService:

    REQUIRED_SHEETS_ORDER = [
        "CERTIFICATE",
        "CALCULATIONS",
        "LOG BOOK FIGURES"
    ]

    SCALE_MAP = {
        "CERTIFICATE": 70,
        "CALCULATIONS": 80,
        "LOG BOOK FIGURES": 80
    }

    # =========================================================
    # MAIN — Router calls this
    # =========================================================
    def generate_pdf_from_excel(self, excel_path: str) -> str:

        if not excel_path:
            raise ValueError("excel_path is required")

        if not os.path.exists(excel_path):
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        # 1) Prepare workbook print settings (WITHOUT breaking formulas)
        self._prepare_print(excel_path)

        # 2) Convert to PDF
        pdf_path = self._convert_excel_to_pdf(excel_path)

        if not pdf_path or not os.path.exists(pdf_path):
            raise RuntimeError("PDF generation failed")

        return pdf_path

    # =========================================================
    # PREPARE PRINT — keep formulas + hide non-required sheets
    # =========================================================
    def _prepare_print(self, excel_path: str):

        wb = load_workbook(excel_path)

        try:
            # -------------------------------------------------
            # Validate required sheets exist
            # -------------------------------------------------
            missing = [s for s in self.REQUIRED_SHEETS_ORDER if s not in wb.sheetnames]
            if missing:
                raise Exception(f"Missing required sheets: {', '.join(missing)}")

            # -------------------------------------------------
            # Hide non-required sheets (DO NOT REMOVE)
            # -------------------------------------------------
            for ws in wb.worksheets:
                if ws.title in self.REQUIRED_SHEETS_ORDER:
                    ws.sheet_state = "visible"
                else:
                    ws.sheet_state = "hidden"

            # -------------------------------------------------
            # Reorder: required sheets first, keep others after
            # -------------------------------------------------
            visible = [wb[s] for s in self.REQUIRED_SHEETS_ORDER]
            rest = [wb[s] for s in wb.sheetnames if s not in self.REQUIRED_SHEETS_ORDER]
            wb._sheets = visible + rest

            # -------------------------------------------------
            # Configure each printable sheet
            # -------------------------------------------------
            for ws in visible:
                self._fix_invalid_print_area_defined_name(wb, ws)

                # If template already has a print_area and it's valid -> keep it
                # Otherwise, set a sane print_area (A1:last_cell)
                current_pa = None
                try:
                    current_pa = ws.print_area
                except Exception:
                    current_pa = None

                if not current_pa or "#N/A" in str(current_pa):
                    # 🔥 FIX: do NOT use ws.max_row/max_column (can include formatted/empty rows)
                    # Instead detect last REAL cell with value so CERTIFICATE doesn't create blank page 2.
                    area = self._calc_real_print_area(ws)
                    self._safe_set_print_area(wb, ws, area)

                # 🔥 Ensure no manual page breaks force an extra blank page
                try:
                    ws.row_breaks = []
                    ws.col_breaks = []
                except Exception:
                    pass

                # --- Page setup: 1 page only ---
                ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
                ws.page_setup.fitToWidth = 1
                ws.page_setup.fitToHeight = False  # 🔥 forces ONLY page 1

                # Scale per sheet (Excel uses either fitTo* or scale; we keep both set)
                ws.page_setup.scale = int(self.SCALE_MAP.get(ws.title, 80))

                # Margins (close to left, vertical alignment)
                ws.page_margins = PageMargins(
                    left=0.15,
                    right=0.15,
                    top=0.25,
                    bottom=0.25,
                    header=0.10,
                    footer=0.10
                )

                ws.page_setup.horizontalCentered = False
                ws.page_setup.verticalCentered = False

            wb.save(excel_path)

        finally:
            wb.close()

    # =========================================================
    # Compute print area based on REAL content (values only)
    # =========================================================
    def _calc_real_print_area(self, ws) -> str:
        """
        Returns an A1:?? range based on the last cell that truly has a value.
        This prevents "phantom" rows/cols (formatting/old content) from expanding print area
        and creating a blank second page (common on CERTIFICATE templates).
        """
        last_row = 1
        last_col = 1

        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if v is None:
                    continue
                if isinstance(v, str) and v.strip() == "":
                    continue

                if cell.row > last_row:
                    last_row = cell.row
                if cell.column > last_col:
                    last_col = cell.column

        col_letter = get_column_letter(max(1, last_col))
        last_row = max(1, last_row)

        return f"A1:{col_letter}{last_row}"

    # =========================================================
    # Remove/repair invalid Print_Area defined name (#N/A)
    # =========================================================
    def _fix_invalid_print_area_defined_name(self, wb, ws):
        """
        Some templates store print areas as defined names.
        If that defined name is broken (#N/A), openpyxl warns and printing fails.
        We'll remove any broken Print_Area for this sheet.
        """
        try:
            defined = wb.defined_names.get("Print_Area")
            if not defined:
                return

            # defined is a DefinedName, may contain multiple destinations
            # We'll filter destinations for this sheet and drop ones with #N/A
            new_dests = []
            for title, coord in list(defined.destinations):
                if title != ws.title:
                    new_dests.append((title, coord))
                    continue

                # same sheet
                if coord and "#N/A" not in str(coord):
                    new_dests.append((title, coord))
                # else: drop it

            # If nothing left for Print_Area, remove name completely
            if not new_dests:
                try:
                    del wb.defined_names["Print_Area"]
                except Exception:
                    pass
                return

            # Otherwise rebuild the defined name destinations
            # Easiest: delete and re-add via safe setter later if needed
            try:
                del wb.defined_names["Print_Area"]
            except Exception:
                pass

            # Re-add remaining destinations by setting print_area on those sheets later,
            # BUT for our case we only care about this ws; so we just let _safe_set_print_area handle it.

        except Exception:
            # Never crash just for print area cleanup
            return

    # =========================================================
    # Safe set print area without inheriting broken defined name
    # =========================================================
    def _safe_set_print_area(self, wb, ws, area: str):
        """
        Ensures ws.print_area is a valid range and does not keep a broken defined name.
        """
        if not area or "#N/A" in str(area):
            return

        # Attempt to clear any per-sheet Print_Area leftovers first
        try:
            # openpyxl stores Print_Area as a defined name; removing global and resetting is safest
            if "Print_Area" in wb.defined_names:
                # don't nuke other sheets; just let set overwrite properly
                pass
        except Exception:
            pass

        try:
            ws.print_area = area
        except Exception:
            # As a fallback, don't block PDF generation
            return

    # =========================================================
    # CONVERT (LibreOffice)
    # =========================================================
    def _convert_excel_to_pdf(self, excel_path: str) -> str:

        output_dir = tempfile.mkdtemp(prefix="bunker_pdf_")

        command = [
            "soffice",
            "--headless",
            "--nologo",
            "--nodefault",
            "--nolockcheck",
            "--nofirststartwizard",
            "--norestore",
            "--convert-to",
            "pdf",
            excel_path,
            "--outdir",
            output_dir
        ]

        result = subprocess.run(
            command,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True
        )

        if result.returncode != 0:
            raise RuntimeError(
                f"LibreOffice conversion failed:\n{result.stderr or result.stdout}"
            )

        pdf_files = list(Path(output_dir).glob("*.pdf"))
        if not pdf_files:
            raise RuntimeError("PDF not created.")

        return str(pdf_files[0])
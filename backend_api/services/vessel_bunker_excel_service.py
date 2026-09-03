import os
import tempfile
from datetime import datetime, date
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


TEMPLATE_PATH = os.path.abspath(
    os.path.join(
        os.path.dirname(__file__),
        "..",
        "templates",
        "on_off_spot_template.xlsx"
    )
)


class VesselBunkerExcelGenerator:

    # =========================================================
    # MASTER ORDER (DEBE COINCIDIR EXACTAMENTE CON EL TEMPLATE)
    # ORDEN = FILAS EN LA HOJA "Data from DB"
    # =========================================================
    COLUMN_ORDER = [
        "id",
        "bunker_cert_no",
        "ship_name",
        "port_of_registry",
        "gross_tonnage",
        "report_date",
        "certificate",
        "report_category",
        "client",
        "port",
        "country",
        "berthing_date",
        "commenced_date",
        "dslop_date",
        "dslop_port",
        "dslop_country",
        "bunker_delivery_declared",
        "rob_diff",
        "plus_consumption",
        "generator_until_aps",
        "cons_dept",
        "me_to_sea_buoy",
        "remarks",
        "draft",
        "draft_fwd",
        "draft_aft",
        "trim",
        "list",
    ]

    # VLSFO 1–20
    for i in range(1, 21):
        COLUMN_ORDER.extend([
            f"vlsfo_tank_{i}_name",
            f"vlsfo_tank_{i}_dist_mtrs",
            f"vlsfo_tank_{i}_gauge_mtrs",
            f"vlsfo_tank_{i}_volume_m3",
            f"vlsfo_tank_{i}_temp_c",
            f"vlsfo_tank_{i}_temp_f",
            f"vlsfo_tank_{i}_density_15c",
            f"vlsfo_tank_{i}_weight_mt",
        ])

    # MGO 1–20
    for i in range(1, 21):
        COLUMN_ORDER.extend([
            f"mgo_tank_{i}_name",
            f"mgo_tank_{i}_dist_mtrs",
            f"mgo_tank_{i}_gauge_mtrs",
            f"mgo_tank_{i}_volume_m3",
            f"mgo_tank_{i}_temp_c",
            f"mgo_tank_{i}_temp_f",
            f"mgo_tank_{i}_density_15c",
            f"mgo_tank_{i}_weight_mt",
        ])

    # Bunker Figures 1–10
    for i in range(1, 11):
        COLUMN_ORDER.extend([
            f"bunker_figure_{i}_name",
            f"bunker_figure_{i}_ifo",
            f"bunker_figure_{i}_vlsfo",
            f"bunker_figure_{i}_lsmgo",
        ])

    # EXTRA FIELDS
    COLUMN_ORDER.extend([
        "antecedent_arrived_dt",
        "antecedent_survey_date_from",
        "antecedent_survey_date_to",
        "inspection_with",
        "workflow_status",
        "status",
        "created_at",
        "updated_at",
    ])

    # =========================================================
    # SAFE SETTER
    # =========================================================
    def _coerce_excel_value(self, value):
        if value in [None, ""]:
            return None

        if isinstance(value, bool):
            return "YES" if value else "NO"

        if isinstance(value, (int, float, datetime, date)):
            return value

        if isinstance(value, str):
            text = value.strip()
            if not text:
                return None

            candidate = text.replace("\u00a0", "").replace(" ", "")
            if "," in candidate and "." in candidate:
                if candidate.rfind(",") > candidate.rfind("."):
                    candidate = candidate.replace(".", "").replace(",", ".")
                else:
                    candidate = candidate.replace(",", "")
            elif "," in candidate:
                candidate = candidate.replace(",", ".")

            try:
                return float(candidate)
            except ValueError:
                return text

        return value

    def _safe_set(self, ws: Worksheet, cell: str, value):

        value = self._coerce_excel_value(value)
        if value is None:
            return

        ws[cell].value = value

    def _restore_calculation_formulas(self, wb):
        if "CALCULATIONS" not in wb.sheetnames:
            return

        ws = wb["CALCULATIONS"]

        for row in range(11, 22):
            ws[f"H{row}"] = f"=G{row}*1.8+32"
            ws[f"J{row}"] = f"=ROUND(F{row}*(I{row}-(0.00063*(G{row}-15))),2)"

        for row in range(29, 32):
            ws[f"H{row}"] = f"=G{row}*1.8+32"
            ws[f"J{row}"] = f"=ROUND(F{row}*(I{row}-(0.00063*(G{row}-15))),2)"

        ws["I6"] = "=\"TRIM=\" &  G6-E6"
        ws["C22"] = '=+J22&" MT "'
        ws["J22"] = "=SUM(J11:J21)"
        ws["C32"] = '=+J32 &" MT "'
        ws["J32"] = "=SUM(J29:J31)"
        ws["E34"] = "=J22+J32"

        try:
            ws.print_area = "A1:K46"
        except Exception:
            pass

    def _repair_broken_print_areas(self, wb):
        try:
            defined = wb.defined_names.get("Print_Area")
            if not defined:
                return

            if "#N/A" in str(defined.value):
                try:
                    del wb.defined_names["Print_Area"]
                except Exception:
                    return
        except Exception:
            return

    # =========================================================
    # GENERATE (OFICIAL)
    # =========================================================
    def generate(self, payload: dict) -> str:

        if not os.path.exists(TEMPLATE_PATH):
            raise FileNotFoundError("on_off_spot_template.xlsx not found.")

        wb = load_workbook(TEMPLATE_PATH)

        if "Data from DB" not in wb.sheetnames:
            raise Exception("Sheet 'Data from DB' not found in template.")

        ws = wb["Data from DB"]

        for row in range(2, ws.max_row + 1):
            field_name = ws[f"A{row}"].value
            if not field_name:
                continue

            field_name = str(field_name).strip()
            value = (payload or {}).get(field_name)
            self._safe_set(ws, f"B{row}", value)

        tmp_dir = tempfile.mkdtemp(prefix="bunker_excel_")
        tmp_path = os.path.join(tmp_dir, "vessel_bunker_report.xlsx")

        self._repair_broken_print_areas(wb)
        self._restore_calculation_formulas(wb)
        wb.save(tmp_path)

        return tmp_path

    # =========================================================
    # GENERATE PREVIEW (NO DB — VISUALIZAR DESDE FORM)
    # =========================================================
    def generate_preview(self, payload: dict) -> str:

        if not isinstance(payload, dict):
            raise ValueError("Invalid payload for preview.")

        return self.generate(payload)


# =========================================================
# SERVICE WRAPPER (🔥 ESTO ERA LO QUE FALTABA)
# =========================================================
class VesselBunkerExcelService:
    """
    Wrapper service requerido por el router.
    No toca DB.
    Solo usa el generator.
    """

    def __init__(self):
        self.generator = VesselBunkerExcelGenerator()

    def generate_excel_from_payload(self, payload: dict) -> str:

        if not isinstance(payload, dict):
            raise ValueError("Invalid payload for Excel preview")

        file_path = self.generator.generate_preview(payload)

        if not file_path or not os.path.exists(file_path):
            raise Exception("Excel file was not generated")

        return file_path

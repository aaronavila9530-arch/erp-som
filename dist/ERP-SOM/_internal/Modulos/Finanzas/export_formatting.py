from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


INVOICE_TEXT_KEYS = {
    "factura",
    "numero_factura",
    "num_factura",
    "numero_documento",
    "número documento",
    "numero documento",
    "documento",
    "número",
    "numero",
}


def as_export_text(value):
    if value is None:
        return ""
    return str(value).strip()


def is_invoice_text_column(column_name):
    name = str(column_name or "").strip().lower()
    if name in INVOICE_TEXT_KEYS:
        return True
    if "factura" in name and "fecha" not in name and "valor" not in name:
        return True
    return name in {"documento", "numero_documento", "número documento", "numero documento"}


def normalize_invoice_text_columns(ws, columns, min_width=28):
    invoice_indexes = [
        idx for idx, column_name in enumerate(columns, start=1)
        if is_invoice_text_column(column_name)
    ]

    for col_idx in invoice_indexes:
        letter = get_column_letter(col_idx)
        current_width = ws.column_dimensions[letter].width or 0
        ws.column_dimensions[letter].width = max(current_width, min_width)

        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = as_export_text(cell.value)
            cell.data_type = "s"
            cell.number_format = "@"
            cell.alignment = Alignment(horizontal="left")

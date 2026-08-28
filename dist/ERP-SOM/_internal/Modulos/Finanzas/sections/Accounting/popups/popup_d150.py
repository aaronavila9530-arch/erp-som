# -*- coding: utf-8 -*-
import tkinter as tk
from datetime import date
from tkinter import ttk, filedialog, messagebox

from api_client import get_accounting_iva_api, get_tax_book_api, get_tax_iva_api


PURPLE = "#312f91"
TEXT = "#28313d"
MUTED = "#646b74"
BG = "#f7f7f8"
CARD = "#ffffff"
BORDER = "#e2e2e6"
PANEL = "#f4f6fb"
INPUT_BG = "#ffffff"


class PopupD150(tk.Toplevel):
    """Formulario visual 150 - Impuesto al valor agregado."""

    COMPANY_ID = "3102920372"
    COMPANY_NAME = "MSL MARINE SURVEYORS AND LOGISTICS GROUP\nSOCIEDAD DE RESPONSABILIDAD LIMITADA"

    SALES_RATE_FIELDS = [
        ("Tarifa 0.5%", "sales_005", 0.005),
        ("Tarifa 1%", "sales_01", 0.01),
        ("Tarifa 2%", "sales_02", 0.02),
        ("Tarifa 3%", "sales_03", 0.03),
        ("Tarifa 4%", "sales_04", 0.04),
        ("Tarifa 4% (Servicios aéreos internacionales exclusivos)", "sales_04_air", 0.04),
        ("Tarifa 13%", "sales_13", 0.13),
    ]
    PURCHASE_RATE_FIELDS = [
        ("Compras a 0.5%", "purchase_005", 0.005),
        ("Compras a 1%", "purchase_01", 0.01),
        ("Compras a 2%", "purchase_02", 0.02),
        ("Compras a 3%", "purchase_03", 0.03),
        ("Compras a 4%", "purchase_04", 0.04),
        ("Compras a 13%", "purchase_13", 0.13),
    ]
    OTHER_SALES = [
        ("Total ventas exentas con derecho a crédito pleno", "sales_exempt_credit"),
        ("Total ventas exoneradas con derecho a crédito pleno", "sales_exonerated_credit"),
        ("Total ventas no sujetas con derecho a crédito pleno", "sales_non_subject_credit"),
        ("Total ventas exentas sin derecho a crédito", "sales_exempt_no_credit"),
        ("Total ventas exoneradas sin derecho a crédito", "sales_exonerated_no_credit"),
        ("Total ventas no sujetas sin derecho a crédito", "sales_non_subject_no_credit"),
        ("Total ventas de bienes de capital exentas sin derecho a crédito", "sales_capital_exempt_no_credit"),
    ]
    OTHER_PURCHASES = [
        ("Bienes y servicios adquiridos a otros regímenes", "purchase_other_regimes"),
        ("Bienes y servicios exentos", "purchase_exempt"),
        ("Bienes y servicios no sujetos", "purchase_non_subject"),
        ("Compras autorizadas sin impuesto", "purchase_authorized_no_tax"),
        ("Bienes y servicios según artículo 19 de la LIVA", "purchase_article19"),
        ("Compras bajo el sistema especial", "purchase_special"),
    ]

    def __init__(self, parent, period):
        super().__init__(parent)
        self.source_period = period or date.today().strftime("%Y-%m")
        self.period = self.source_period
        self.step = 0
        self.title(f"150 - Impuesto al valor agregado ({self.period})")
        self.geometry("1320x740")
        self.minsize(1120, 650)
        self.configure(bg=BG)

        self.amount_vars = {}
        self.summary_vars = {}
        self.step_labels = []
        self.content = None
        self.main_card = None
        self.canvas = None
        self.inner = None
        self.deferred_var = tk.StringVar(value="No lo utilizaré")
        self.data = {}
        self.automation_notes = []

        self._init_amount_vars()
        self._load_data()
        self._build_shell()
        self._render_step()
        self._refresh_summary()

    # ---------------- data ----------------
    def _init_amount_vars(self):
        keys = [key for _label, key, _rate in self.SALES_RATE_FIELDS + self.PURCHASE_RATE_FIELDS]
        keys.extend(key for _label, key in self.OTHER_SALES + self.OTHER_PURCHASES)
        keys.append("tax_self_assessed_debit")
        for key in keys:
            self.amount_vars[key] = tk.StringVar(value="0,00")

    def _load_data(self):
        sales_book = {}
        purchase_book = {}
        try:
            payload = get_tax_iva_api(self.period)
            fiscal = payload.get("fiscal") or {}
            sales_tax = float(fiscal.get("sales_tax") or 0)
            purchase_credit = float(fiscal.get("purchase_tax_credit") or 0)
            sales_total = float(fiscal.get("sales_total") or 0)
            purchase_total = float(fiscal.get("purchase_total") or 0)
            self.data = payload
            self.automation_notes.append(f"IVA cargado desde Centro fiscal para mes vencido {self.period}.")
        except Exception:
            try:
                legacy = get_accounting_iva_api(self.period)
                sales_tax = float(legacy.get("iva_por_pagar") or legacy.get("iva_total") or 0)
                purchase_credit = float(legacy.get("iva_credito") or 0)
                sales_total = 0.0
                purchase_total = 0.0
                self.data = {"legacy": legacy}
            except Exception as exc:
                messagebox.showerror("Formulario 150", f"No se pudo cargar IVA:\n{exc}", parent=self)
                sales_tax = purchase_credit = sales_total = purchase_total = 0.0

        try:
            sales_book = get_tax_book_api("SALE", self.period) or {}
            purchase_book = get_tax_book_api("PURCHASE", self.period) or {}
            self.automation_notes.append("Libros de ventas y compras usados para bases exentas/sin IVA.")
        except Exception as exc:
            self.automation_notes.append(f"Libros fiscales no disponibles para autollenado detallado: {exc}")

        sales_13 = sales_tax / 0.13 if sales_tax else 0.0
        sales_net_total = max(sales_total - sales_tax, 0.0)
        sales_exempt = self._book_exempt_total(sales_book) or max(sales_net_total - sales_13, 0.0)

        purchase_13 = purchase_credit / 0.13 if purchase_credit else 0.0
        purchase_net_total = max(purchase_total - purchase_credit, 0.0)
        purchase_exempt = self._book_exempt_total(purchase_book) or max(purchase_net_total - purchase_13, 0.0)

        self._set_amount("sales_13", sales_13)
        self._set_amount("sales_exempt_credit", sales_exempt)
        self._set_amount("purchase_13", purchase_13)
        self._set_amount("purchase_exempt", purchase_exempt)

    @staticmethod
    def _book_exempt_total(book):
        rows = (book or {}).get("data") or []
        total = 0.0
        for row in rows:
            try:
                total += float(row.get("exempt_amount") or 0)
            except Exception:
                pass
        return total

    # ---------------- UI shell ----------------
    def _build_shell(self):
        header = tk.Frame(self, bg=BG)
        header.pack(fill="x", padx=14, pady=(12, 16))
        tk.Button(
            header,
            text="<",
            font=("Segoe UI", 18, "bold"),
            width=2,
            bd=0,
            bg="#ededee",
            fg=TEXT,
            activebackground="#e5e5e7",
            command=self.destroy,
        ).pack(side="left")
        tk.Label(
            header,
            text="150 - Impuesto al valor agregado",
            font=("Segoe UI", 18, "bold"),
            bg=BG,
            fg=TEXT,
        ).pack(side="left", padx=14)
        tk.Label(
            header,
            text="Declaracion asistida, editable y exportable",
            font=("Segoe UI", 10),
            bg=BG,
            fg=MUTED,
        ).pack(side="left", padx=(0, 14))
        ttk.Button(header, text="Descargar Excel", command=self._export_excel).pack(side="right", padx=(8, 0))
        ttk.Button(header, text="Recalcular", command=self._render_step).pack(side="right", padx=(8, 0))
        tk.Label(
            header,
            text=f"Mes vencido: {self.period}",
            font=("Segoe UI", 10, "bold"),
            bg=BG,
            fg=MUTED,
        ).pack(side="right", padx=(0, 14))

        info = tk.Frame(self, bg="#eef4ff", padx=12, pady=7)
        info.pack(fill="x", padx=14, pady=(0, 8))
        tk.Label(
            info,
            text=(
                "Autollenado IVA: ventas, compras y credito fiscal desde Centro fiscal. "
                "Campos permanecen editables para conciliacion antes de presentar."
            ),
            bg="#eef4ff",
            fg="#254064",
            font=("Segoe UI", 9, "bold"),
        ).pack(anchor="w")

        body = tk.Frame(self, bg=BG)
        body.pack(fill="both", expand=True)

        sidebar = tk.Frame(body, bg=BG, width=205)
        sidebar.pack(side="left", fill="y", padx=(0, 10))
        sidebar.pack_propagate(False)
        self._build_stepper(sidebar)

        self.content = tk.Frame(body, bg=BG)
        self.content.pack(side="left", fill="both", expand=True)

        summary = tk.Frame(body, bg=BG, width=330)
        summary.pack(side="right", fill="y", padx=(14, 18))
        summary.pack_propagate(False)
        self._build_summary(summary)

    def _build_stepper(self, parent):
        steps = [
            "Ventas generales",
            "Pago diferido",
            "Compras totales",
            "Crédito fiscal",
            "Cálculo del\nimpuesto",
        ]
        for index, text in enumerate(steps):
            row = tk.Frame(parent, bg=BG)
            row.pack(fill="x", pady=(0, 10), padx=(6, 0))
            dot = tk.Label(
                row,
                text=str(index + 1),
                bg="#e8e8ee",
                fg=MUTED,
                font=("Segoe UI", 9, "bold"),
                width=3,
                pady=4,
            )
            dot.pack(side="left", anchor="n")
            label = tk.Label(row, text=text, bg=BG, fg=MUTED, justify="left", font=("Segoe UI", 10))
            label.pack(side="left", padx=(8, 0), anchor="w")
            self.step_labels.append((dot, label))

    def _build_summary(self, parent):
        card = tk.Frame(parent, bg=CARD, padx=18, pady=18)
        card.pack(fill="both", expand=True, pady=(14, 34))
        tk.Label(card, text="Resumen", font=("Segoe UI", 14, "bold"), bg=CARD, fg=TEXT).pack(anchor="w")

        info = tk.Frame(card, bg=CARD)
        info.pack(fill="x", pady=(20, 8))
        self._summary_text(info, "Identificación", self.COMPANY_ID)
        self._summary_text(info, "Nombre", self.COMPANY_NAME)
        self._summary_text(info, "Periodo", self.period.replace("-", ""))
        self._summary_text(info, "Declaración", "150 - Impuesto al valor agregado")
        dates = tk.Frame(info, bg=CARD)
        dates.pack(fill="x", pady=(10, 0))
        self._summary_text(dates, "Fecha inicio", self._period_start(), side="left")
        self._summary_text(dates, "Fecha fin", self._period_end(), side="left")
        tk.Label(card, text=". " * 34, bg=CARD, fg="#dddddf").pack(fill="x", pady=(10, 8))

        for label in [
            "Total monto del impuesto",
            "Total crédito fiscal",
            "Total gasto para utilidades",
            "Impuesto determinado",
            "Saldo a favor",
        ]:
            var = tk.StringVar(value="0,00")
            self.summary_vars[label] = var
            row = tk.Frame(card, bg=CARD)
            row.pack(fill="x", pady=8)
            tk.Label(row, text=label, bg=CARD, fg=MUTED, justify="left", font=("Segoe UI", 10)).pack(side="left")
            tk.Label(row, textvariable=var, bg=CARD, fg=TEXT, font=("Segoe UI", 10, "bold")).pack(side="right")

    def _summary_text(self, parent, title, value, side=None):
        box = tk.Frame(parent, bg=CARD)
        pack_args = {"anchor": "w", "pady": (0, 9)}
        if side:
            pack_args = {"side": side, "anchor": "n", "padx": (0, 16)}
        box.pack(**pack_args)
        tk.Label(box, text=title, bg=CARD, fg="#8a8a8f", font=("Segoe UI", 8)).pack(anchor="w")
        tk.Label(box, text=value, bg=CARD, fg=MUTED, justify="left", font=("Segoe UI", 9)).pack(anchor="w")

    # ---------------- steps ----------------
    def _render_step(self):
        self._refresh_summary()
        for widget in self.content.winfo_children():
            widget.destroy()
        for index, (dot, label) in enumerate(self.step_labels):
            active = index == self.step
            dot.configure(bg=PURPLE if active else "#e8e8ee", fg="white" if active else MUTED)
            label.configure(fg=PURPLE if active else MUTED, font=("Segoe UI", 10, "bold" if active else "normal"))

        self.main_card = tk.Frame(self.content, bg=CARD, padx=26, pady=22)
        self.main_card.pack(fill="both", expand=True, pady=(10, 30))

        self.canvas = tk.Canvas(self.main_card, bg=CARD, highlightthickness=0)
        scrollbar = ttk.Scrollbar(self.main_card, orient="vertical", command=self.canvas.yview)
        self.inner = tk.Frame(self.canvas, bg=CARD)
        self.inner.bind("<Configure>", lambda _e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
        self.canvas.create_window((0, 0), window=self.inner, anchor="nw")
        self.canvas.configure(yscrollcommand=scrollbar.set)
        self.canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        self.canvas.bind_all("<MouseWheel>", self._on_mousewheel)

        if self.step == 0:
            self._step_sales()
        elif self.step == 1:
            self._step_deferred()
        elif self.step == 2:
            self._step_purchases()
        elif self.step == 3:
            self._step_credit()
        else:
            self._step_tax()
        self._nav_buttons()

    def _step_sales(self):
        self._section_title("Ventas generales", "Declaración rectificada")
        self._paragraph(
            "En esta sección complete el importe de las ventas que ha realizado a cada una de las tarifas. "
            "El formulario calcula de forma automática el importe del monto del impuesto para cada una de ellas.\n\n"
            "Asimismo, si le corresponde, complete las casillas de las ventas sin impuesto que dan derecho a crédito pleno "
            "y las ventas sin impuesto que no dan derecho a crédito."
        )
        self._dotted()
        for title, key, rate in self.SALES_RATE_FIELDS:
            self._rate_group(title, f"Total ventas a {self._rate_label(rate)}", key, rate, "Monto de impuesto a")
        self._group_title("Otras ventas sin IVA con derecho a crédito pleno")
        for label, key in self.OTHER_SALES[:3]:
            self._amount_row(label, key)
        self._group_title("Otras ventas sin IVA sin derecho a crédito")
        for label, key in self.OTHER_SALES[3:]:
            self._amount_row(label, key)
        self._dotted()
        self._readonly_row("Total ventas generales", self._sales_total())
        self._readonly_row("Total ventas generales gravadas", self._sales_taxable_total())
        self._readonly_row("Monto del impuesto ventas generales", self._sales_tax_total())

    def _step_deferred(self):
        self._section_title("Pago diferido del impuesto por ventas a crédito del periodo a presentar o de periodos anteriores", "Declaración rectificada")
        self._paragraph(
            "En este apartado seleccione la opción “Sí lo utilizaré” si desea acogerse al esquema del pago diferido del "
            "impuesto por ventas a crédito o si requiere cancelar el impuesto de periodos anteriores bajo esta modalidad.\n\n"
            "Si marca la opción “Sí lo utilizaré”, debe completar los campos según corresponda; de lo contrario marque la opción “No lo utilizaré”."
        )
        self._dotted()
        row = tk.Frame(self.inner, bg=CARD)
        row.pack(fill="x", pady=22)
        tk.Label(
            row,
            text="Quiero realizar el pago diferido del\nimpuesto por ventas a crédito del periodo",
            bg=CARD,
            fg=TEXT,
            font=("Segoe UI", 11),
            justify="left",
        ).pack(side="left")
        tk.Label(row, textvariable=self.deferred_var, bg=CARD, fg=TEXT, font=("Segoe UI", 11)).pack(side="left", padx=34)
        ttk.Combobox(
            row,
            textvariable=self.deferred_var,
            values=["No lo utilizaré", "Sí lo utilizaré"],
            width=32,
            state="readonly",
        ).pack(side="right", padx=70)

    def _step_purchases(self):
        self._section_title("Compras totales", "Declaración rectificada")
        self._paragraph(
            "En esta sección incluya el monto de las compras realizadas en este periodo por cada tarifa. "
            "El sistema calcula de forma automática el impuesto soportado en cada una."
        )
        self._dotted()
        for title, key, rate in self.PURCHASE_RATE_FIELDS:
            self._rate_group(title, f"Total importe compras a {self._rate_label(rate)}", key, rate, "Impuesto soportado a")
        self._group_title("Compras sin IVA soportado o no acreditable")
        for label, key in self.OTHER_PURCHASES:
            self._amount_row(label, key)
        self._dotted()
        self._readonly_row("Total importe compras", self._purchase_total())
        self._readonly_row("Total importe compras sin IVA soportado o no acreditable", self._purchase_no_iva_total())
        self._readonly_row("Total importe de compras con IVA soportado", self._purchase_taxable_total())
        self._readonly_row("Total impuesto soportado", self._purchase_tax_total())

    def _step_credit(self):
        self._section_title("Crédito fiscal", "Declaración rectificada")
        self._paragraph(
            "En esta sección del formulario se calcula el crédito fiscal para el IVA de la siguiente forma:\n\n"
            "* Si usted vende a una única tarifa reducida y no es con derecho a crédito pleno, la tarifa de IVA aplicada es la tarifa menor de entre la tarifa de compras y la tarifa de ventas.\n"
            "* Si usted vende a una única tarifa que es con derecho a crédito pleno, la tarifa de IVA aplicada es la tarifa que ha soportado en sus compras.\n"
            "* Si usted vende a varias tarifas con derecho a crédito pleno exentas o no, la tarifa de IVA aplicada es la que ha soportado en sus compras.\n\n"
            "La diferencia entre el monto del impuesto soportado y el crédito fiscal para el IVA será el importe de costo o gasto para utilidades."
        )
        self._dotted()
        self._group_title("Compras a 13%")
        self._readonly_row("Total importe de compras con IVA soportado", self._purchase_taxable_total())
        self._readonly_row("Total impuesto soportado", self._purchase_tax_total())
        self._readonly_row("Total crédito fiscal del periodo", self._credit_period())
        self._readonly_row("Total importe de gasto para utilidades", self._utility_expense())

    def _step_tax(self):
        self._section_title("Cálculo del impuesto", "Declaración rectificada")
        self._paragraph("Se relaciona el monto total del impuesto calculado para el periodo")
        self._readonly_row("Monto del impuesto ventas generales", self._sales_tax_total())
        self._readonly_row("Total monto del impuesto", self._sales_tax_total())
        self._dotted()
        self._paragraph(
            "Se relacionan los créditos fiscales calculados para el periodo\n\n"
            "Para determinar el saldo a favor o el importe determinado para el impuesto se calcula la diferencia entre "
            "el total monto del impuesto y el total crédito fiscal para el IVA."
        )
        self._readonly_row("Total crédito fiscal para el IVA", self._credit_period())
        self._readonly_row("Total gasto para utilidades", self._utility_expense())
        self._amount_row("Débito por auto repercusión del impuesto", "tax_self_assessed_debit")
        self._readonly_row("Impuesto determinado", self._determined_tax(), bold=True)
        self._readonly_row("Saldo a favor", self._balance_credit(), bold=True)

    # ---------------- widgets ----------------
    def _section_title(self, title, right_text=None):
        row = tk.Frame(self.inner, bg=CARD)
        row.pack(fill="x", pady=(0, 20))
        tk.Label(row, text=title, bg=CARD, fg=TEXT, font=("Segoe UI", 15, "bold"), justify="left").pack(side="left", anchor="w")
        if right_text:
            tk.Label(
                row,
                text=right_text,
                bg="#eef4ff",
                fg="#254064",
                font=("Segoe UI", 9, "bold"),
                padx=10,
                pady=5,
            ).pack(side="right", padx=(0, 60))

    def _paragraph(self, text):
        tk.Label(self.inner, text=text, bg=CARD, fg=TEXT, font=("Segoe UI", 11), justify="left", wraplength=820).pack(anchor="w", pady=(0, 18))

    def _dotted(self):
        tk.Frame(self.inner, bg=BORDER, height=1).pack(fill="x", pady=(10, 20))

    def _group_title(self, title):
        band = tk.Frame(self.inner, bg=PANEL, padx=12, pady=8)
        band.pack(fill="x", pady=(18, 8))
        tk.Label(band, text=title, bg=PANEL, fg=TEXT, font=("Segoe UI", 11, "bold"), justify="left").pack(anchor="w")

    def _rate_group(self, title, total_label, key, rate, tax_prefix):
        self._group_title(title)
        self._amount_row(total_label, key)
        self._readonly_row(f"{tax_prefix} {self._rate_label(rate)}", self._amount(key) * rate)

    def _amount_row(self, label, key):
        row = tk.Frame(self.inner, bg="#fbfbfd", padx=12, pady=8, highlightbackground=BORDER, highlightthickness=1)
        row.pack(fill="x", pady=5, padx=(0, 36))
        tk.Label(row, text=label, bg="#fbfbfd", fg=TEXT, font=("Segoe UI", 10), wraplength=520, justify="left").pack(side="left", anchor="w")
        ent = tk.Entry(
            row,
            textvariable=self.amount_vars[key],
            width=18,
            justify="right",
            relief="solid",
            bd=1,
            bg=INPUT_BG,
            font=("Segoe UI", 10),
        )
        ent.pack(side="right", ipady=6)

    def _readonly_row(self, label, value, bold=False):
        row = tk.Frame(self.inner, bg="#fbfbfd", padx=12, pady=8, highlightbackground=BORDER, highlightthickness=1)
        row.pack(fill="x", pady=5, padx=(0, 36))
        font = ("Segoe UI", 12, "bold") if bold else ("Segoe UI", 11)
        tk.Label(row, text=label, bg="#fbfbfd", fg=TEXT, font=("Segoe UI", 10), wraplength=520, justify="left").pack(side="left", anchor="w")
        tk.Label(row, text=self._fmt(value), bg="#fbfbfd", fg=TEXT, font=font, width=18, anchor="e").pack(side="right")

    def _nav_buttons(self):
        bar = tk.Frame(self.content, bg=BG)
        bar.pack(fill="x", padx=18, pady=(0, 18))
        if self.step > 0:
            ttk.Button(bar, text="<  Anterior", command=self._back).pack(side="left", fill="x", expand=True, padx=(0, 10))
        ttk.Button(bar, text="Siguiente  >" if self.step < 4 else "Presentar", command=self._next_or_present).pack(
            side="right", fill="x", expand=True, padx=(10, 0)
        )

    def _on_mousewheel(self, event):
        if self.canvas and self.canvas.winfo_exists():
            self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    # ---------------- calculations ----------------
    def _amount(self, key):
        return self._parse(self.amount_vars[key].get())

    def _set_amount(self, key, value):
        self.amount_vars[key].set(self._fmt(value))

    def _sales_taxable_total(self):
        return sum(self._amount(key) for _label, key, _rate in self.SALES_RATE_FIELDS)

    def _sales_tax_total(self):
        return sum(self._amount(key) * rate for _label, key, rate in self.SALES_RATE_FIELDS)

    def _sales_no_tax_total(self):
        return sum(self._amount(key) for _label, key in self.OTHER_SALES)

    def _sales_total(self):
        return self._sales_taxable_total() + self._sales_no_tax_total()

    def _purchase_taxable_total(self):
        return sum(self._amount(key) for _label, key, _rate in self.PURCHASE_RATE_FIELDS)

    def _purchase_tax_total(self):
        return sum(self._amount(key) * rate for _label, key, rate in self.PURCHASE_RATE_FIELDS)

    def _purchase_no_iva_total(self):
        return sum(self._amount(key) for _label, key in self.OTHER_PURCHASES)

    def _purchase_total(self):
        return self._purchase_taxable_total() + self._purchase_no_iva_total()

    def _credit_period(self):
        return self._purchase_tax_total()

    def _utility_expense(self):
        return max(self._purchase_tax_total() - self._credit_period(), 0.0)

    def _determined_tax(self):
        return max(self._sales_tax_total() - self._credit_period() + self._amount("tax_self_assessed_debit"), 0.0)

    def _balance_credit(self):
        return max(self._credit_period() - self._sales_tax_total() - self._amount("tax_self_assessed_debit"), 0.0)

    def _refresh_summary(self):
        if not self.summary_vars:
            return
        self.summary_vars["Total monto del impuesto"].set(self._fmt(self._sales_tax_total()))
        self.summary_vars["Total crédito fiscal"].set(self._fmt(self._credit_period()))
        self.summary_vars["Total gasto para utilidades"].set(self._fmt(self._utility_expense()))
        self.summary_vars["Impuesto determinado"].set(self._fmt(self._determined_tax()))
        self.summary_vars["Saldo a favor"].set(self._fmt(self._balance_credit()))

    # ---------------- Excel ----------------
    def _export_excel(self):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
        except Exception as exc:
            messagebox.showerror("Formulario 150", f"No se pudo cargar openpyxl:\n{exc}", parent=self)
            return

        path = filedialog.asksaveasfilename(
            parent=self,
            title="Descargar D150 en Excel",
            defaultextension=".xlsx",
            initialfile=f"D150_IVA_{self.period.replace('-', '')}.xlsx",
            filetypes=[("Excel", "*.xlsx")],
        )
        if not path:
            return

        wb = Workbook()
        styles = {
            "title_fill": PatternFill("solid", fgColor="312F91"),
            "header_fill": PatternFill("solid", fgColor="E9E9F6"),
            "bold": Font(bold=True, color="28313D"),
            "white_bold": Font(bold=True, color="FFFFFF"),
            "money": '#,##0.00',
        }
        self._excel_summary(wb.active, styles)
        self._excel_dashboard(wb.create_sheet("Dashboard"), styles)
        self._excel_sales(wb.create_sheet("Ventas generales"), styles)
        self._excel_deferred(wb.create_sheet("Pago diferido"), styles)
        self._excel_purchases(wb.create_sheet("Compras totales"), styles)
        self._excel_credit(wb.create_sheet("Credito fiscal"), styles)
        self._excel_tax(wb.create_sheet("Calculo impuesto"), styles)
        self._excel_source(wb.create_sheet("Fuente automatica"), styles)
        self._excel_review(wb.create_sheet("Revision"), styles)

        for ws in wb.worksheets:
            for column_cells in ws.columns:
                width = 12
                col = get_column_letter(column_cells[0].column)
                for cell in column_cells:
                    width = max(width, min(len("" if cell.value is None else str(cell.value)) + 2, 58))
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                ws.column_dimensions[col].width = width
            ws.freeze_panes = "A3"

        try:
            wb.save(path)
        except PermissionError:
            messagebox.showerror("Formulario 150", "Cierre el Excel si está abierto y vuelva a intentar.", parent=self)
            return
        except Exception as exc:
            messagebox.showerror("Formulario 150", f"No se pudo generar el Excel:\n{exc}", parent=self)
            return
        messagebox.showinfo("Formulario 150", "Excel D150 generado correctamente.", parent=self)

    def _excel_dashboard(self, ws, styles):
        self._excel_title(ws, "Dashboard ejecutivo D150", styles, 4)
        self._excel_headers(ws, ["Indicador", "Monto", "Estado", "Comentario"], styles)
        rows = [
            ("Debito fiscal ventas", self._sales_tax_total(), "Calculado", "Base gravada por tarifa"),
            ("Credito fiscal compras", self._credit_period(), "Calculado", "Segun compras con IVA soportado"),
            ("Gasto para utilidades", self._utility_expense(), "Calculado", "IVA no acreditable"),
            ("Impuesto determinado", self._determined_tax(), "Por pagar" if self._determined_tax() else "Sin saldo por pagar", "Resultado D150"),
            ("Saldo a favor", self._balance_credit(), "A favor" if self._balance_credit() else "Sin saldo a favor", "Resultado D150"),
        ]
        for label, value, status, comment in rows:
            ws.append([label, value, status, comment])
            self._money_if_number(ws.cell(ws.max_row, 2), styles)

    def _excel_summary(self, ws, styles):
        ws.title = "D150 Resumen"
        self._excel_title(ws, "Formulario 150 - Impuesto al valor agregado", styles, 4)
        rows = [
            ("Identificación", self.COMPANY_ID),
            ("Nombre", self.COMPANY_NAME.replace("\n", " ")),
            ("Periodo", self.period.replace("-", "")),
            ("Mes seleccionado en ERP", self.source_period),
            ("Regla aplicada", "D150 se presenta contra mes vencido"),
            ("Declaración", "150 - Impuesto al valor agregado"),
            ("Fecha inicio", self._period_start()),
            ("Fecha fin", self._period_end()),
            ("Total monto del impuesto", self._sales_tax_total()),
            ("Total crédito fiscal", self._credit_period()),
            ("Total gasto para utilidades", self._utility_expense()),
            ("Impuesto determinado", self._determined_tax()),
            ("Saldo a favor", self._balance_credit()),
        ]
        ws.append([])
        for label, value in rows:
            ws.append([label, value])
            ws.cell(ws.max_row, 1).font = styles["bold"]
            self._money_if_number(ws.cell(ws.max_row, 2), styles)

    def _excel_sales(self, ws, styles):
        self._excel_title(ws, "Ventas generales", styles, 5)
        self._excel_headers(ws, ["Sección", "Rubro", "Base", "Tarifa", "Impuesto"], styles)
        for title, key, rate in self.SALES_RATE_FIELDS:
            ws.append([title, f"Total ventas a {self._rate_label(rate)}", self._amount(key), rate, self._amount(key) * rate])
            self._format_money_row(ws, [3, 5], styles)
            ws.cell(ws.max_row, 4).number_format = "0.00%"
        for label, key in self.OTHER_SALES:
            ws.append(["Ventas sin IVA", label, self._amount(key), 0, 0])
            self._format_money_row(ws, [3, 5], styles)
        self._excel_total(ws, "Total ventas generales", self._sales_total(), styles)
        self._excel_total(ws, "Total ventas generales gravadas", self._sales_taxable_total(), styles)
        self._excel_total(ws, "Monto del impuesto ventas generales", self._sales_tax_total(), styles)

    def _excel_deferred(self, ws, styles):
        self._excel_title(ws, "Pago diferido", styles, 3)
        self._excel_headers(ws, ["Campo", "Valor", "Comentario"], styles)
        ws.append(["Quiero realizar el pago diferido del impuesto por ventas a crédito del periodo", self.deferred_var.get(), "Seleccionable"])

    def _excel_purchases(self, ws, styles):
        self._excel_title(ws, "Compras totales", styles, 5)
        self._excel_headers(ws, ["Sección", "Rubro", "Base", "Tarifa", "Impuesto soportado"], styles)
        for title, key, rate in self.PURCHASE_RATE_FIELDS:
            ws.append([title, f"Total importe compras a {self._rate_label(rate)}", self._amount(key), rate, self._amount(key) * rate])
            self._format_money_row(ws, [3, 5], styles)
            ws.cell(ws.max_row, 4).number_format = "0.00%"
        for label, key in self.OTHER_PURCHASES:
            ws.append(["Compras sin IVA soportado", label, self._amount(key), 0, 0])
            self._format_money_row(ws, [3, 5], styles)
        self._excel_total(ws, "Total importe compras", self._purchase_total(), styles)
        self._excel_total(ws, "Total importe compras sin IVA soportado o no acreditable", self._purchase_no_iva_total(), styles)
        self._excel_total(ws, "Total importe de compras con IVA soportado", self._purchase_taxable_total(), styles)
        self._excel_total(ws, "Total impuesto soportado", self._purchase_tax_total(), styles)

    def _excel_credit(self, ws, styles):
        self._excel_title(ws, "Crédito fiscal", styles, 3)
        self._excel_headers(ws, ["Concepto", "Monto", "Nota"], styles)
        for label, value in [
            ("Total importe de compras con IVA soportado", self._purchase_taxable_total()),
            ("Total impuesto soportado", self._purchase_tax_total()),
            ("Total crédito fiscal del periodo", self._credit_period()),
            ("Total importe de gasto para utilidades", self._utility_expense()),
        ]:
            ws.append([label, value, "Calculado"])
            self._money_if_number(ws.cell(ws.max_row, 2), styles)

    def _excel_tax(self, ws, styles):
        self._excel_title(ws, "Cálculo del impuesto", styles, 3)
        self._excel_headers(ws, ["Concepto", "Monto", "Nota"], styles)
        rows = [
            ("Monto del impuesto ventas generales", self._sales_tax_total(), "Ventas"),
            ("Total monto del impuesto", self._sales_tax_total(), "Calculado"),
            ("Total crédito fiscal para el IVA", self._credit_period(), "Compras"),
            ("Total gasto para utilidades", self._utility_expense(), "Diferencia no acreditable"),
            ("Débito por auto repercusión del impuesto", self._amount("tax_self_assessed_debit"), "Editable"),
            ("Impuesto determinado", self._determined_tax(), "Resultado"),
            ("Saldo a favor", self._balance_credit(), "Resultado"),
        ]
        for label, value, note in rows:
            ws.append([label, value, note])
            self._money_if_number(ws.cell(ws.max_row, 2), styles)

    def _excel_review(self, ws, styles):
        self._excel_title(ws, "Checklist de revisión D150", styles, 4)
        self._excel_headers(ws, ["Estado", "Punto de revisión", "Responsable", "Comentario"], styles)
        for item in [
            "Validar ventas gravadas, exentas, exoneradas y no sujetas contra libro de ventas.",
            "Validar compras con IVA soportado contra libro de compras y XML aceptados.",
            "Confirmar crédito fiscal aplicado y gasto para utilidades.",
            "Comparar impuesto determinado o saldo a favor contra TRIBU-CR.",
            "Revisar documentos pendientes de Hacienda antes de presentar.",
        ]:
            ws.append(["Pendiente", item, "", ""])

    def _excel_source(self, ws, styles):
        self._excel_title(ws, "Fuente automatica y reglas IVA", styles, 4)
        self._excel_headers(ws, ["Tipo", "Detalle", "Regla", "Comentario"], styles)
        for note in self.automation_notes or ["Sin notas de autollenado."]:
            ws.append(["Autollenado", note, "Centro fiscal / libros fiscales", "Editable antes de presentar"])
        for label, _key, rate in self.SALES_RATE_FIELDS:
            ws.append(["IVA ventas", label, f"Base x {rate:.2%}", "Tarifa configurada en formulario D150"])
        for label, _key, rate in self.PURCHASE_RATE_FIELDS:
            ws.append(["IVA compras", label, f"Base x {rate:.2%}", "Tarifa configurada en formulario D150"])
        ws.append(["Resultado", "Impuesto determinado", "Debito fiscal - credito fiscal + auto repercusion", "Saldo positivo por pagar"])
        ws.append(["Resultado", "Saldo a favor", "Credito fiscal - debito fiscal - auto repercusion", "Saldo positivo a favor"])

    # ---------------- helpers ----------------
    def _next_or_present(self):
        self._refresh_summary()
        if self.step >= 4:
            messagebox.showinfo("Formulario 150", "Formulario D150 preparado. Use Descargar Excel para revisar el detalle.", parent=self)
            return
        self.step += 1
        self._render_step()

    def _back(self):
        self._refresh_summary()
        self.step = max(self.step - 1, 0)
        self._render_step()

    def _period_start(self):
        try:
            year, month = self.period.split("-")
            return f"01/{month}/{year}"
        except Exception:
            return ""

    def _period_end(self):
        import calendar

        try:
            year, month = [int(x) for x in self.period.split("-")]
            return f"{calendar.monthrange(year, month)[1]:02d}/{month:02d}/{year}"
        except Exception:
            return ""

    @staticmethod
    def _previous_period(period):
        try:
            year, month = [int(x) for x in str(period).split("-")[:2]]
        except Exception:
            today = date.today()
            year, month = today.year, today.month
        month -= 1
        if month == 0:
            month = 12
            year -= 1
        return f"{year:04d}-{month:02d}"

    @staticmethod
    def _rate_label(rate):
        value = rate * 100
        return f"{value:g}%"

    @staticmethod
    def _parse(text):
        clean = str(text or "").strip().replace("₡", "").replace(" ", "")
        if not clean:
            return 0.0
        if "," in clean and "." in clean:
            clean = clean.replace(".", "").replace(",", ".")
        elif "," in clean:
            clean = clean.replace(",", ".")
        try:
            return float(clean)
        except ValueError:
            return 0.0

    @staticmethod
    def _fmt(value):
        text = f"{float(value or 0):,.2f}"
        return text.replace(",", "X").replace(".", ",").replace("X", ".")

    @staticmethod
    def _excel_title(ws, text, styles, columns):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=columns)
        cell = ws.cell(1, 1, text)
        cell.fill = styles["title_fill"]
        cell.font = styles["white_bold"]

    @staticmethod
    def _excel_headers(ws, headers, styles):
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(ws.max_row, col)
            cell.fill = styles["header_fill"]
            cell.font = styles["bold"]

    @staticmethod
    def _money_if_number(cell, styles):
        if isinstance(cell.value, (int, float)):
            cell.number_format = styles["money"]

    def _format_money_row(self, ws, columns, styles):
        for col in columns:
            self._money_if_number(ws.cell(ws.max_row, col), styles)

    def _excel_total(self, ws, label, value, styles):
        ws.append(["TOTAL", label, value, None, None])
        ws.cell(ws.max_row, 1).font = styles["bold"]
        ws.cell(ws.max_row, 2).font = styles["bold"]
        ws.cell(ws.max_row, 3).font = styles["bold"]
        ws.cell(ws.max_row, 3).number_format = styles["money"]

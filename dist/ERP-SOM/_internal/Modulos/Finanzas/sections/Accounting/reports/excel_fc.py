from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from tkinter import filedialog
from typing import Any, Dict, List

from Modulos.Finanzas.sections.Accounting.reports.build_fc_from_trial_balance import (
    build_fc_from_trial_balance
)

THIN = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def export_fc_excel_from_fc(fc_or_rows: Any):
    """
    Exporta Estado de Flujo de Efectivo (Método Indirecto)

    ✔ Acepta FC ya construido (dict)
    ✔ Acepta accounting_lines (list) y construye FC automáticamente
    ✔ Save As interno
    ✔ Blindado total
    """

    # ==================================================
    # NORMALIZACIÓN DE ENTRADA
    # ==================================================
    if isinstance(fc_or_rows, list):
        fc = build_fc_from_trial_balance(fc_or_rows)
    elif isinstance(fc_or_rows, dict):
        fc = fc_or_rows
    else:
        raise ValueError(
            "export_fc_excel_from_fc esperaba un FC (dict) "
            "o accounting_lines (list)"
        )

    # ==================================================
    # VALIDACIÓN FUERTE DE FC
    # ==================================================
    required_keys = [
        "operacion",
        "neto_operacion",
        "inversion",
        "neto_inversion",
        "financiamiento",
        "neto_financiamiento",
        "variacion_efectivo",
        "efectivo_inicio",
        "efectivo_final",
        "period_label",
    ]

    for k in required_keys:
        if k not in fc:
            raise ValueError(f"Flujo de Efectivo inválido. Falta clave '{k}'")

    # ==================================================
    # NORMALIZADORES
    # ==================================================
    def safe_list(val) -> List[dict]:
        return val if isinstance(val, list) else []

    def safe_amount(val) -> float:
        try:
            return float(val)
        except Exception:
            return 0.0

    # ==================================================
    # SAVE AS
    # ==================================================
    file_path = filedialog.asksaveasfilename(
        title="Guardar Estado de Flujo de Efectivo",
        defaultextension=".xlsx",
        filetypes=[("Excel", "*.xlsx")]
    )
    if not file_path:
        return None

    # ==================================================
    # WORKBOOK
    # ==================================================
    wb = Workbook()
    ws = wb.active
    ws.title = "Flujo de Efectivo"

    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 20

    row = 1

    # ==================================================
    # ENCABEZADO
    # ==================================================
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.cell(row=row, column=1, value="ESTADO DE FLUJO DE EFECTIVO")
    ws.cell(row=row, column=1).font = Font(bold=True, size=14)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.cell(
        row=row,
        column=1,
        value=f"Por el período terminado el {fc['period_label']}"
    )
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
    row += 2

    # ==================================================
    # HELPERS
    # ==================================================
    def section(title: str):
        nonlocal row
        ws.cell(row=row, column=1, value=title).font = Font(bold=True)
        row += 1

    def line(label: str, amount: Any, bold: bool = False):
        nonlocal row
        amount = safe_amount(amount)

        ws.cell(row=row, column=1, value=str(label))
        ws.cell(row=row, column=2, value=round(amount, 2))
        ws.cell(row=row, column=2).number_format = "#,##0.00"
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="right")

        if bold:
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=2).font = Font(bold=True)

        row += 1

    def spacer():
        nonlocal row
        row += 1

    # ==================================================
    # OPERACIÓN
    # ==================================================
    section("FLUJO DE EFECTIVO DE ACTIVIDADES DE OPERACIÓN")
    for item in safe_list(fc["operacion"]):
        line(item.get("label", ""), item.get("amount", 0))

    line(
        "Flujo Neto de Actividades de Operación",
        fc["neto_operacion"],
        bold=True
    )
    spacer()
    spacer()

    # ==================================================
    # INVERSIÓN
    # ==================================================
    section("FLUJO DE EFECTIVO DE ACTIVIDADES DE INVERSIÓN")
    for item in safe_list(fc["inversion"]):
        line(item.get("label", ""), item.get("amount", 0))

    line(
        "Flujo Neto de Actividades de Inversión",
        fc["neto_inversion"],
        bold=True
    )
    spacer()
    spacer()

    # ==================================================
    # FINANCIAMIENTO
    # ==================================================
    section("FLUJO DE EFECTIVO DE ACTIVIDADES DE FINANCIAMIENTO")
    for item in safe_list(fc["financiamiento"]):
        line(item.get("label", ""), item.get("amount", 0))

    line(
        "Flujo Neto de Actividades de Financiamiento",
        fc["neto_financiamiento"],
        bold=True
    )
    spacer()
    spacer()

    # ==================================================
    # EFECTIVO
    # ==================================================
    line(
        "AUMENTO / (DISMINUCIÓN) NETO DE EFECTIVO",
        fc["variacion_efectivo"],
        bold=True
    )

    line("Efectivo al Inicio del Periodo", fc["efectivo_inicio"])
    line(
        "EFECTIVO AL FINAL DEL PERIODO",
        fc["efectivo_final"],
        bold=True
    )

    # ==================================================
    # BORDES (BLINDADO)
    # ==================================================
    for r in ws.iter_rows(min_row=1, max_row=row, min_col=1, max_col=2):
        for cell in r:
            if cell.value is not None:
                try:
                    cell.border = BORDER
                except Exception:
                    pass

    wb.save(file_path)
    return file_path


# 🔑 Alias para compatibilidad
export_fc_excel = export_fc_excel_from_fc

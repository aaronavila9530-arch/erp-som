from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from tkinter import filedialog
from typing import Any, Dict, List

from Modulos.Finanzas.sections.Accounting.reports.build_mayor_from_lines import (
    build_mayor_from_lines
)

THIN = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def export_mayor_excel_from_lines(rows_or_built: Any):
    """
    Exporta Libro Mayor a Excel

    ✔ Acepta mayor ya construido (dict)
    ✔ Acepta accounting_lines (list)
    ✔ Save As interno
    ✔ SIN company_code
    ✔ Style blindado
    """

    # ==================================================
    # NORMALIZACIÓN DE ENTRADA
    # ==================================================
    if isinstance(rows_or_built, list):
        mayor = build_mayor_from_lines(rows_or_built)
    elif isinstance(rows_or_built, dict):
        mayor = rows_or_built
    else:
        raise ValueError(
            "export_mayor_excel_from_lines esperaba dict o list"
        )

    # ==================================================
    # VALIDACIÓN
    # ==================================================
    if "accounts" not in mayor:
        raise ValueError("Libro Mayor inválido")

    # ==================================================
    # SAVE AS
    # ==================================================
    file_path = filedialog.asksaveasfilename(
        title="Guardar Libro Mayor",
        defaultextension=".xlsx",
        filetypes=[("Excel", "*.xlsx")]
    )
    if not file_path:
        return None

    # ==================================================
    # WORKBOOK
    # ==================================================
    wb = Workbook()
    ws = wb.active
    ws.title = "Libro Mayor"

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18

    row = 1

    # ==================================================
    # ENCABEZADO
    # ==================================================
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row=row, column=1, value="LIBRO MAYOR")
    ws.cell(row=row, column=1).font = Font(bold=True, size=14)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(
        row=row,
        column=1,
        value=f"Periodo {mayor['period_label']}"
    )
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
    row += 2

    # ==================================================
    # HELPERS
    # ==================================================
    def header():
        nonlocal row
        headers = ["Fecha", "Asiento", "Detalle", "Debe", "Haber"]
        ws.append(headers)
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        row += 1

    def line(values: list):
        nonlocal row
        ws.append(values)
        ws.cell(row=row, column=4).number_format = "#,##0.00"
        ws.cell(row=row, column=5).number_format = "#,##0.00"
        ws.cell(row=row, column=4).alignment = Alignment(horizontal="right")
        ws.cell(row=row, column=5).alignment = Alignment(horizontal="right")
        row += 1

    def spacer():
        nonlocal row
        row += 1

    # ==================================================
    # DATA
    # ==================================================
    for acc in mayor["accounts"]:

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.cell(row=row, column=1, value=acc["account"])
        ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1

        header()

        for l in acc["lines"]:
            line([
                l.get("date"),
                l.get("entry_id"),
                l.get("detail"),
                l.get("debit", 0),
                l.get("credit", 0),
            ])

        ws.append([
            "",
            "",
            "TOTAL",
            acc["total_debit"],
            acc["total_credit"],
        ])

        ws.cell(row=row, column=3).font = Font(bold=True)
        ws.cell(row=row, column=4).font = Font(bold=True)
        ws.cell(row=row, column=5).font = Font(bold=True)
        row += 2

    # ==================================================
    # BORDES (BLINDADO)
    # ==================================================
    for r in ws.iter_rows(min_row=1, max_row=row, min_col=1, max_col=5):
        for cell in r:
            if cell.value is not None:
                try:
                    cell.border = BORDER
                except Exception:
                    pass

    wb.save(file_path)
    return file_path


# 🔑 Alias compatible con el popup actual
export_mayor_excel = export_mayor_excel_from_lines

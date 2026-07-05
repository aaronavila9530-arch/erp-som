from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from tkinter import filedialog
from typing import Dict, Any, List


THIN = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def export_er_excel_from_er(
    er: Dict[str, Any],
    fiscal_year: int,
    period: int
):
    """
    Exporta Estado de Resultados a Excel
    desde un ER ya construido (build_er_from_lines)

    ✔ NO recibe accounting_lines
    ✔ NO recalcula
    ✔ SOLO renderiza
    """

    # ==================================================
    # VALIDACIÓN FUERTE
    # ==================================================
    if not isinstance(er, dict):
        raise ValueError("ER inválido: se esperaba dict.")

    required_keys = [
        "ingresos", "total_ingresos",
        "costos", "total_costos",
        "utilidad_bruta",
        "gastos_operativos", "total_gastos_operativos",
        "utilidad_operativa",
        "otros",
        "utilidad_antes_impuestos",
        "impuesto_renta",
        "utilidad_neta",
    ]

    for k in required_keys:
        if k not in er:
            raise ValueError(f"ER inválido. Falta clave '{k}'")

    # Asegurar listas
    def safe_list(v) -> List[dict]:
        return v if isinstance(v, list) else []

    ingresos = safe_list(er["ingresos"])
    costos = safe_list(er["costos"])
    gastos_operativos = safe_list(er["gastos_operativos"])
    otros = safe_list(er["otros"])

    # ==================================================
    # SAVE AS
    # ==================================================
    file_path = filedialog.asksaveasfilename(
        title="Guardar Estado de Resultados",
        defaultextension=".xlsx",
        filetypes=[("Excel", "*.xlsx")]
    )
    if not file_path:
        return None

    period_label = f"{int(fiscal_year)}-{int(period):02d}"

    wb = Workbook()
    ws = wb.active
    ws.title = "Estado de Resultados"

    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 20

    row = 1

    # ==================================================
    # ENCABEZADO (🔥 AQUÍ ESTABA EL ERROR)
    # ==================================================
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.cell(row=row, column=1, value="ESTADO DE RESULTADOS")
    ws.cell(row=row, column=1).font = Font(bold=True, size=14)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.cell(
        row=row,
        column=1,
        value=f"Por el período terminado el {period_label}"
    )
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
    row += 2

    # ==================================================
    # HELPERS
    # ==================================================
    def section(title):
        nonlocal row
        ws.cell(row=row, column=1, value=title).font = Font(bold=True)
        row += 1

    def line(label, amount, bold=False):
        nonlocal row
        try:
            amount = float(amount)
        except Exception:
            amount = 0.0

        ws.cell(row=row, column=1, value=str(label))
        ws.cell(row=row, column=2, value=round(amount, 2))
        ws.cell(row=row, column=2).number_format = "#,##0.00"
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="right")

        if bold:
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=2).font = Font(bold=True)

        row += 1

    def spacer():
        nonlocal row
        row += 1

    # ==================================================
    # INGRESOS
    # ==================================================
    section("INGRESOS")
    for item in ingresos:
        if isinstance(item, dict):
            line(item.get("label", ""), item.get("amount", 0))

    line("Total Ingresos", er["total_ingresos"], bold=True)
    spacer()

    # ==================================================
    # COSTOS
    # ==================================================
    section("(-) COSTOS")
    for item in costos:
        if isinstance(item, dict):
            line(item.get("label", ""), item.get("amount", 0))

    line("Total Costos", er["total_costos"], bold=True)
    spacer()

    # ==================================================
    # UTILIDAD BRUTA
    # ==================================================
    line("UTILIDAD BRUTA", er["utilidad_bruta"], bold=True)
    spacer()
    spacer()

    # ==================================================
    # GASTOS OPERATIVOS
    # ==================================================
    section("(-) GASTOS OPERATIVOS")
    for item in gastos_operativos:
        if isinstance(item, dict):
            line(item.get("label", ""), item.get("amount", 0))

    line("Total Gastos Operativos", er["total_gastos_operativos"], bold=True)
    spacer()

    # ==================================================
    # OTROS
    # ==================================================
    if otros:
        section("(+/-) OTROS INGRESOS Y GASTOS")
        for item in otros:
            if isinstance(item, dict):
                line(item.get("label", ""), item.get("amount", 0))
        spacer()

    # ==================================================
    # RESULTADOS FINALES
    # ==================================================
    line("UTILIDAD ANTES DE IMPUESTOS", er["utilidad_antes_impuestos"], bold=True)
    spacer()

    line("(-) IMPUESTO SOBRE LA RENTA", er["impuesto_renta"])
    spacer()

    line("UTILIDAD NETA DEL PERIODO", er["utilidad_neta"], bold=True)

    # ==================================================
    # BORDES
    # ==================================================
    for r in ws.iter_rows(min_row=1, max_row=row, min_col=1, max_col=2):
        for cell in r:
            if cell.value is not None:
                cell.border = BORDER

    wb.save(file_path)
    return file_path

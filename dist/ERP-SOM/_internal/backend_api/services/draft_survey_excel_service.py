import os
import tempfile
from datetime import datetime, date
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


TEMPLATE_PATH = os.path.abspath(
    os.path.join(
        os.path.dirname(__file__),
        "..",
        "templates",
        "draft_survey_template.xlsx"
    )
)


class DraftSurveyExcelGenerator:

    # =========================================================
    # MASTER MAPPING STRUCTURE (FULL TEMPLATE 1:1)
    # =========================================================
    EXCEL_MAPPING = {
        "General": {
            "date_fields": [],
            "fields": {

                # HEADER
                "vessel_mv": "X1",
                "survey_no": "X3",
                "call_letters": "AI3",
                "vessel_previous_names": "J6",

                # REGISTRY BLOCK
                "flag": "C8",
                "registry": "M8",
                "built_year": "W8",
                "by": "AA8",

                # OFFICERS
                "master": "H10",
                "chief_officer": "H11",
                "chief_engineer": "H12",
                "witness_draughts": "H13",
                "witness_sounding": "H14",

                # SURVEY INFO
                "initial_surveyors": "AB10",
                "final_surveyors": "AB11",
                "survey_requested_by": "AB12",
                "on_account_of": "AB13",
                "attended_also_by": "AB14",

                # LOCATIONS
                "init_ships_location": "H16",
                "final_ships_location": "AB16",

                # DIMENSIONS
                "length_overall": "M18",
                "length_between_pp": "M19",
                "extreme_breadth": "M20",
                "moulded_breadth": "M21",
                "depth_overall_incl_keel_plate": "M22",
                "moulded_depth": "M23",
                "summer_draught": "M24",
                "summer_freeboard": "M25",

                # CONSTANTS / TONNAGE
                "constant_declared": "AG18",
                "constant_calculated": "AG19",
                "light_displacement": "AG20",
                "light_shipweight_plan": "AG21",
                "summer_displacement": "AG22",
                "summer_deadweight": "AG23",
                "net_register_tons": "AG24",
                "gross_register_tons": "AG25",

                # HYDROSTATIC
                "hydro_tables_issued": "L32",
            }
        },

        "Draft": {
            "date_fields": ["init_date", "final_date"],
            "fields": {

                # =====================================================
                # INITIAL â€” TOP
                # =====================================================

                "init_date": "T2",
                "init_time_from": "T3",
                "init_time_to": "T4",
                "cargo": "AB2",
                "port_from": "AB3",
                "port_to": "AB4",

                # =====================================================
                # INITIAL â€” DRAFT READINGS
                # =====================================================

                "init_draft_fwd_port": "A8",
                "init_draft_fwd_stb": "F8",
                "init_draft_mid_port": "A9",
                "init_draft_mid_stb": "F9",
                "init_draft_aft_port": "A10",
                "init_draft_aft_stb": "F10",

                "init_draft_fwd_marks": "V8",
                "init_draft_mid_marks": "V9",
                "init_draft_aft_marks": "V10",

                "init_sg": "I13",
                "init_lpp": "M13",

                # =====================================================
                # INITIAL â€” FIGURES
                # =====================================================

                "init_tpc_p": "Q16",
                "init_tpc_s": "U16",
                "init_light_ship": "M29",
                "init_historic_constant": "M30",
                "init_bl_figure": "M35",

                # =====================================================
                # INITIAL â€” DEDUCTIONS
                # =====================================================

                "init_ballast": "G18",
                "init_fresh_water": "G19",
                "init_fuel_oil": "G20",
                "init_diesel_oil": "G21",
                "init_lub_oil": "G22",
                "init_slop": "G23",
                "init_swimming_pool": "G24",
                "init_others": "G25",

                # =====================================================
                # INITIAL â€” HYDRO 1 (CUADRO 1)
                # 4 filas EXACTAS:
                #   F1: draft/disp/tpc/lcf
                #   F2: draft/disp/tpc/lcf
                #   F3: draft_mtc / mtc+50 / mtc-50
                #   F4: mtc+50 / mtc-50
                # =====================================================

                # Fila 1
                "init_hydro1_draft_1": "BG27",
                "init_hydro1_disp_1":  "BH27",
                "init_hydro1_tpc_1":   "BI27",
                "init_hydro1_lcf_1":   "BJ27",

                # Fila 2
                "init_hydro1_draft_2": "BG29",
                "init_hydro1_disp_2":  "BH29",
                "init_hydro1_tpc_2":   "BI29",
                "init_hydro1_lcf_2":   "BJ29",

                # Fila 3
                "init_hydro1_draft_mtc":  "BG31",
                "init_hydro1_mtc_p50_1":  "BH31",
                "init_hydro1_mtc_m50_1":  "BJ31",

                # Fila 4
                "init_hydro1_mtc_p50_2":  "BH33",
                "init_hydro1_mtc_m50_2":  "BJ33",


                # =====================================================
                # INITIAL â€” HYDRO 2 (CUADRO 2)
                # MISMA ESTRUCTURA, SOLO CAMBIA EL BLOQUE DE CELDAS
                # =====================================================

                # Fila 1
                "init_hydro2_draft_1": "BG36",
                "init_hydro2_disp_1":  "BH36",
                "init_hydro2_tpc_1":   "BI36",
                "init_hydro2_lcf_1":   "BJ36",

                # Fila 2
                "init_hydro2_draft_2": "BG38",
                "init_hydro2_disp_2":  "BH38",
                "init_hydro2_tpc_2":   "BI38",
                "init_hydro2_lcf_2":   "BJ38",

                # Fila 3
                "init_hydro2_draft_mtc":  "BG40",
                "init_hydro2_mtc_p50_1":  "BH40",
                "init_hydro2_mtc_m50_1":  "BJ40",

                # Fila 4
                "init_hydro2_mtc_p50_2":  "BH42",
                "init_hydro2_mtc_m50_2":  "BJ42",

                "final_hydro1_draft_1": "BG27",
                "final_hydro1_disp_1":  "BH27",
                "final_hydro1_tpc_1":   "BI27",
                "final_hydro1_lcf_1":   "BJ27",
                "final_hydro1_draft_2": "BG29",
                "final_hydro1_disp_2":  "BH29",
                "final_hydro1_tpc_2":   "BI29",
                "final_hydro1_lcf_2":   "BJ29",
                "final_hydro1_draft_mtc":  "BG31",
                "final_hydro1_mtc_p50_1":  "BH31",
                "final_hydro1_mtc_m50_1":  "BJ31",
                "final_hydro1_mtc_p50_2":  "BH33",
                "final_hydro1_mtc_m50_2":  "BJ33",

                "final_hydro2_draft_1": "BG36",
                "final_hydro2_disp_1":  "BH36",
                "final_hydro2_tpc_1":   "BI36",
                "final_hydro2_lcf_1":   "BJ36",
                "final_hydro2_draft_2": "BG38",
                "final_hydro2_disp_2":  "BH38",
                "final_hydro2_tpc_2":   "BI38",
                "final_hydro2_lcf_2":   "BJ38",
                "final_hydro2_draft_mtc":  "BG40",
                "final_hydro2_mtc_p50_1":  "BH40",
                "final_hydro2_mtc_m50_1":  "BJ40",
                "final_hydro2_mtc_p50_2":  "BH42",
                "final_hydro2_mtc_m50_2":  "BJ42",

                # =====================================================
                # FINAL â€” TOP
                # =====================================================

                "final_date": "AS2",
                "final_time_from": "AS3",
                "final_time_to": "AS4",

                # =====================================================
                # FINAL â€” DRAFT READINGS
                # =====================================================

                "final_draft_fwd_port": "Z8",
                "final_draft_fwd_stb": "AD8",
                "final_draft_mid_port": "Z9",
                "final_draft_mid_stb": "AD9",
                "final_draft_aft_port": "Z10",
                "final_draft_aft_stb": "AD10",

                "final_draft_fwd_marks": "AU8",
                "final_draft_mid_marks": "AU9",
                "final_draft_aft_marks": "AU10",

                "final_sg": "AH13",
                "final_lpp": "AL13",

                # =====================================================
                # FINAL â€” FIGURES
                # =====================================================

                "final_tpc_p": "AP16",
                "final_tpc_s": "AT16",
                "final_light_ship": "AG29",
                "final_historic_constant": "AG30",
                "final_bl_figure": "AG35",

                # =====================================================
                # FINAL â€” DEDUCTIONS
                # =====================================================

                "final_fuel_oil": "AS20",
                "final_diesel_oil": "AS21",
                "final_lub_oil": "AS22",
                "final_slop": "AS23",
                "final_swimming_pool": "AS24",
                "final_others": "AS25",

                # =====================================================
                # FINAL â€” HYDRO 1 (CUADRO 1 â€” 4 FILAS)
                # =====================================================

                # Fila 1

                # Fila 2

                # Fila 3

                # Fila 4


                # =====================================================
                # FINAL â€” HYDRO 2 (CUADRO 2 â€” 4 FILAS)
                # =====================================================

                # Fila 1

                # Fila 2

                # Fila 3

                # Fila 4

                # =====================================================
                # SIGNATURES
                # =====================================================

                "chief_officer": "AN29",
                "master": "AN32",
                "msl_surveyor": "AN35",
            }
        }   # â† cierra "Draft"
        ,

        "Deductions": {
            "date_fields": [],
            "fields": {

            }
        }
    }

    # =========================================================
    # SAFE SETTERS (MERGE SAFE + NUMERIC SAFE + BOOL SAFE)
    # =========================================================
    def _coerce_excel_value(self, value):
        if value in [None, ""]:
            return value

        if isinstance(value, bool):
            return "YES" if value else "NO"

        if isinstance(value, (int, float)):
            return value

        if not isinstance(value, str):
            return value

        text = value.strip()
        if not text:
            return value
        if text.startswith("="):
            return text

        number_text = text.replace("\u00a0", "").replace(" ", "")
        allowed = set("0123456789+-.,")
        if any(ch not in allowed for ch in number_text):
            return value
        if any(sign in number_text[1:] for sign in "+-"):
            return value

        if "," in number_text and "." in number_text:
            if number_text.rfind(",") > number_text.rfind("."):
                number_text = number_text.replace(".", "").replace(",", ".")
            else:
                number_text = number_text.replace(",", "")
        elif "," in number_text:
            parts = number_text.split(",")
            if len(parts) == 2:
                number_text = f"{parts[0]}.{parts[1]}"
            else:
                number_text = number_text.replace(",", "")

        try:
            number = float(number_text)
        except Exception:
            return value

        if number.is_integer() and "." not in number_text:
            return int(number)
        return number

    def _coerce_number(self, value):
        coerced = self._coerce_excel_value(value)
        if isinstance(coerced, (int, float)):
            return float(coerced)
        return None

    def _is_empty(self, value) -> bool:
        return value is None or (isinstance(value, str) and value.strip() == "")

    def _prepare_hydrostatic_payload(self, payload: dict) -> dict:
        prepared = dict(payload or {})

        # The template needs the MTC draft base. If the user leaves it blank,
        # derive it from the first hydrostatic draft plus 0.50, as in the
        # vessel spreadsheet reference.
        for prefix in ("init", "final"):
            for table_no in (1, 2):
                draft_key = f"{prefix}_hydro{table_no}_draft_1"
                mtc_draft_key = f"{prefix}_hydro{table_no}_draft_mtc"
                if self._is_empty(prepared.get(mtc_draft_key)):
                    draft_value = self._coerce_number(prepared.get(draft_key))
                    if draft_value is not None:
                        prepared[mtc_draft_key] = round(draft_value + 0.5, 6)

        # Final Draft tables for this workflow expect LCF to carry the negative
        # sign when the hydrostatic table is supplied as an absolute value.
        for table_no in (1, 2):
            for row_no in (1, 2):
                lcf_key = f"final_hydro{table_no}_lcf_{row_no}"
                lcf_value = self._coerce_number(prepared.get(lcf_key))
                if lcf_value is not None and lcf_value > 0:
                    prepared[lcf_key] = -abs(lcf_value)

        return prepared

    def _safe_set(self, ws: Worksheet, cell: str, value):

        if value in [None, ""]:
            return

        value = self._coerce_excel_value(value)

        for merged in ws.merged_cells.ranges:
            if cell in merged:
                ws.cell(row=merged.min_row, column=merged.min_col).value = value
                return

        ws[cell].value = value

    def _safe_set_date(self, ws: Worksheet, cell: str, value):

        if not value:
            return

        parsed = None

        try:
            # -------------------------------------------------
            # 1) Si ya es datetime o date â†’ usar directo
            # -------------------------------------------------
            if isinstance(value, (datetime, date)):
                parsed = value

            # -------------------------------------------------
            # 2) Si es string â†’ intentar mÃºltiples formatos
            # -------------------------------------------------
            elif isinstance(value, str):

                v = " ".join(value.strip().replace(",", " ").split())

                # Intentar formatos comunes usados en el ERP
                date_formats = [
                    "%m-%d-%Y",  # 02-20-2026  â† DateEntry actual
                    "%d-%m-%Y",  # 20-02-2026
                    "%Y-%m-%d",  # 2026-02-20
                    "%m/%d/%Y",
                    "%d/%m/%Y",
                    "%Y/%m/%d",
                    "%B %d %Y",
                    "%b %d %Y",
                    "%Y-%m-%d %H:%M",
                    "%Y-%m-%d %H:%M:%S",
                    "%Y-%m-%dT%H:%M",
                    "%Y-%m-%dT%H:%M:%S",
                ]

                for fmt in date_formats:
                    try:
                        parsed = datetime.strptime(v, fmt)
                        break
                    except Exception:
                        continue

        except Exception:
            parsed = None

        # -------------------------------------------------
        # Si no logrÃ³ parsear, salir silenciosamente
        # -------------------------------------------------
        if not parsed:
            return

        # -------------------------------------------------
        # Manejo seguro de merged cells
        # -------------------------------------------------
        for merged in ws.merged_cells.ranges:
            if cell in merged:
                c = ws.cell(row=merged.min_row, column=merged.min_col)
                c.value = parsed
                c.number_format = "DD-MM-YYYY"
                return

        # -------------------------------------------------
        # Celda normal
        # -------------------------------------------------
        ws[cell].value = parsed
        ws[cell].number_format = "DD-MM-YYYY"


    # =========================================================
    # GENERATE EXCEL FROM TEMPLATE (PUBLIC METHOD) â€” BLINDADO
    # =========================================================
    def generate(self, payload: dict) -> str:

        if not os.path.exists(TEMPLATE_PATH):
            raise FileNotFoundError("Draft Survey template not found.")

        payload = self._prepare_hydrostatic_payload(payload or {})

        wb = load_workbook(TEMPLATE_PATH)

        # =====================================================
        # 1) FILL STATIC MAPPING
        # =====================================================
        for sheet_name, config in self.EXCEL_MAPPING.items():

            if sheet_name not in wb.sheetnames:
                continue

            ws = wb[sheet_name]

            fields = config.get("fields", {})
            date_fields = config.get("date_fields", [])

            for key, cell in fields.items():

                value = (payload or {}).get(key)

                if key in date_fields:
                    self._safe_set_date(ws, cell, value)
                else:
                    self._safe_set(ws, cell, value)

        # =====================================================
        # 2) DEDUCTIONS â€” DYNAMIC (ALINEADO CON FRONTEND)
        # payload["ballast"]["init"/"final"]
        # payload["fresh_water"]["init"/"final"]
        # =====================================================
        try:
            if "Deductions" in wb.sheetnames:
                ws_ded = wb["Deductions"]

                # -------------------------------------------------
                # HELPERS
                # -------------------------------------------------
                def _get_nested_list(section_key, phase_key):
                    section = (payload or {}).get(section_key)

                    if not isinstance(section, dict):
                        return []

                    items = section.get(phase_key)

                    if isinstance(items, list):
                        return items

                    return []

                def _value_present(value) -> bool:
                    return value is not None and str(value).strip() != ""

                def _with_defaults(items, section_key, phase_key, initial_items=None):
                    normalized = []
                    initial_items = initial_items or []

                    for i, item in enumerate((items or [])[:20]):
                        item = dict(item or {})
                        initial_item = (
                            initial_items[i]
                            if i < len(initial_items) and isinstance(initial_items[i], dict)
                            else {}
                        )

                        if phase_key == "final":
                            if not _value_present(item.get("tank_name")) and _value_present(initial_item.get("tank_name")):
                                item["tank_name"] = initial_item.get("tank_name")
                            if not _value_present(item.get("height")) and _value_present(initial_item.get("height")):
                                item["height"] = initial_item.get("height")

                        if section_key == "ballast":
                            if not _value_present(item.get("height")):
                                item["height"] = 0

                        if section_key == "fresh_water":
                            if _value_present(item.get("volume")) and not _value_present(item.get("density")):
                                item["density"] = 1

                        normalized.append(item)

                    return normalized

                def _clear_block(start_row, cols, max_rows=20):
                    for i in range(max_rows):
                        row = start_row + i
                        for col in cols:
                            try:
                                ws_ded[f"{col}{row}"].value = None
                            except Exception:
                                pass

                def _fill_block(items, start_row, col_name, col_height, col_sounding, col_volume, col_density):
                    for i, item in enumerate((items or [])[:20]):
                        row = start_row + i
                        item = item or {}

                        self._safe_set(ws_ded, f"{col_name}{row}", item.get("tank_name"))
                        self._safe_set(ws_ded, f"{col_height}{row}", item.get("height"))
                        self._safe_set(ws_ded, f"{col_sounding}{row}", item.get("sounding"))
                        self._safe_set(ws_ded, f"{col_volume}{row}", item.get("volume"))
                        self._safe_set(ws_ded, f"{col_density}{row}", item.get("density"))

                # -------------------------------------------------
                # BALLAST INITIAL
                # A11 / D11 / G11 / J11 / M11
                # -------------------------------------------------
                ballast_initial = _with_defaults(
                    _get_nested_list("ballast", "init"),
                    "ballast",
                    "init"
                )

                _clear_block(11, ["A", "D", "G", "J", "M"])

                _fill_block(
                    ballast_initial,
                    11,
                    "A", "D", "G", "J", "M"
                )

                # -------------------------------------------------
                # BALLAST FINAL
                # T11 / W11 / Z11 / AC11 / AF11
                # -------------------------------------------------
                ballast_final = _with_defaults(
                    _get_nested_list("ballast", "final"),
                    "ballast",
                    "final",
                    ballast_initial
                )

                _clear_block(11, ["T", "W", "Z", "AC", "AF"])

                _fill_block(
                    ballast_final,
                    11,
                    "T", "W", "Z", "AC", "AF"
                )

                # -------------------------------------------------
                # FRESH WATER INITIAL
                # A47 / D47 / G47 / J47 / M47
                # -------------------------------------------------
                fw_initial = _with_defaults(
                    _get_nested_list("fresh_water", "init"),
                    "fresh_water",
                    "init"
                )

                _clear_block(47, ["A", "D", "G", "J", "M"])

                _fill_block(
                    fw_initial,
                    47,
                    "A", "D", "G", "J", "M"
                )

                # -------------------------------------------------
                # FRESH WATER FINAL
                # T47 / W47 / Z47 / AC47 / AF47
                # -------------------------------------------------
                fw_final = _with_defaults(
                    _get_nested_list("fresh_water", "final"),
                    "fresh_water",
                    "final",
                    fw_initial
                )

                _clear_block(47, ["T", "W", "Z", "AC", "AF"])

                _fill_block(
                    fw_final,
                    47,
                    "T", "W", "Z", "AC", "AF"
                )

        except Exception as e:
            print("ERROR DEDUCTIONS:", e)

        # =====================================================
        # 3) SAVE TEMP FILE
        # =====================================================
        tmp_dir = tempfile.mkdtemp(prefix="draft_excel_")
        tmp_path = os.path.join(tmp_dir, "draft_survey.xlsx")

        wb.save(tmp_path)

        return tmp_path


    # =========================================================
    # VISUALIZAR DRAFT (ULTRA BLINDADO + AT1 AISLADA DEFINITIVA)
    # =========================================================
    def _visualizar_draft(self):

        # -----------------------------------------------------
        # ðŸ”’ Forzar actualizaciÃ³n de widgets
        # -----------------------------------------------------
        try:
            self.update_idletasks()
            self.update()
        except Exception:
            pass

        # -----------------------------------------------------
        # 1) Preguntar variante
        # -----------------------------------------------------
        try:
            variant = self._ask_draft_variant()
            if not variant:
                return
        except Exception:
            return

        # -----------------------------------------------------
        # 2) Obtener payload
        # -----------------------------------------------------
        try:
            payload = self.get_payload() or {}
        except Exception as e:
            messagebox.showerror(
                "Error",
                f"No se pudo obtener datos del formulario:\n{e}"
            )
            return

        # -----------------------------------------------------
        # 3) Sanitizar payload completo
        # -----------------------------------------------------
        try:
            sanitized_payload = {
                k: self._sanitize_for_excel(v)
                for k, v in payload.items()
            }
        except Exception:
            sanitized_payload = payload

        # -----------------------------------------------------
        # 4) ValidaciÃ³n mÃ­nima
        # -----------------------------------------------------
        if not sanitized_payload.get("vessel_mv"):
            messagebox.showwarning(
                "ValidaciÃ³n",
                "Debe seleccionar un servicio antes de visualizar."
            )
            return

        # -----------------------------------------------------
        # 5) Generar Excel desde template
        # -----------------------------------------------------
        try:
            from backend_api.services.draft_survey_excel_service import DraftSurveyExcelGenerator

            generator = DraftSurveyExcelGenerator()
            tmp_path = generator.generate(sanitized_payload)

            if not tmp_path or not os.path.exists(tmp_path):
                raise Exception("No se generÃ³ archivo temporal vÃ¡lido.")

        except Exception as e:
            messagebox.showerror(
                "Error",
                f"No se pudo generar el preview:\n{e}"
            )
            return

        # -----------------------------------------------------
        # 6) Abrir Excel y aplicar modo blindado
        # -----------------------------------------------------
        try:
            import win32com.client
            import pythoncom
            import time

            pythoncom.CoInitialize()

            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = True
            excel.DisplayAlerts = False
            excel.ScreenUpdating = True

            # Seguridad adicional
            try:
                excel.EnableEvents = False
            except Exception:
                pass

            try:
                excel.AskToUpdateLinks = False
            except Exception:
                pass

            workbook = excel.Workbooks.Open(
                tmp_path,
                UpdateLinks=0,
                ReadOnly=False
            )

            # -------------------------------------------------
            # Determinar tÃ­tulo
            # -------------------------------------------------
            variant_norm = (variant or "").strip().lower()
            if variant_norm == "intermediate":
                title_text = "INTERMEDIATE DRAFT SURVEY"
            else:
                title_text = "FINAL DRAFT SURVEY"

            # -------------------------------------------------
            # Validar hoja Draft
            # -------------------------------------------------
            try:
                ws = workbook.Worksheets("Draft")
            except Exception:
                raise Exception("La hoja 'Draft' no existe en el template.")

            # -------------------------------------------------
            # Recalcular con timeout seguro
            # -------------------------------------------------
            try:
                excel.Calculation = -4105  # xlCalculationAutomatic
                excel.CalculateFullRebuild()

                timeout = 10
                start = time.time()

                while excel.CalculationState != 0:
                    if time.time() - start > timeout:
                        break
                    time.sleep(0.1)

            except Exception:
                pass

            # -------------------------------------------------
            # Aislar AT1 completamente
            # -------------------------------------------------
            try:
                rng_at1 = ws.Range("AT1")

                try:
                    if rng_at1.HasFormula:
                        rng_at1.Formula = ""
                except Exception:
                    pass

                try:
                    rng_at1.ClearContents()
                except Exception:
                    pass

                rng_at1.NumberFormat = "@"
                rng_at1.Value = title_text

                try:
                    rng_at1.Locked = True
                except Exception:
                    pass

            except Exception as e:
                messagebox.showwarning(
                    "Aviso",
                    f"No se pudo forzar AT1:\n{e}"
                )

            # -------------------------------------------------
            # Opcional: mantener Z6 sincronizada
            # -------------------------------------------------
            try:
                rng_z6 = ws.Range("Z6")
                rng_z6.NumberFormat = "@"
                rng_z6.Value = title_text
            except Exception:
                pass

            # -------------------------------------------------
            # Guardar cambios
            # -------------------------------------------------
            try:
                workbook.Save()
                workbook.Saved = True
            except Exception:
                pass

            # -------------------------------------------------
            # Proteger estructura workbook
            # -------------------------------------------------
            try:
                workbook.Protect(
                    Password="msl_view_only",
                    Structure=True,
                    Windows=False
                )
            except Exception:
                pass

            # -------------------------------------------------
            # Proteger cada hoja
            # -------------------------------------------------
            for sheet in workbook.Worksheets:
                try:
                    sheet.Protect(
                        Password="msl_view_only",
                        DrawingObjects=True,
                        Contents=True,
                        Scenarios=True
                    )
                except Exception:
                    try:
                        sheet.Protect(Password="msl_view_only")
                    except Exception:
                        pass

                try:
                    sheet.EnableSelection = 0  # xlNoSelection
                except Exception:
                    pass

            # -------------------------------------------------
            # Bloqueo visual UI
            # -------------------------------------------------
            try:
                excel.DisplayFormulaBar = False
            except Exception:
                pass

            try:
                excel.ExecuteExcel4Macro('SHOW.TOOLBAR("Ribbon",False)')
            except Exception:
                pass

            try:
                excel.EnableEvents = True
            except Exception:
                pass

        except Exception as e:
            messagebox.showwarning(
                "Aviso",
                f"El Excel se abriÃ³, pero no se pudo aplicar el modo bloqueado completo:\n{e}"
            )



# =========================================================
# BACKWARD COMPATIBILITY
# =========================================================

def generate_draft_survey_excel(*args, **kwargs):
    """
    Wrapper para compatibilidad con imports antiguos.
    Soporta llamadas como:
        generate_draft_survey_excel(payload)
        generate_draft_survey_excel(payload, variant="FINAL")
        generate_draft_survey_excel(payload=..., variant=...)
    """

    payload = None

    # Caso 1: payload viene como primer argumento
    if args:
        payload = args[0]

    # Caso 2: payload viene como keyword
    if payload is None:
        payload = kwargs.get("payload")

    generator = DraftSurveyExcelGenerator()
    return generator.generate(payload or {})

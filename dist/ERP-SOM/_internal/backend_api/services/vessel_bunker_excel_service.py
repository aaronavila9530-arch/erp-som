import os
import tempfile
from datetime import datetime, date
from openpyxl import load_workbook
try:
    from services.template_autofit import apply_workbook_autofit
except ModuleNotFoundError:
    from backend_api.services.template_autofit import apply_workbook_autofit
from openpyxl.worksheet.worksheet import Worksheet


TEMPLATE_PATH = os.path.abspath(
    os.path.join(
        os.path.dirname(__file__),
        "..",
        "templates",
        "on_off_spot_template.xlsx"
    )
)


class VesselBunkerExcelGenerator:

    # =========================================================
    # MASTER ORDER (DEBE COINCIDIR EXACTAMENTE CON EL TEMPLATE)
    # ORDEN = FILAS EN LA HOJA "Data from DB"
    # =========================================================
    COLUMN_ORDER = [
        "id",
        "bunker_cert_no",
        "ship_name",
        "port_of_registry",
        "gross_tonnage",
        "report_date",
        "certificate",
        "report_category",
        "client",
        "port",
        "country",
        "berthing_date",
        "commenced_date",
        "dslop_date",
        "dslop_port",
        "dslop_country",
        "bunker_delivery_declared",
        "rob_diff",
        "plus_consumption",
        "generator_until_aps",
        "cons_dept",
        "me_to_sea_buoy",
        "remarks",
        "draft",
        "draft_fwd",
        "draft_aft",
        "trim",
        "list",
    ]

    # VLSFO 1–20
    for i in range(1, 21):
        COLUMN_ORDER.extend([
            f"vlsfo_tank_{i}_name",
            f"vlsfo_tank_{i}_dist_mtrs",
            f"vlsfo_tank_{i}_gauge_mtrs",
            f"vlsfo_tank_{i}_volume_m3",
            f"vlsfo_tank_{i}_temp_c",
            f"vlsfo_tank_{i}_temp_f",
            f"vlsfo_tank_{i}_density_15c",
            f"vlsfo_tank_{i}_weight_mt",
        ])

    # MGO 1–20
    for i in range(1, 21):
        COLUMN_ORDER.extend([
            f"mgo_tank_{i}_name",
            f"mgo_tank_{i}_dist_mtrs",
            f"mgo_tank_{i}_gauge_mtrs",
            f"mgo_tank_{i}_volume_m3",
            f"mgo_tank_{i}_temp_c",
            f"mgo_tank_{i}_temp_f",
            f"mgo_tank_{i}_density_15c",
            f"mgo_tank_{i}_weight_mt",
        ])

    # Bunker Figures 1–10
    for i in range(1, 11):
        COLUMN_ORDER.extend([
            f"bunker_figure_{i}_name",
            f"bunker_figure_{i}_ifo",
            f"bunker_figure_{i}_vlsfo",
            f"bunker_figure_{i}_lsmgo",
        ])

    # EXTRA FIELDS
    COLUMN_ORDER.extend([
        "antecedent_arrived_dt",
        "antecedent_survey_date_from",
        "antecedent_survey_date_to",
        "inspection_with",
        "workflow_status",
        "status",
        "created_at",
        "updated_at",
    ])

    # =========================================================
    # SAFE SETTER
    # =========================================================
    def _safe_set(self, ws: Worksheet, cell: str, value):

        if value in [None, ""]:
            return

        if isinstance(value, bool):
            value = "YES" if value else "NO"

        try:
            if isinstance(value, str):
                v = value.replace(",", ".")
                if v.replace(".", "", 1).isdigit():
                    value = float(v)
        except Exception:
            pass

        ws[cell].value = value

    # =========================================================
    # GENERATE (OFICIAL)
    # =========================================================
    def generate(self, payload: dict) -> str:

        if not os.path.exists(TEMPLATE_PATH):
            raise FileNotFoundError("on_off_spot_template.xlsx not found.")

        wb = load_workbook(TEMPLATE_PATH)

        if "Data from DB" not in wb.sheetnames:
            raise Exception("Sheet 'Data from DB' not found in template.")

        ws = wb["Data from DB"]

        start_row = 2

        for idx, field_name in enumerate(self.COLUMN_ORDER):

            row = start_row + idx
            cell = f"B{row}"

            value = (payload or {}).get(field_name)

            self._safe_set(ws, cell, value)

        tmp_dir = tempfile.mkdtemp(prefix="bunker_excel_")
        tmp_path = os.path.join(tmp_dir, "vessel_bunker_report.xlsx")

        apply_workbook_autofit(wb)
        wb.save(tmp_path)

        return tmp_path

    # =========================================================
    # GENERATE PREVIEW (NO DB — VISUALIZAR DESDE FORM)
    # =========================================================
    def generate_preview(self, payload: dict) -> str:

        if not isinstance(payload, dict):
            raise ValueError("Invalid payload for preview.")

        return self.generate(payload)


# =========================================================
# SERVICE WRAPPER (🔥 ESTO ERA LO QUE FALTABA)
# =========================================================
class VesselBunkerExcelService:
    """
    Wrapper service requerido por el router.
    No toca DB.
    Solo usa el generator.
    """

    def __init__(self):
        self.generator = VesselBunkerExcelGenerator()

    def generate_excel_from_payload(self, payload: dict) -> str:

        if not isinstance(payload, dict):
            raise ValueError("Invalid payload for Excel preview")

        file_path = self.generator.generate_preview(payload)

        if not file_path or not os.path.exists(file_path):
            raise Exception("Excel file was not generated")

        return file_path

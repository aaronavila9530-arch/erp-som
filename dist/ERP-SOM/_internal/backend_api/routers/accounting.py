from fastapi import (
    APIRouter,
    Depends,
    HTTPException,
    Header,
    Query,
)
from fastapi.responses import FileResponse
from psycopg2.extras import Json, RealDictCursor
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
import os
import tempfile

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from database import get_db
from rbac_service import has_permission
from services.finance_audit import (
    actor_from_headers,
    audit_event,
    ensure_finance_audit_schema,
    row_to_dict,
    rows_to_dicts,
)
from services.accounting_bank_rules import (
    backfill_missing_bank_accounts,
    canonical_bac_account,
    canonical_bcr_account,
)
from services.accounting_rule_engine import (
    ensure_posting_rule_schema,
    list_posting_rules,
    seed_default_posting_rules,
)
from services.accounting_fiscal_rules import ensure_accounting_fiscal_schema


router = APIRouter(
    prefix="/accounting",
    tags=["Accounting"]
)

MONEY_QUANT = Decimal("0.01")
ENTRY_STATUSES = {"DRAFT", "IN_REVIEW", "APPROVED", "POSTED", "REVERSED"}


def _company_code(value: str | None = None, header_value: str | None = None) -> str:
    code = str(value or header_value or "MSL-CR").strip().upper()
    return code or "MSL-CR"


def _money(value, field="amount"):
    try:
        result = Decimal(str(value or 0)).quantize(MONEY_QUANT, rounding=ROUND_HALF_UP)
    except (InvalidOperation, ValueError, TypeError):
        raise HTTPException(status_code=400, detail=f"Invalid {field}")
    if result < 0:
        raise HTTPException(status_code=400, detail=f"{field} cannot be negative")
    return result


def _ensure_accounting_professional_schema(conn):
    """Idempotent, additive migration for the professional accounting foundation."""
    with conn.cursor() as cur:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS accounting_accounts (
                id BIGSERIAL PRIMARY KEY,
                account_code VARCHAR(50) NOT NULL UNIQUE,
                account_name TEXT NOT NULL,
                account_type VARCHAR(30) NOT NULL,
                normal_balance VARCHAR(10) NOT NULL DEFAULT 'DEBIT',
                account_level INTEGER NOT NULL DEFAULT 1,
                parent_account VARCHAR(50),
                accepts_posting BOOLEAN NOT NULL DEFAULT TRUE,
                requires_third_party BOOLEAN NOT NULL DEFAULT FALSE,
                requires_cost_center BOOLEAN NOT NULL DEFAULT FALSE,
                currency_code VARCHAR(3),
                financial_statement_line TEXT,
                tax_mapping TEXT,
                active BOOLEAN NOT NULL DEFAULT TRUE,
                locked BOOLEAN NOT NULL DEFAULT FALSE,
                locked_by TEXT,
                locked_at TIMESTAMP,
                created_by TEXT,
                created_at TIMESTAMP NOT NULL DEFAULT NOW(),
                updated_by TEXT,
                updated_at TIMESTAMP NOT NULL DEFAULT NOW()
            )
        """)
        cur.execute("""
            INSERT INTO accounting_accounts (
                account_code, account_name, account_type, account_level,
                parent_account, normal_balance, active
            )
            SELECT DISTINCT ON (account_code)
                account_code, account_name,
                COALESCE(NULLIF(account_type, ''), 'UNCLASSIFIED'),
                COALESCE(account_level, 1), parent_account,
                CASE WHEN COALESCE(account_type, '') IN ('PASIVO','PATRIMONIO','INGRESO','REVENUE','LIABILITY','EQUITY')
                     THEN 'CREDIT' ELSE 'DEBIT' END,
                COALESCE(active, TRUE)
            FROM accounting_ledger
            WHERE account_code IS NOT NULL AND BTRIM(account_code) <> ''
              AND account_code <> '1.1.02.01'
              AND LOWER(COALESCE(account_name, '')) NOT LIKE '%banco nacional%'
            ON CONFLICT (account_code) DO NOTHING
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS accounting_period_controls (
                id BIGSERIAL PRIMARY KEY,
                company_code VARCHAR(30) NOT NULL DEFAULT 'MSL-CR',
                period VARCHAR(7) NOT NULL,
                status VARCHAR(20) NOT NULL DEFAULT 'OPEN',
                closed_by TEXT,
                closed_at TIMESTAMP,
                reopened_by TEXT,
                reopened_at TIMESTAMP,
                reopen_reason TEXT,
                updated_at TIMESTAMP NOT NULL DEFAULT NOW(),
                UNIQUE(company_code, period),
                CHECK (status IN ('OPEN','SOFT_CLOSED','CLOSED'))
            )
        """)
        for statement in (
            "ALTER TABLE accounting_entries ADD COLUMN IF NOT EXISTS workflow_status VARCHAR(20) NOT NULL DEFAULT 'POSTED'",
            "ALTER TABLE accounting_entries ADD COLUMN IF NOT EXISTS version INTEGER NOT NULL DEFAULT 1",
            "ALTER TABLE accounting_entries ADD COLUMN IF NOT EXISTS company_code VARCHAR(30) NOT NULL DEFAULT 'MSL-CR'",
            "ALTER TABLE accounting_entries ADD COLUMN IF NOT EXISTS currency_code VARCHAR(3) NOT NULL DEFAULT 'CRC'",
            "ALTER TABLE accounting_entries ADD COLUMN IF NOT EXISTS exchange_rate NUMERIC(18,6) NOT NULL DEFAULT 1",
            "ALTER TABLE accounting_entries ADD COLUMN IF NOT EXISTS submitted_by TEXT",
            "ALTER TABLE accounting_entries ADD COLUMN IF NOT EXISTS submitted_at TIMESTAMP",
            "ALTER TABLE accounting_entries ADD COLUMN IF NOT EXISTS approved_by TEXT",
            "ALTER TABLE accounting_entries ADD COLUMN IF NOT EXISTS approved_at TIMESTAMP",
            "ALTER TABLE accounting_entries ADD COLUMN IF NOT EXISTS posted_by TEXT",
            "ALTER TABLE accounting_entries ADD COLUMN IF NOT EXISTS posted_at TIMESTAMP",
            "ALTER TABLE accounting_entries ADD COLUMN IF NOT EXISTS updated_at TIMESTAMP NOT NULL DEFAULT NOW()",
            "ALTER TABLE accounting_entries ADD COLUMN IF NOT EXISTS posting_rule_code VARCHAR(100)",
            "ALTER TABLE accounting_entries ADD COLUMN IF NOT EXISTS posting_metadata JSONB NOT NULL DEFAULT '{}'::jsonb",
            "ALTER TABLE accounting_accounts ADD COLUMN IF NOT EXISTS locked BOOLEAN NOT NULL DEFAULT FALSE",
            "ALTER TABLE accounting_accounts ADD COLUMN IF NOT EXISTS locked_by TEXT",
            "ALTER TABLE accounting_accounts ADD COLUMN IF NOT EXISTS locked_at TIMESTAMP",
            "ALTER TABLE accounting_lines ADD COLUMN IF NOT EXISTS third_party_type VARCHAR(30)",
            "ALTER TABLE accounting_lines ADD COLUMN IF NOT EXISTS third_party_id TEXT",
            "ALTER TABLE accounting_lines ADD COLUMN IF NOT EXISTS cost_center_code VARCHAR(50)",
            "ALTER TABLE accounting_lines ADD COLUMN IF NOT EXISTS project_code VARCHAR(50)",
            "ALTER TABLE accounting_lines ADD COLUMN IF NOT EXISTS vessel_code VARCHAR(100)",
            "ALTER TABLE accounting_lines ADD COLUMN IF NOT EXISTS tax_role VARCHAR(40)",
        ):
            cur.execute(statement)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS accounting_audit_log (
                id BIGSERIAL PRIMARY KEY,
                entry_id INTEGER REFERENCES accounting_entries(id) ON DELETE SET NULL,
                action VARCHAR(40) NOT NULL,
                previous_status VARCHAR(20),
                new_status VARCHAR(20),
                performed_by TEXT,
                reason TEXT,
                snapshot JSONB,
                created_at TIMESTAMP NOT NULL DEFAULT NOW()
            )
        """)
        cur.execute("CREATE INDEX IF NOT EXISTS idx_accounting_audit_entry ON accounting_audit_log(entry_id, created_at DESC)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_accounting_entries_workflow ON accounting_entries(period, workflow_status)")
        ensure_finance_audit_schema(cur)
        ensure_posting_rule_schema(cur)
        ensure_accounting_fiscal_schema(cur)
    conn.commit()


def _assert_period_open(cur, period, company_code="MSL-CR"):
    cur.execute("""
        SELECT status FROM accounting_period_controls
        WHERE company_code = %s AND period = %s
    """, (company_code, period))
    row = cur.fetchone()
    if row and row.get("status") == "CLOSED":
        raise HTTPException(status_code=409, detail=f"Accounting period {period} is closed")
    fiscal_year, fiscal_month = (int(value) for value in period.split("-"))
    cur.execute("""
        SELECT period_closed FROM closing_status
        WHERE company_code = %s AND fiscal_year = %s AND period = %s
          AND ledger = '0L'
    """, (company_code, fiscal_year, fiscal_month))
    legacy_close = cur.fetchone()
    if legacy_close and legacy_close.get("period_closed"):
        raise HTTPException(status_code=409, detail=f"Accounting period {period} is closed")


def _audit(cur, entry_id, action, user, previous_status=None, new_status=None, reason=None):
    cur.execute("SELECT * FROM accounting_entries WHERE id = %s", (entry_id,))
    entry = cur.fetchone()
    cur.execute("SELECT * FROM accounting_lines WHERE entry_id = %s ORDER BY id", (entry_id,))
    lines = cur.fetchall()
    snapshot = {"entry": dict(entry or {}), "lines": [dict(line) for line in lines]}
    cur.execute("""
        INSERT INTO accounting_audit_log
            (entry_id, action, previous_status, new_status, performed_by, reason, snapshot)
        VALUES (%s, %s, %s, %s, %s, %s, %s)
    """, (entry_id, action, previous_status, new_status, user, reason, Json(snapshot, dumps=lambda v: __import__('json').dumps(v, default=str))))


def _entry_validation(cur, entry_id):
    cur.execute("""
        SELECT
            COALESCE(SUM(debit), 0) AS total_debit,
            COALESCE(SUM(credit), 0) AS total_credit,
            COUNT(*) AS line_count,
            COUNT(*) FILTER (
                WHERE COALESCE(debit, 0) > 0 AND COALESCE(credit, 0) > 0
            ) AS both_sides_count,
            COUNT(*) FILTER (
                WHERE COALESCE(debit, 0) = 0 AND COALESCE(credit, 0) = 0
            ) AS zero_lines_count
        FROM accounting_lines
        WHERE entry_id = %s
    """, (entry_id,))
    row = cur.fetchone() or {}
    debit = Decimal(str(row.get("total_debit") or 0)).quantize(MONEY_QUANT, rounding=ROUND_HALF_UP)
    credit = Decimal(str(row.get("total_credit") or 0)).quantize(MONEY_QUANT, rounding=ROUND_HALF_UP)
    return {
        "total_debit": debit,
        "total_credit": credit,
        "difference": (debit - credit).quantize(MONEY_QUANT, rounding=ROUND_HALF_UP),
        "line_count": int(row.get("line_count") or 0),
        "both_sides_count": int(row.get("both_sides_count") or 0),
        "zero_lines_count": int(row.get("zero_lines_count") or 0),
    }


def _assert_entry_can_advance(cur, entry_id):
    validation = _entry_validation(cur, entry_id)
    if validation["line_count"] < 2:
        raise HTTPException(409, "El asiento debe tener al menos dos lineas contables")
    if validation["both_sides_count"]:
        raise HTTPException(409, "El asiento tiene lineas con Debe y Haber simultaneamente")
    if validation["zero_lines_count"]:
        raise HTTPException(409, "El asiento tiene lineas sin monto")
    if validation["difference"] != Decimal("0.00"):
        raise HTTPException(
            409,
            f"Asiento descuadrado. Debe={validation['total_debit']}, "
            f"Haber={validation['total_credit']}, Diferencia={validation['difference']}"
        )
    return validation

# ============================================================
# RBAC GUARD
# ============================================================
def require_permission(module: str, action: str):
    def checker(
        x_user_role: str = Header(..., alias="X-User-Role")
    ):
        if not has_permission(x_user_role, module, action):
            raise HTTPException(
                status_code=403,
                detail="No autorizado"
            )
    return checker


def _accounting_entry_stats(conn):
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute("""
        SELECT
            COUNT(*) AS total_entries,
            MAX(period) AS latest_period
        FROM accounting_entries
    """)
    stats = cur.fetchone() or {}

    cur.execute("""
        SELECT period, COUNT(*) AS count
        FROM accounting_entries
        GROUP BY period
        ORDER BY period DESC
        LIMIT 12
    """)
    periods = cur.fetchall()

    return {
        "total_entries": int(stats.get("total_entries") or 0),
        "latest_period": stats.get("latest_period"),
        "period_counts": [
            {"period": row["period"], "count": int(row["count"] or 0)}
            for row in periods
        ]
    }


def _report_title(report: str | None):
    titles = {
        "ASIENTOS": "Asientos contables",
        "MAYOR": "Mayor general",
        "BC": "Balance de comprobacion",
        "ESF": "Estado de situacion financiera",
        "ER": "Estado de resultados",
        "FC": "Flujo de caja"
    }
    key = (report or "ASIENTOS").upper()
    return titles.get(key, key)


def _append_account_filter(conditions, params, account_code: str | None):
    if not account_code or account_code == "TODOS":
        return
    code = str(account_code).strip()
    if not code:
        return
    equivalent_accounts = {
        "1.1.04.01": ["1101"],
        "1101": ["1.1.04.01"],
        "4.1.01": ["4101"],
        "4101": ["4.1.01"],
        "2.1.02.03": ["2108"],
        "2108": ["2.1.02.03"],
        "2105": ["2.1.05"],
        "2.1.05": ["2105"],
    }
    codes = [code, *equivalent_accounts.get(code, [])]
    clauses = []
    values = []
    for item in codes:
        clauses.extend(["l.account_code = %s", "l.account_code LIKE %s"])
        values.extend([item, f"{item}.%"])
    if code == "1.1.02.02":
        clauses.extend([
            "LOWER(l.account_name) LIKE %s",
            "l.account_code = %s",
            "l.account_code = %s",
        ])
        values.extend(["%bac%", "110-002-002-001", "1.1.02"])
    elif code == "1.1.02.04":
        clauses.extend([
            "LOWER(l.account_name) LIKE %s",
            "l.account_code = %s",
        ])
        values.extend(["%banco de costa rica%", "1.1.02"])
    elif code == "1.1.02":
        clauses.append("LOWER(l.account_name) LIKE %s")
        values.append("%banco%")
    conditions.append("(" + " OR ".join(clauses) + ")")
    params.extend(values)


def _fetch_accounting_report_lines(
    conn,
    period: str | None = None,
    period_from: str | None = None,
    period_to: str | None = None,
    origin: str | None = None,
    account_code: str | None = None
):
    _ensure_accounting_professional_schema(conn)
    cur = conn.cursor(cursor_factory=RealDictCursor)
    conditions = ["e.workflow_status = 'POSTED'", "e.entry_date <= CURRENT_DATE"]
    params = []

    if period:
        conditions.append("e.period = %s")
        params.append(period)

    if period_from:
        conditions.append("e.period >= %s")
        params.append(period_from)

    if period_to:
        conditions.append("e.period <= %s")
        params.append(period_to)

    if origin and origin != "TODOS":
        conditions.append("e.origin = %s")
        params.append(origin)

    _append_account_filter(conditions, params, account_code)

    where_clause = f"WHERE {' AND '.join(conditions)}" if conditions else ""

    cur.execute(f"""
        SELECT
            e.entry_date,
            e.id AS entry_id,
            e.period,
            e.origin,
            e.origin_id,
            e.description AS entry_description,
            l.account_code,
            l.account_name,
            l.line_description,
            l.debit,
            l.credit
        FROM accounting_entries e
        JOIN accounting_lines l ON l.entry_id = e.id
        {where_clause}
        ORDER BY e.period DESC, e.entry_date DESC, e.id DESC, l.id ASC
    """, params)

    return cur.fetchall()


def _report_filename(extension: str, report: str | None, period: str | None, period_from: str | None, period_to: str | None):
    scope = period or (f"{period_from or 'inicio'}_{period_to or 'fin'}" if period_from or period_to else "todos")
    safe_report = (report or "ASIENTOS").lower().replace(" ", "_")
    return f"accounting_{safe_report}_{scope}.{extension}"


@router.get("/periods")
def get_accounting_periods(
    company_code: str | None = None,
    x_company_code: str | None = Header(None, alias="X-Company-Code"),
    conn=Depends(get_db),
):
    """
    Devuelve solamente los periodos que realmente tienen movimientos contables.
    Evita mostrar meses futuros o meses sin data en Accounting.
    """
    company = _company_code(company_code, x_company_code)
    _ensure_accounting_professional_schema(conn)
    current_period = date.today().strftime("%Y-%m")
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute("""
        SELECT period, COUNT(*) AS count
        FROM accounting_entries
        WHERE period IS NOT NULL
          AND company_code = %s
          AND period ~ '^[0-9]{4}-[0-9]{2}$'
          AND period >= '2025-01'
          AND period <= %s
          AND entry_date <= CURRENT_DATE
        GROUP BY period
        HAVING COUNT(*) > 0
        ORDER BY period ASC
    """, (company, current_period))
    rows = cur.fetchall()
    return {
        "data": [row["period"] for row in rows],
        "period_counts": [
            {"period": row["period"], "count": int(row["count"] or 0)}
            for row in rows
        ]
    }


@router.post("/manual-entry")
def create_manual_entry(
    payload: dict,
    conn=Depends(get_db),
    x_user: str | None = Header(None, alias="X-User"),
    x_role: str | None = Header(None, alias="X-Role"),
    x_user_role: str | None = Header(None, alias="X-User-Role"),
    x_company_code: str | None = Header(None, alias="X-Company-Code"),
):
    """
    payload:
    {
        entry_date,
        description,
        lines: [
            {account_code, account_name, debit, credit, line_description}
        ]
    }
    """

    _ensure_accounting_professional_schema(conn)
    lines = payload.get("lines", [])
    if not lines:
        raise HTTPException(400, "No accounting lines provided")

    total_debit = sum((_money(l.get("debit"), "debit") for l in lines), Decimal("0"))
    total_credit = sum((_money(l.get("credit"), "credit") for l in lines), Decimal("0"))

    if total_debit != total_credit or total_debit == 0:
        raise HTTPException(400, "Entry does not balance")

    entry_date = date.fromisoformat(payload["entry_date"])
    period = entry_date.strftime("%Y-%m")

    header_user, header_role = actor_from_headers(x_user, x_role, x_user_role)
    created_by = str(payload.get("created_by") or header_user or "unknown").strip()
    company_code = _company_code(payload.get("company_code"), x_company_code)
    currency_code = str(payload.get("currency_code") or "CRC").strip().upper()
    exchange_rate = Decimal(str(payload.get("exchange_rate") or 1))
    if exchange_rate <= 0:
        raise HTTPException(400, "exchange_rate must be greater than zero")

    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        _assert_period_open(cur, period, company_code)
        validated = []
        for line in lines:
            debit = _money(line.get("debit"), "debit")
            credit = _money(line.get("credit"), "credit")
            if (debit > 0) == (credit > 0):
                raise HTTPException(400, "Each line must contain debit or credit, but not both")
            cur.execute("""
                SELECT * FROM accounting_accounts
                WHERE account_code = %s AND active = TRUE AND accepts_posting = TRUE
            """, (line.get("account_code"),))
            account = cur.fetchone()
            if not account:
                raise HTTPException(400, f"Invalid or non-postable account: {line.get('account_code')}")
            if account.get("requires_third_party") and not line.get("third_party_id"):
                raise HTTPException(400, f"Account {account['account_code']} requires a third party")
            if account.get("requires_cost_center") and not line.get("cost_center_code"):
                raise HTTPException(400, f"Account {account['account_code']} requires a cost center")
            validated.append((line, account, debit, credit))

        cur.execute("""
            INSERT INTO accounting_entries (
                entry_date, period, description, origin, created_by,
                workflow_status, company_code, currency_code, exchange_rate,
                posting_rule_code, posting_metadata
            ) VALUES (%s, %s, %s, 'MANUAL', %s, 'DRAFT', %s, %s, %s, %s, %s)
            RETURNING id, workflow_status, version
        """, (
            entry_date,
            period,
            payload.get("description"),
            created_by,
            company_code,
            currency_code,
            exchange_rate,
            payload.get("posting_rule_code") or "MANUAL_ADJUSTMENT",
            Json({
                "origin": "Manual",
                "event_type": "Manual adjustment",
                "source": "desktop_manual_entry",
            }),
        ))
        entry = cur.fetchone()
        for line, account, debit, credit in validated:
            cur.execute("""
                INSERT INTO accounting_lines (
                    entry_id, account_code, account_name, debit, credit, line_description,
                    third_party_type, third_party_id, cost_center_code, project_code, vessel_code
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (entry["id"], account["account_code"], account["account_name"], debit, credit,
                  line.get("line_description"), line.get("third_party_type"), line.get("third_party_id"),
                  line.get("cost_center_code"), line.get("project_code"), line.get("vessel_code")))
        _audit(cur, entry["id"], "CREATE_DRAFT", created_by, None, "DRAFT")
        cur.execute("SELECT * FROM accounting_entries WHERE id = %s", (entry["id"],))
        after_entry = row_to_dict(cur.fetchone())
        cur.execute("SELECT * FROM accounting_lines WHERE entry_id = %s ORDER BY id", (entry["id"],))
        after_lines = rows_to_dicts(cur.fetchall())
        audit_event(
            cur,
            module="accounting",
            action="MANUAL_ENTRY_CREATED",
            entity_type="accounting_entry",
            entity_id=entry["id"],
            performed_by=created_by,
            performed_role=header_role,
            after={"entry": after_entry, "lines": after_lines},
            metadata={"period": period, "total_debit": str(total_debit), "total_credit": str(total_credit)},
        )
        conn.commit()
        return {"status": "ok", "entry_id": entry["id"], "workflow_status": "DRAFT", "version": entry["version"]}
    except HTTPException:
        conn.rollback()
        raise
    except Exception:
        conn.rollback()
        raise



@router.post("/reverse/{entry_id}")
def reverse_entry(
    entry_id: int,
    conn=Depends(get_db),
    x_user: str | None = Header(None, alias="X-User"),
    x_role: str | None = Header(None, alias="X-Role"),
    x_user_role: str | None = Header(None, alias="X-User-Role"),
):
    _ensure_accounting_professional_schema(conn)
    cur = conn.cursor(cursor_factory=RealDictCursor)
    performed_by, performed_role = actor_from_headers(x_user, x_role, x_user_role)

    # 1️⃣ Validar asiento original
    cur.execute("""
        SELECT *
        FROM accounting_entries
        WHERE id = %s
          AND COALESCE(reversed, FALSE) = FALSE
          AND workflow_status = 'POSTED'
    """, (entry_id,))
    entry = cur.fetchone()

    if not entry:
        raise HTTPException(
            status_code=400,
            detail="El asiento no existe o ya fue revertido"
        )

    # 2️⃣ Traer líneas originales
    _assert_period_open(cur, entry["period"], entry.get("company_code") or "MSL-CR")

    cur.execute("""
        SELECT *
        FROM accounting_lines
        WHERE entry_id = %s
    """, (entry_id,))
    lines = cur.fetchall()

    if not lines:
        raise HTTPException(400, "El asiento no tiene líneas")

    # 3️⃣ Crear asiento de reverso (NO marcado como reversed)
    before_snapshot = {"entry": row_to_dict(entry), "lines": rows_to_dicts(lines)}

    cur.execute("""
        INSERT INTO accounting_entries
        (entry_date, period, description, origin, origin_id, reversed)
        VALUES (CURRENT_DATE, %s, %s, 'REVERSAL', %s, FALSE)
        RETURNING id
    """, (
        entry["period"],
        f"Asiento de reversa del asiento {entry_id}",
        entry_id
    ))
    reversal_id = cur.fetchone()["id"]

    # 4️⃣ Insertar líneas INVERTIDAS
    for l in lines:
        cur.execute("""
            INSERT INTO accounting_lines
            (entry_id, account_code, account_name, debit, credit, line_description)
            VALUES (%s, %s, %s, %s, %s, %s)
        """, (
            reversal_id,
            l["account_code"],
            l["account_name"],
            l["credit"],   # 👈 INVERTIDO
            l["debit"],    # 👈 INVERTIDO
            f"Reverso de línea {l['id']}"
        ))

    # 5️⃣ Marcar SOLO el original como revertido
    cur.execute("""
        UPDATE accounting_entries
        SET reversed = TRUE,
            reversal_entry_id = %s,
            workflow_status = 'REVERSED',
            version = version + 1,
            updated_at = NOW()
        WHERE id = %s
    """, (reversal_id, entry_id))

    _audit(cur, entry_id, "REVERSE", performed_by, "POSTED", "REVERSED")
    _audit(cur, reversal_id, "CREATE_REVERSAL", performed_by, None, "POSTED")
    cur.execute("SELECT * FROM accounting_entries WHERE id IN (%s, %s) ORDER BY id", (entry_id, reversal_id))
    after_entries = rows_to_dicts(cur.fetchall())
    cur.execute("SELECT * FROM accounting_lines WHERE entry_id IN (%s, %s) ORDER BY entry_id, id", (entry_id, reversal_id))
    after_lines = rows_to_dicts(cur.fetchall())
    audit_event(
        cur,
        module="accounting",
        action="ENTRY_REVERSED",
        entity_type="accounting_entry",
        entity_id=entry_id,
        performed_by=performed_by,
        performed_role=performed_role,
        before=before_snapshot,
        after={"entries": after_entries, "lines": after_lines},
        metadata={"reversal_entry_id": reversal_id},
    )

    conn.commit()

    return {
        "status": "ok",
        "original_entry_id": entry_id,
        "reversal_entry_id": reversal_id
    }




# ============================================================
# CHART OF ACCOUNTS (Catalogo Contable)
# ============================================================
@router.get("/accounts")
def get_accounting_accounts(include_inactive: bool = False, conn=Depends(get_db)):
    """
    Devuelve el catálogo contable desde accounting_ledger
    para uso en combobox (UI / Popup de ajustes)
    """

    _ensure_accounting_professional_schema(conn)
    cur = conn.cursor(cursor_factory=RealDictCursor)

    query = """
        SELECT
            account_code,
            account_name,
            account_type,
            account_level,
            parent_account,
            normal_balance,
            accepts_posting,
            requires_third_party,
            requires_cost_center,
            currency_code,
            financial_statement_line,
            tax_mapping,
            active,
            locked,
            locked_by,
            locked_at
        FROM accounting_accounts
        WHERE (%s = TRUE OR active = TRUE)
        ORDER BY account_code
    """

    cur.execute(query, (include_inactive,))
    rows = cur.fetchall()

    return {
        "data": rows
    }


@router.get("/bank-accounts")
def get_accounting_bank_accounts(conn=Depends(get_db)):
    """
    Devuelve cuentas bancarias posteables para pagos de Collections e ITP.
    Usa el plan contable como fuente unica de verdad.
    """

    _ensure_accounting_professional_schema(conn)
    cur = conn.cursor(cursor_factory=RealDictCursor)
    canonical_codes = set()
    canonical_rows = []
    for row in (canonical_bac_account(cur), canonical_bcr_account(cur)):
        if row and row["account_code"] not in canonical_codes:
            canonical_codes.add(row["account_code"])
            canonical_rows.append({
                "account_code": row["account_code"],
                "account_name": row["account_name"],
                "account_type": "ASSET",
                "parent_account": "1.1.02",
                "active": True,
            })

    cur.execute("""
        SELECT
            account_code,
            account_name,
            account_type,
            parent_account,
            active
        FROM accounting_accounts
        WHERE active = TRUE
          AND account_code NOT IN ('1.1.01', '1.1.02', '1.1.02.01', '110-002-002-001')
          AND LOWER(account_name) NOT LIKE '%banco nacional%'
          AND (
                account_code LIKE '1.1.02.%'
             OR account_code LIKE '1.1.01.%'
             OR LOWER(account_name) LIKE 'banco%%'
          )
        ORDER BY account_code
    """)
    rows = []
    seen = set(canonical_codes)
    for row in cur.fetchall() or []:
        name = (row.get("account_name") or "").lower()
        if "banco de costa rica" in name and row.get("account_code") != "1.1.02.04":
            continue
        if "bac" in name and row.get("account_code") != "1.1.02.02":
            continue
        if row["account_code"] in seen:
            continue
        seen.add(row["account_code"])
        rows.append(row)
    return {"data": sorted(canonical_rows + rows, key=lambda item: item["account_code"])}


@router.post("/accounts")
def create_accounting_account(
    payload: dict,
    conn=Depends(get_db),
    x_user: str | None = Header(None, alias="X-User"),
    x_role: str | None = Header(None, alias="X-Role"),
    x_user_role: str | None = Header(None, alias="X-User-Role"),
):
    _ensure_accounting_professional_schema(conn)
    code = str(payload.get("account_code") or "").strip()
    name = str(payload.get("account_name") or "").strip()
    account_type = str(payload.get("account_type") or "").strip().upper()
    if not code or not name or not account_type:
        raise HTTPException(400, "account_code, account_name and account_type are required")
    header_user, header_role = actor_from_headers(x_user, x_role, x_user_role)
    user = str(payload.get("updated_by") or payload.get("created_by") or header_user or "unknown")
    try:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                INSERT INTO accounting_accounts (
                    account_code, account_name, account_type, normal_balance, account_level,
                    parent_account, accepts_posting, requires_third_party, requires_cost_center,
                    currency_code, financial_statement_line, tax_mapping, active, locked, created_by, updated_by
                ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                RETURNING *
            """, (code, name, account_type, str(payload.get("normal_balance") or "DEBIT").upper(),
                  int(payload.get("account_level") or 1), payload.get("parent_account"),
                  bool(payload.get("accepts_posting", True)), bool(payload.get("requires_third_party", False)),
                  bool(payload.get("requires_cost_center", False)), payload.get("currency_code"),
                  payload.get("financial_statement_line"), payload.get("tax_mapping"),
                  bool(payload.get("active", True)), bool(payload.get("locked", False)), user, user))
            row = cur.fetchone()
            audit_event(
                cur,
                module="accounting",
                action="ACCOUNT_CREATED",
                entity_type="accounting_account",
                entity_id=code,
                performed_by=user,
                performed_role=header_role,
                after=row_to_dict(row),
            )
        conn.commit()
        return {"status": "ok", "account": row}
    except Exception as exc:
        conn.rollback()
        if getattr(exc, "pgcode", None) == "23505":
            raise HTTPException(409, f"Account {code} already exists")
        raise


@router.put("/accounts/{account_code}")
def update_accounting_account(
    account_code: str,
    payload: dict,
    conn=Depends(get_db),
    x_user: str | None = Header(None, alias="X-User"),
    x_role: str | None = Header(None, alias="X-Role"),
    x_user_role: str | None = Header(None, alias="X-User-Role"),
):
    _ensure_accounting_professional_schema(conn)
    header_user, header_role = actor_from_headers(x_user, x_role, x_user_role)
    user = str(payload.get("updated_by") or header_user or "unknown")
    with conn.cursor(cursor_factory=RealDictCursor) as cur:
        cur.execute("SELECT * FROM accounting_accounts WHERE account_code = %s", (account_code,))
        before = row_to_dict(cur.fetchone())
        if not before:
            conn.rollback()
            raise HTTPException(404, "Account not found")
        protected_fields = {
            "account_name",
            "account_type",
            "normal_balance",
            "account_level",
            "parent_account",
            "accepts_posting",
            "requires_third_party",
            "requires_cost_center",
            "currency_code",
            "financial_statement_line",
            "tax_mapping",
        }
        changed_locked_fields = [
            field for field in protected_fields
            if field in payload and payload.get(field) != before.get(field)
        ]
        if before.get("locked") and changed_locked_fields and not payload.get("allow_locked_update"):
            raise HTTPException(
                409,
                "Account is locked. Only active/inactive can be changed without controlled override.",
            )
        cur.execute("""
            UPDATE accounting_accounts SET
                account_name = COALESCE(%s, account_name),
                account_type = COALESCE(%s, account_type),
                normal_balance = COALESCE(%s, normal_balance),
                account_level = COALESCE(%s, account_level),
                parent_account = %s,
                accepts_posting = COALESCE(%s, accepts_posting),
                requires_third_party = COALESCE(%s, requires_third_party),
                requires_cost_center = COALESCE(%s, requires_cost_center),
                currency_code = %s,
                financial_statement_line = %s,
                tax_mapping = %s,
                active = COALESCE(%s, active),
                locked = COALESCE(%s, locked),
                locked_by = CASE WHEN %s = TRUE AND locked = FALSE THEN %s ELSE locked_by END,
                locked_at = CASE WHEN %s = TRUE AND locked = FALSE THEN NOW() ELSE locked_at END,
                updated_by = %s, updated_at = NOW()
            WHERE account_code = %s RETURNING *
        """, (payload.get("account_name"), payload.get("account_type"), payload.get("normal_balance"),
              payload.get("account_level"), payload.get("parent_account"), payload.get("accepts_posting"),
              payload.get("requires_third_party"), payload.get("requires_cost_center"), payload.get("currency_code"),
              payload.get("financial_statement_line"), payload.get("tax_mapping"), payload.get("active"),
              payload.get("locked"), payload.get("locked"), user, payload.get("locked"), user, account_code))
        row = cur.fetchone()
        audit_event(
            cur,
            module="accounting",
            action="ACCOUNT_UPDATED",
            entity_type="accounting_account",
            entity_id=account_code,
            performed_by=user,
            performed_role=header_role,
            before=before,
            after=row_to_dict(row),
            metadata={"changed_fields": sorted(payload.keys())},
        )
    conn.commit()
    return {"status": "ok", "account": row}


@router.post("/accounts/harden")
def harden_accounting_chart(
    payload: dict | None = None,
    conn=Depends(get_db),
    x_user: str | None = Header(None, alias="X-User"),
    x_role: str | None = Header(None, alias="X-Role"),
    x_user_role: str | None = Header(None, alias="X-User-Role"),
):
    _ensure_accounting_professional_schema(conn)
    payload = payload or {}
    header_user, header_role = actor_from_headers(x_user, x_role, x_user_role)
    user = str(payload.get("user") or header_user or "system")
    with conn.cursor(cursor_factory=RealDictCursor) as cur:
        cur.execute("SELECT * FROM accounting_accounts ORDER BY account_code")
        before_accounts = rows_to_dicts(cur.fetchall())

        cur.execute("""
            SELECT a.account_code, COUNT(l.id) AS lines,
                   COUNT(c.id) AS collection_refs,
                   COUNT(p.id) AS itp_refs
            FROM accounting_accounts a
            LEFT JOIN accounting_lines l ON l.account_code = a.account_code
            LEFT JOIN cash_app c ON c.bank_account_code = a.account_code
            LEFT JOIN payment_obligations p ON p.payment_bank_account_code = a.account_code
            WHERE LOWER(a.account_name) LIKE '%banco nacional%'
               OR a.account_code = '1.1.02.01'
            GROUP BY a.account_code
        """)
        banco_nacional = cur.fetchall() or []
        deleted_bn = []
        inactive_bn = []
        for row in banco_nacional:
            if int(row.get("lines") or 0) == 0 and int(row.get("collection_refs") or 0) == 0 and int(row.get("itp_refs") or 0) == 0:
                cur.execute("DELETE FROM accounting_accounts WHERE account_code = %s RETURNING *", (row["account_code"],))
                deleted_bn.extend(rows_to_dicts(cur.fetchall()))
            else:
                cur.execute("""
                    UPDATE accounting_accounts
                    SET active = FALSE,
                        accepts_posting = FALSE,
                        locked = TRUE,
                        locked_by = %s,
                        locked_at = COALESCE(locked_at, NOW()),
                        updated_by = %s,
                        updated_at = NOW()
                    WHERE account_code = %s
                    RETURNING *
                """, (user, user, row["account_code"]))
                inactive_bn.extend(rows_to_dicts(cur.fetchall()))
        cur.execute("""
            DELETE FROM accounting_ledger
            WHERE account_code = '1.1.02.01'
               OR LOWER(COALESCE(account_name, '')) LIKE '%banco nacional%'
        """)

        cur.execute("""
            WITH used_accounts AS (
                SELECT DISTINCT account_code
                FROM accounting_lines
                WHERE account_code IS NOT NULL AND BTRIM(account_code) <> ''
            )
            UPDATE accounting_accounts a
            SET active = FALSE,
                accepts_posting = FALSE,
                locked = TRUE,
                locked_by = %s,
                locked_at = COALESCE(locked_at, NOW()),
                updated_by = %s,
                updated_at = NOW()
            WHERE a.active = TRUE
              AND NOT EXISTS (
                  SELECT 1 FROM used_accounts u WHERE u.account_code = a.account_code
              )
            RETURNING *
        """, (user, user))
        inactivated_unused = rows_to_dicts(cur.fetchall())

        cur.execute("""
            UPDATE accounting_accounts
            SET locked = TRUE,
                locked_by = COALESCE(locked_by, %s),
                locked_at = COALESCE(locked_at, NOW()),
                updated_by = %s,
                updated_at = NOW()
            WHERE active = TRUE
            RETURNING *
        """, (user, user))
        locked_active = rows_to_dicts(cur.fetchall())

        audit_event(
            cur,
            module="accounting",
            action="CHART_OF_ACCOUNTS_HARDENED",
            entity_type="accounting_chart",
            entity_id="MSL-CR",
            performed_by=user,
            performed_role=header_role,
            before={"accounts": before_accounts},
            after={
                "deleted_banco_nacional": deleted_bn,
                "inactive_banco_nacional": inactive_bn,
                "inactivated_unused": inactivated_unused,
                "locked_active_count": len(locked_active),
            },
            metadata={"reason": payload.get("reason") or "Definitive chart cleanup"},
        )
    conn.commit()
    return {
        "status": "ok",
        "deleted_banco_nacional": len(deleted_bn),
        "inactive_banco_nacional": len(inactive_bn),
        "inactivated_unused": len(inactivated_unused),
        "locked_active": len(locked_active),
    }


@router.get("/posting-rules")
def get_accounting_posting_rules(
    origin: str | None = None,
    include_inactive: bool = False,
    conn=Depends(get_db),
):
    _ensure_accounting_professional_schema(conn)
    return {"data": list_posting_rules(conn, origin=origin, include_inactive=include_inactive)}


@router.post("/posting-rules/seed")
def seed_accounting_posting_rules(
    payload: dict | None = None,
    conn=Depends(get_db),
    x_user: str | None = Header(None, alias="X-User"),
    x_role: str | None = Header(None, alias="X-Role"),
    x_user_role: str | None = Header(None, alias="X-User-Role"),
):
    _ensure_accounting_professional_schema(conn)
    payload = payload or {}
    header_user, header_role = actor_from_headers(x_user, x_role, x_user_role)
    user = str(payload.get("user") or header_user or "system")
    with conn.cursor(cursor_factory=RealDictCursor) as cur:
        before = rows_to_dicts(list_posting_rules(conn, include_inactive=True))
        result = seed_default_posting_rules(cur, user=user)
        after = rows_to_dicts(list_posting_rules(conn, include_inactive=True))
        audit_event(
            cur,
            module="accounting",
            action="POSTING_RULES_SEEDED",
            entity_type="accounting_posting_rules",
            entity_id="FORMAL_ENGINE",
            performed_by=user,
            performed_role=header_role,
            before={"rules": before},
            after={"rules": after},
            metadata={"reason": payload.get("reason") or "Seed formal accounting engine"},
        )
    conn.commit()
    return {"status": "ok", **result}


@router.get("/period-controls")
def list_accounting_period_controls(
    company_code: str | None = None,
    x_company_code: str | None = Header(None, alias="X-Company-Code"),
    conn=Depends(get_db),
):
    _ensure_accounting_professional_schema(conn)
    company = _company_code(company_code, x_company_code)
    with conn.cursor(cursor_factory=RealDictCursor) as cur:
        cur.execute(
            "SELECT * FROM accounting_period_controls WHERE company_code=%s ORDER BY period DESC",
            (company,),
        )
        return {"data": cur.fetchall()}


@router.post("/period-controls/{period}/close")
def close_accounting_period(period: str, payload: dict, conn=Depends(get_db)):
    _ensure_accounting_professional_schema(conn)
    company = str(payload.get("company_code") or "MSL-CR")
    user = str(payload.get("user") or "unknown")
    if len(period) != 7 or period[4] != "-":
        raise HTTPException(400, "period must use YYYY-MM")
    try:
        from routers.accounting_workspace import _close_snapshot, _ensure_schema as _ensure_workspace_schema
        _ensure_workspace_schema(conn)
    except HTTPException:
        raise
    with conn.cursor(cursor_factory=RealDictCursor) as cur:
        checklist, validation, summary, blockers = _close_snapshot(cur, period, company)
        if blockers:
            raise HTTPException(409, f"Robust close blocked: {len(blockers)} mandatory controls are pending or failing.")
        if validation["counts"]["critical"]:
            raise HTTPException(409, f"Robust close blocked: {validation['counts']['critical']} critical validations remain.")
        cur.execute(
            "SELECT COUNT(*) AS count FROM accounting_entries WHERE company_code=%s AND period=%s AND workflow_status <> 'POSTED'",
            (company, period),
        )
        pending = int(cur.fetchone()["count"])
        if pending:
            raise HTTPException(409, f"Period has {pending} entries that are not posted")
        cur.execute("""
            INSERT INTO accounting_period_controls(company_code, period, status, closed_by, closed_at)
            VALUES (%s,%s,'CLOSED',%s,NOW())
            ON CONFLICT(company_code,period) DO UPDATE SET
                status='CLOSED', closed_by=EXCLUDED.closed_by, closed_at=NOW(), updated_at=NOW()
            RETURNING *
        """, (company, period, user))
        row = cur.fetchone()
    conn.commit()
    return {"status": "ok", "period": row}


@router.post("/period-controls/{period}/reopen")
def reopen_accounting_period(period: str, payload: dict, conn=Depends(get_db)):
    _ensure_accounting_professional_schema(conn)
    reason = str(payload.get("reason") or "").strip()
    if not reason:
        raise HTTPException(400, "A reopen reason is required")
    company = str(payload.get("company_code") or "MSL-CR")
    user = str(payload.get("user") or "unknown")
    with conn.cursor(cursor_factory=RealDictCursor) as cur:
        cur.execute("""
            UPDATE accounting_period_controls SET status='OPEN', reopened_by=%s,
                reopened_at=NOW(), reopen_reason=%s, updated_at=NOW()
            WHERE company_code=%s AND period=%s RETURNING *
        """, (user, reason, company, period))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, "Period control not found")
    conn.commit()
    return {"status": "ok", "period": row}


# ============================================================
# GET SINGLE ACCOUNTING ENTRY (FOR POPUP EDIT)
# ============================================================
@router.get("/entry/{entry_id}")
def get_accounting_entry(
    entry_id: int,
    company_code: str | None = None,
    x_company_code: str | None = Header(None, alias="X-Company-Code"),
    conn=Depends(get_db)
):
    """
    Devuelve un asiento contable completo (cabecera + líneas)
    para edición en popup
    """

    _ensure_accounting_professional_schema(conn)
    company = _company_code(company_code, x_company_code)
    cur = conn.cursor(cursor_factory=RealDictCursor)

    # --------------------------------------------------------
    # 1. Traer cabecera
    # --------------------------------------------------------
    cur.execute("""
        SELECT
            id AS entry_id,
            entry_date,
            period,
            description,
            origin,
            origin_id,
            workflow_status,
            version,
            company_code,
            currency_code,
            exchange_rate,
            created_by,
            submitted_by,
            approved_by,
            posted_by
        FROM accounting_entries
        WHERE id = %s AND company_code = %s
    """, (entry_id, company))

    entry = cur.fetchone()

    if not entry:
        raise HTTPException(status_code=404, detail="Asiento no encontrado")

    # --------------------------------------------------------
    # 2. Traer líneas
    # --------------------------------------------------------
    cur.execute("""
        SELECT
            id AS line_id,
            account_code,
            account_name,
            debit,
            credit,
            line_description
        FROM accounting_lines
        WHERE entry_id = %s
        ORDER BY id
    """, (entry_id,))

    lines = cur.fetchall()

    # --------------------------------------------------------
    # 3. Respuesta final
    # --------------------------------------------------------
    return {
        "entry_id": entry["entry_id"],
        "entry_date": entry["entry_date"],
        "period": entry["period"],
        "description": entry["description"],
        "origin": entry["origin"],
        "origin_id": entry["origin_id"],
        "workflow_status": entry["workflow_status"],
        "version": entry["version"],
        "company_code": entry["company_code"],
        "currency_code": entry["currency_code"],
        "exchange_rate": float(entry["exchange_rate"] or 1),
        "created_by": entry["created_by"],
        "submitted_by": entry["submitted_by"],
        "approved_by": entry["approved_by"],
        "posted_by": entry["posted_by"],
        "lines": [
            {
                "line_id": l["line_id"],
                "account_code": l["account_code"],
                "account_name": l["account_name"],
                "debit": float(l["debit"] or 0),
                "credit": float(l["credit"] or 0),
                "line_description": l["line_description"]
            }
            for l in lines
        ]
    }


# ============================================================
# UPDATE ACCOUNTING ENTRY (POPUP EDIT)
# ============================================================
@router.put("/entry/{entry_id}")
def update_accounting_entry(
    entry_id: int,
    payload: dict,
    conn=Depends(get_db),
    x_user: str | None = Header(None, alias="X-User"),
    x_role: str | None = Header(None, alias="X-Role"),
    x_user_role: str | None = Header(None, alias="X-User-Role"),
):
    """
    Actualiza descripción del asiento y líneas contables.
    Valida partida doble.
    """

    _ensure_accounting_professional_schema(conn)
    _ensure_accounting_professional_schema(conn)
    cur = conn.cursor(cursor_factory=RealDictCursor)
    performed_by, performed_role = actor_from_headers(x_user, x_role, x_user_role)

    description = payload.get("description")
    lines = payload.get("lines", [])

    if not lines:
        raise HTTPException(status_code=400, detail="No se enviaron líneas")

    # --------------------------------------------------------
    # 1. VALIDACIONES CONTABLES
    # --------------------------------------------------------
    total_debit = Decimal("0.00")
    total_credit = Decimal("0.00")

    for line in lines:
        debit = _money(line.get("debit"), "debit")
        credit = _money(line.get("credit"), "credit")

        if debit > 0 and credit > 0:
            raise HTTPException(
                status_code=400,
                detail="Una línea no puede tener Debe y Haber simultáneamente"
            )

        total_debit += debit
        total_credit += credit

    if total_debit != total_credit or total_debit == 0:
        raise HTTPException(
            status_code=400,
            detail="La partida no está balanceada (Debe ≠ Haber)"
        )

    # --------------------------------------------------------
    # 2. VALIDAR QUE EL ASIENTO EXISTE
    # --------------------------------------------------------
    cur.execute(
        "SELECT * FROM accounting_entries WHERE id = %s FOR UPDATE",
        (entry_id,)
    )
    existing = cur.fetchone()
    if not existing:
        raise HTTPException(status_code=404, detail="Asiento no encontrado")
    if existing.get("workflow_status") != "DRAFT":
        raise HTTPException(status_code=409, detail="Only draft entries can be edited; use an adjustment or reversal")
    _assert_period_open(cur, existing["period"], existing.get("company_code") or "MSL-CR")
    expected_version = payload.get("expected_version")
    if expected_version is not None and int(expected_version) != int(existing.get("version") or 1):
        raise HTTPException(status_code=409, detail="The entry changed. Reload before saving")

    cur.execute("SELECT * FROM accounting_lines WHERE entry_id = %s ORDER BY id", (entry_id,))
    before_snapshot = {"entry": row_to_dict(existing), "lines": rows_to_dicts(cur.fetchall())}

    # --------------------------------------------------------
    # 3. ACTUALIZAR CABECERA
    # --------------------------------------------------------
    if description is not None:
        cur.execute("""
            UPDATE accounting_entries
            SET description = %s, version = version + 1, updated_at = NOW()
            WHERE id = %s
        """, (description, entry_id))

    # --------------------------------------------------------
    # 4. ACTUALIZAR LÍNEAS
    # --------------------------------------------------------
    for line in lines:
        line_id = line.get("line_id")

        if not line_id:
            raise HTTPException(
                status_code=400,
                detail="line_id es obligatorio"
            )

        # validar cuenta contable
        cur.execute("""
            SELECT account_name
            FROM accounting_accounts
            WHERE account_code = %s AND active = TRUE AND accepts_posting = TRUE
        """, (line["account_code"],))

        acc = cur.fetchone()
        if not acc:
            raise HTTPException(
                status_code=400,
                detail=f"Cuenta contable inválida: {line['account_code']}"
            )

        cur.execute("""
            UPDATE accounting_lines
            SET
                account_code = %s,
                account_name = %s,
                debit = %s,
                credit = %s,
                line_description = %s
            WHERE id = %s
              AND entry_id = %s
        """, (
            line["account_code"],
            acc["account_name"],
            line.get("debit", 0),
            line.get("credit", 0),
            line.get("line_description"),
            line_id,
            entry_id
        ))

    cur.execute("SELECT * FROM accounting_entries WHERE id = %s", (entry_id,))
    after_entry = row_to_dict(cur.fetchone())
    cur.execute("SELECT * FROM accounting_lines WHERE entry_id = %s ORDER BY id", (entry_id,))
    after_lines = rows_to_dicts(cur.fetchall())
    audit_event(
        cur,
        module="accounting",
        action="ENTRY_UPDATED",
        entity_type="accounting_entry",
        entity_id=entry_id,
        performed_by=performed_by,
        performed_role=performed_role,
        before=before_snapshot,
        after={"entry": after_entry, "lines": after_lines},
        metadata={"total_debit": str(total_debit), "total_credit": str(total_credit)},
    )

    conn.commit()

    return {
        "status": "ok",
        "message": "Asiento actualizado correctamente"
    }


def _transition_entry(conn, entry_id, payload, expected_status, new_status, action):
    _ensure_accounting_professional_schema(conn)
    user = str((payload or {}).get("user") or "unknown")
    role = str((payload or {}).get("role") or "").strip().lower() or None
    reason = str((payload or {}).get("reason") or "").strip() or None
    try:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("SELECT * FROM accounting_entries WHERE id=%s FOR UPDATE", (entry_id,))
            entry = cur.fetchone()
            if not entry:
                raise HTTPException(404, "Accounting entry not found")
            current = entry.get("workflow_status") or "POSTED"
            if current != expected_status:
                raise HTTPException(409, f"Entry is {current}; expected {expected_status}")
            _assert_period_open(cur, entry["period"], entry.get("company_code") or "MSL-CR")
            if new_status in ("APPROVED", "POSTED"):
                _assert_entry_can_advance(cur, entry_id)
            if new_status == "IN_REVIEW":
                fields = "submitted_by=%s, submitted_at=NOW()"
            elif new_status == "APPROVED":
                if entry.get("created_by") and entry.get("created_by") == user:
                    raise HTTPException(409, "The preparer cannot approve the same entry")
                fields = "approved_by=%s, approved_at=NOW()"
            else:
                fields = "posted_by=%s, posted_at=NOW()"
            cur.execute(f"""
                UPDATE accounting_entries SET workflow_status=%s, {fields},
                    version=version+1, updated_at=NOW()
                WHERE id=%s RETURNING *
            """, (new_status, user, entry_id))
            updated = cur.fetchone()
            _audit(cur, entry_id, action, user, current, new_status, reason)
            audit_event(
                cur,
                module="accounting",
                action=f"ENTRY_{action}",
                entity_type="accounting_entry",
                entity_id=entry_id,
                performed_by=user,
                performed_role=role,
                reason=reason,
                before=row_to_dict(entry),
                after=row_to_dict(updated),
                metadata={"previous_status": current, "new_status": new_status},
            )
        conn.commit()
        return {"status": "ok", "entry": updated}
    except HTTPException:
        conn.rollback()
        raise
    except Exception:
        conn.rollback()
        raise


@router.post("/entry/{entry_id}/submit")
def submit_accounting_entry(entry_id: int, payload: dict, conn=Depends(get_db)):
    return _transition_entry(conn, entry_id, payload, "DRAFT", "IN_REVIEW", "SUBMIT")


@router.post("/entry/{entry_id}/approve")
def approve_accounting_entry(entry_id: int, payload: dict, conn=Depends(get_db)):
    return _transition_entry(conn, entry_id, payload, "IN_REVIEW", "APPROVED", "APPROVE")


@router.post("/entry/{entry_id}/post")
def post_accounting_entry(entry_id: int, payload: dict, conn=Depends(get_db)):
    return _transition_entry(conn, entry_id, payload, "APPROVED", "POSTED", "POST")


@router.get("/entry/{entry_id}/audit")
def get_accounting_entry_audit(entry_id: int, conn=Depends(get_db)):
    _ensure_accounting_professional_schema(conn)
    with conn.cursor(cursor_factory=RealDictCursor) as cur:
        cur.execute("""
            SELECT id, entry_id, action, previous_status, new_status,
                   performed_by, reason, created_at
            FROM accounting_audit_log WHERE entry_id=%s ORDER BY created_at DESC, id DESC
        """, (entry_id,))
        return {"data": cur.fetchall()}


@router.get("/audit")
def get_finance_audit(
    module: str | None = None,
    entity_type: str | None = None,
    entity_id: str | None = None,
    performed_by: str | None = None,
    limit: int = 200,
    conn=Depends(get_db),
):
    _ensure_accounting_professional_schema(conn)
    limit = max(1, min(int(limit or 200), 1000))
    conditions = []
    params = []
    if module:
        conditions.append("module = %s")
        params.append(module)
    if entity_type:
        conditions.append("entity_type = %s")
        params.append(entity_type)
    if entity_id:
        conditions.append("entity_id = %s")
        params.append(str(entity_id))
    if performed_by:
        conditions.append("LOWER(COALESCE(performed_by,'')) LIKE LOWER(%s)")
        params.append(f"%{performed_by}%")
    where = "WHERE " + " AND ".join(conditions) if conditions else ""
    with conn.cursor(cursor_factory=RealDictCursor) as cur:
        ensure_finance_audit_schema(cur)
        cur.execute(f"""
            SELECT id, module, action, entity_type, entity_id, performed_by,
                   performed_role, reason, before_snapshot, after_snapshot,
                   metadata, created_at
            FROM finance_audit_log
            {where}
            ORDER BY created_at DESC, id DESC
            LIMIT %s
        """, [*params, limit])
        return {"data": cur.fetchall()}


@router.get("/audit/users")
def get_finance_audit_users(conn=Depends(get_db)):
    _ensure_accounting_professional_schema(conn)
    finance_roles = {"admin", "master", "accounting", "finance", "finanzas", "gerencia"}
    finance_users = {"admin", "aaron01", "gerencia1", "accountant"}
    with conn.cursor(cursor_factory=RealDictCursor) as cur:
        cur.execute("""
            SELECT column_name
            FROM information_schema.columns
            WHERE table_schema = 'public' AND table_name = 'usuarios'
        """)
        columns = {row["column_name"] for row in cur.fetchall()}
        user_col = next((c for c in ("usuario", "username", "user_name") if c in columns), None)
        name_col = next((c for c in ("nombre", "name", "full_name") if c in columns), None)
        role_col = next((c for c in ("rol", "role", "role_code") if c in columns), None)
        active_col = next((c for c in ("activo", "active", "is_active") if c in columns), None)
        if not user_col:
            return {"data": []}
        select_cols = [f"{user_col} AS usuario"]
        select_cols.append(f"{name_col} AS nombre" if name_col else "NULL::text AS nombre")
        select_cols.append(f"{role_col} AS rol" if role_col else "NULL::text AS rol")
        where = ""
        if active_col:
            where = f"WHERE COALESCE({active_col}, TRUE) = TRUE"
        cur.execute(f"""
            SELECT {', '.join(select_cols)}
            FROM usuarios
            {where}
            ORDER BY {user_col}
        """)
        rows = []
        for row in cur.fetchall():
            usuario = str(row.get("usuario") or "").strip()
            rol = str(row.get("rol") or "").strip().lower()
            if not usuario:
                continue
            if usuario.lower() in finance_users or rol in finance_roles:
                rows.append(row)
        return {"data": rows}


@router.get("/validation-alerts")
def get_accounting_validation_alerts(
    period: str | None = None,
    period_from: str | None = None,
    period_to: str | None = None,
    origin: str | None = None,
    company_code: str | None = None,
    x_company_code: str | None = Header(None, alias="X-Company-Code"),
    limit: int = 200,
    conn=Depends(get_db),
):
    _ensure_accounting_professional_schema(conn)
    company = _company_code(company_code, x_company_code)
    limit = max(1, min(int(limit or 200), 1000))
    alerts = []

    def add(severity, code, title, message, entity_type=None, entity_id=None, metadata=None):
        alerts.append({
            "severity": severity,
            "code": code,
            "title": title,
            "message": message,
            "entity_type": entity_type,
            "entity_id": str(entity_id) if entity_id is not None else None,
            "metadata": metadata or {},
        })

    conditions = ["e.company_code = %s"]
    params = [company]
    if period:
        conditions.append("e.period = %s")
        params.append(period)
    if period_from:
        conditions.append("e.period >= %s")
        params.append(period_from)
    if period_to:
        conditions.append("e.period <= %s")
        params.append(period_to)
    if origin and origin != "TODOS":
        conditions.append("e.origin = %s")
        params.append(origin)
    where = "WHERE " + " AND ".join(conditions) if conditions else ""

    with conn.cursor(cursor_factory=RealDictCursor) as cur:
        backfill_missing_bank_accounts(cur)
        conn.commit()
        cur.execute(f"""
            SELECT e.id, e.period, e.origin, e.description,
                   COALESCE(SUM(l.debit),0) AS total_debit,
                   COALESCE(SUM(l.credit),0) AS total_credit,
                   ROUND((COALESCE(SUM(l.debit),0) - COALESCE(SUM(l.credit),0))::numeric, 2) AS difference
            FROM accounting_entries e
            JOIN accounting_lines l ON l.entry_id = e.id
            {where}
            GROUP BY e.id, e.period, e.origin, e.description
            HAVING ROUND((COALESCE(SUM(l.debit),0) - COALESCE(SUM(l.credit),0))::numeric, 2) <> 0
            ORDER BY ABS(ROUND((COALESCE(SUM(l.debit),0) - COALESCE(SUM(l.credit),0))::numeric, 2)) DESC
            LIMIT %s
        """, [*params, limit])
        for row in cur.fetchall():
            add(
                "critical",
                "UNBALANCED_ENTRY",
                "Asiento descuadrado",
                f"Asiento {row['id']} periodo {row['period']} tiene diferencia {row['difference']}.",
                "accounting_entry",
                row["id"],
                row,
            )

        cur.execute(f"""
            SELECT e.id, e.period, e.origin, e.description
            FROM accounting_entries e
            LEFT JOIN accounting_lines l ON l.entry_id = e.id
            {where}
            GROUP BY e.id, e.period, e.origin, e.description
            HAVING COUNT(l.id) = 0
            ORDER BY e.id DESC
            LIMIT %s
        """, [*params, limit])
        for row in cur.fetchall():
            add(
                "critical",
                "ENTRY_WITHOUT_LINES",
                "Asiento sin lineas",
                f"Asiento {row['id']} no tiene lineas contables.",
                "accounting_entry",
                row["id"],
                row,
            )

        cur.execute(f"""
            SELECT l.id, l.entry_id, e.period, l.account_code, l.account_name, l.debit, l.credit
            FROM accounting_lines l
            JOIN accounting_entries e ON e.id = l.entry_id
            {where}
              {"AND" if where else "WHERE"} (
                    (COALESCE(l.debit,0) > 0 AND COALESCE(l.credit,0) > 0)
                 OR (COALESCE(l.debit,0) = 0 AND COALESCE(l.credit,0) = 0)
              )
            ORDER BY l.id DESC
            LIMIT %s
        """, [*params, limit])
        for row in cur.fetchall():
            add(
                "critical",
                "INVALID_LINE_AMOUNT",
                "Linea con monto invalido",
                f"Linea {row['id']} del asiento {row['entry_id']} tiene monto invalido.",
                "accounting_line",
                row["id"],
                row,
            )

        cur.execute(f"""
            SELECT l.id, l.entry_id, e.period, l.account_code, l.account_name
            FROM accounting_lines l
            JOIN accounting_entries e ON e.id = l.entry_id
            LEFT JOIN accounting_accounts a ON a.account_code = l.account_code
            {where}
              {"AND" if where else "WHERE"} (a.account_code IS NULL OR a.active = FALSE)
            ORDER BY l.id DESC
            LIMIT %s
        """, [*params, limit])
        for row in cur.fetchall():
            add(
                "warning",
                "ACCOUNT_NOT_ACTIVE",
                "Cuenta no activa o inexistente",
                f"Linea {row['id']} usa cuenta {row['account_code']} no activa en catalogo contable.",
                "accounting_line",
                row["id"],
                row,
            )

        if company == "MSL-CR" and (not origin or origin in ("TODOS", "CASH_APP")):
            cur.execute("""
                SELECT id, numero_documento, nombre_cliente, banco, bank_account_code, bank_account_name, fecha_pago
                FROM cash_app
                WHERE tipo_aplicacion = 'PAGO'
                  AND (bank_account_code IS NULL OR BTRIM(bank_account_code) = '')
                ORDER BY id DESC
                LIMIT %s
            """, (limit,))
            for row in cur.fetchall():
                add(
                    "warning",
                    "COLLECTION_PAYMENT_WITHOUT_BANK_ACCOUNT",
                    "Pago Collections sin banco contable especifico",
                    f"Pago cash_app {row['id']} no tiene cuenta bancaria contable seleccionada.",
                    "cash_app",
                    row["id"],
                    row,
                )

        if company == "MSL-CR" and (not origin or origin in ("TODOS", "COLLECTIONS")):
            cur.execute("""
                SELECT c.id, c.numero_documento, c.nombre_cliente, c.total, c.moneda, c.fecha_emision
                FROM collections c
                LEFT JOIN accounting_entries e
                  ON e.origin='COLLECTIONS'
                 AND e.origin_id=c.id
                 AND e.workflow_status='POSTED'
                WHERE COALESCE(c.total,0)>0
                  AND COALESCE(c.fecha_emision::date, CURRENT_DATE) <= CURRENT_DATE
                  AND e.id IS NULL
                ORDER BY c.id DESC
                LIMIT %s
            """, (limit,))
            for row in cur.fetchall():
                add(
                    "critical",
                    "COLLECTION_INVOICE_WITHOUT_POSTED_ENTRY",
                    "Factura de venta sin asiento contabilizado",
                    f"Factura {row['numero_documento']} de {row['nombre_cliente']} no tiene asiento POSTED.",
                    "collections",
                    row["id"],
                    row,
                )

            cur.execute("""
                SELECT e.id entry_id, c.id collection_id, c.numero_documento, c.nombre_cliente,
                       COUNT(*) FILTER(WHERE l.account_code IN ('1.1.04.01','1101') AND COALESCE(l.debit,0)>0) ar_lines,
                       COUNT(*) FILTER(WHERE (l.account_code LIKE '4%%' OR l.account_code='4101') AND COALESCE(l.credit,0)>0) revenue_lines
                FROM collections c
                JOIN accounting_entries e ON e.origin='COLLECTIONS' AND e.origin_id=c.id AND e.workflow_status='POSTED'
                LEFT JOIN accounting_lines l ON l.entry_id=e.id
                WHERE COALESCE(c.total,0)>0
                  AND COALESCE(c.fecha_emision::date, CURRENT_DATE) <= CURRENT_DATE
                GROUP BY e.id,c.id,c.numero_documento,c.nombre_cliente
                HAVING COUNT(*) FILTER(WHERE l.account_code IN ('1.1.04.01','1101') AND COALESCE(l.debit,0)>0)=0
                    OR COUNT(*) FILTER(WHERE (l.account_code LIKE '4%%' OR l.account_code='4101') AND COALESCE(l.credit,0)>0)=0
                ORDER BY e.id DESC
                LIMIT %s
            """, (limit,))
            for row in cur.fetchall():
                add(
                    "critical",
                    "COLLECTION_INVOICE_BAD_CLASSIFICATION",
                    "Factura de venta con clasificacion incompleta",
                    f"Factura {row['numero_documento']} debe tener debito a CxC y credito a ingresos.",
                    "accounting_entry",
                    row["entry_id"],
                    row,
                )

        if company == "MSL-CR" and (not origin or origin in ("TODOS", "ITP", "ITP_PAYMENT")):
            cur.execute("""
                SELECT id, payee_name, reference, status, payment_bank_account_code, payment_bank_account_name, last_payment_date
                FROM payment_obligations
                WHERE active = TRUE
                  AND status IN ('PAID', 'PARTIAL')
                  AND last_payment_date IS NOT NULL
                  AND (payment_bank_account_code IS NULL OR BTRIM(payment_bank_account_code) = '')
                ORDER BY id DESC
                LIMIT %s
            """, (limit,))
            for row in cur.fetchall():
                add(
                    "warning",
                    "ITP_PAYMENT_WITHOUT_BANK_ACCOUNT",
                    "Pago ITP sin banco contable especifico",
                    f"Obligacion ITP {row['id']} tiene pago sin cuenta bancaria contable.",
                    "payment_obligation",
                    row["id"],
                    row,
                )

            cur.execute("""
                SELECT p.id, p.payee_name, p.reference, p.total, p.currency, p.issue_date
                FROM payment_obligations p
                LEFT JOIN accounting_entries e
                  ON e.origin='ITP'
                 AND e.origin_id=p.id
                 AND e.workflow_status='POSTED'
                WHERE p.active=TRUE
                  AND p.record_type='OBLIGATION'
                  AND COALESCE(p.total,0)>0
                  AND COALESCE(p.issue_date, CURRENT_DATE) <= CURRENT_DATE
                  AND e.id IS NULL
                ORDER BY p.id DESC
                LIMIT %s
            """, (limit,))
            for row in cur.fetchall():
                add(
                    "critical",
                    "ITP_OBLIGATION_WITHOUT_POSTED_ENTRY",
                    "Factura/obligacion de compra sin asiento contabilizado",
                    f"ITP {row['id']} {row['payee_name']} no tiene asiento POSTED.",
                    "payment_obligation",
                    row["id"],
                    row,
                )

            cur.execute("""
                SELECT e.id entry_id, p.id obligation_id, p.payee_name, p.reference,
                       COUNT(*) FILTER(WHERE l.account_code='2.1.01.01' AND COALESCE(l.credit,0)>0) payable_lines,
                       COUNT(*) FILTER(WHERE (l.account_code LIKE '5%%' OR l.account_code LIKE '1%%') AND COALESCE(l.debit,0)>0) debit_lines
                FROM payment_obligations p
                JOIN accounting_entries e ON e.origin='ITP' AND e.origin_id=p.id AND e.workflow_status='POSTED'
                LEFT JOIN accounting_lines l ON l.entry_id=e.id
                WHERE p.active=TRUE
                  AND p.record_type='OBLIGATION'
                  AND COALESCE(p.total,0)>0
                  AND COALESCE(p.issue_date, CURRENT_DATE) <= CURRENT_DATE
                GROUP BY e.id,p.id,p.payee_name,p.reference
                HAVING COUNT(*) FILTER(WHERE l.account_code='2.1.01.01' AND COALESCE(l.credit,0)>0)=0
                    OR COUNT(*) FILTER(WHERE (l.account_code LIKE '5%%' OR l.account_code LIKE '1%%') AND COALESCE(l.debit,0)>0)=0
                ORDER BY e.id DESC
                LIMIT %s
            """, (limit,))
            for row in cur.fetchall():
                add(
                    "critical",
                    "ITP_OBLIGATION_BAD_CLASSIFICATION",
                    "Factura/obligacion de compra con clasificacion incompleta",
                    f"ITP {row['obligation_id']} debe tener debito a gasto/activo y credito a CxP.",
                    "accounting_entry",
                    row["entry_id"],
                    row,
                )

        cur.execute(f"""
            SELECT e.origin, e.origin_id, COUNT(*) AS count
            FROM accounting_entries e
            {where}
              {"AND" if where else "WHERE"} e.origin_id IS NOT NULL
            GROUP BY e.origin, e.origin_id
            HAVING COUNT(*) > 1
            ORDER BY COUNT(*) DESC
            LIMIT %s
        """, [*params, limit])
        for row in cur.fetchall():
            add(
                "warning",
                "DUPLICATE_ORIGIN_ENTRY",
                "Posible duplicidad de origen",
                f"Origen {row['origin']} #{row['origin_id']} tiene {row['count']} asientos.",
                "accounting_entry",
                row["origin_id"],
                row,
            )

    counts = {
        "critical": sum(1 for item in alerts if item["severity"] == "critical"),
        "warning": sum(1 for item in alerts if item["severity"] == "warning"),
        "info": sum(1 for item in alerts if item["severity"] == "info"),
    }
    return {
        "status": "ok",
        "period": period,
        "period_from": period_from,
        "period_to": period_to,
        "origin": origin,
        "counts": counts,
        "alerts": alerts[:limit],
    }


@router.post("/sync/collections")
def sync_collections(
    x_company_code: str | None = Header(None, alias="X-Company-Code"),
    conn=Depends(get_db),
):
    _ensure_accounting_professional_schema(conn)
    from services.accounting_auto import sync_collections_to_accounting

    sync_collections_to_accounting(conn, company_code_filter=_company_code(header_value=x_company_code))

    return {
        "status": "ok",
        "message": "Collections sincronizadas con Accounting"
    }



@router.post("/sync/cash-app")
def sync_cash_app(conn=Depends(get_db)):
    _ensure_accounting_professional_schema(conn)
    try:
        from services.accounting_auto import sync_cash_app_to_accounting
        sync_cash_app_to_accounting(conn)

        return {
            "status": "ok",
            "message": "Cash App sincronizado a Accounting"
        }

    except Exception as e:
        conn.rollback()
        raise HTTPException(500, repr(e))




@router.post("/sync/itp")
def sync_itp(conn=Depends(get_db)):
    _ensure_accounting_professional_schema(conn)
    try:
        from services.accounting_auto import sync_itp_to_accounting
        sync_itp_to_accounting(conn)

        return {
            "status": "ok",
            "message": "Invoice to Pay sincronizado a Accounting"
        }

    except Exception as e:
        conn.rollback()
        raise HTTPException(500, repr(e))


@router.post("/sync/payroll")
def sync_payroll(conn=Depends(get_db)):
    _ensure_accounting_professional_schema(conn)
    try:
        from services.accounting_auto import sync_payroll_to_accounting
        sync_payroll_to_accounting(conn)

        return {
            "status": "ok",
            "message": "Payroll sincronizado a Accounting"
        }

    except Exception as e:
        conn.rollback()
        raise HTTPException(500, repr(e))


@router.post("/sync/all")
def sync_all_accounting(
    x_company_code: str | None = Header(None, alias="X-Company-Code"),
    conn=Depends(get_db),
):
    _ensure_accounting_professional_schema(conn)
    try:
        from services.accounting_auto import (
            sync_cash_app_to_accounting,
            sync_collections_to_accounting,
            sync_itp_to_accounting,
            sync_payroll_to_accounting
        )

        before = _accounting_entry_stats(conn)
        company = _company_code(header_value=x_company_code)

        sync_collections_to_accounting(conn, company_code_filter=company)
        sync_cash_app_to_accounting(conn)
        sync_itp_to_accounting(conn)
        sync_payroll_to_accounting(conn)

        after = _accounting_entry_stats(conn)

        return {
            "status": "ok",
            "message": "Accounting sincronizado correctamente",
            "created": max(0, after["total_entries"] - before["total_entries"]),
            "before": before,
            "after": after,
            "latest_period": after["latest_period"],
            "period_counts": after["period_counts"]
        }

    except Exception as e:
        conn.rollback()
        raise HTTPException(500, repr(e))



@router.get("/ledger")
def get_accounting_ledger(
    period: str | None = None,
    period_from: str | None = None,
    period_to: str | None = None,
    origin: str | None = None,
    account_code: str | None = None,   # ✅ NUEVO FILTRO
    company_code: str | None = None,
    x_company_code: str | None = Header(None, alias="X-Company-Code"),
    conn=Depends(get_db)
):
    """
    Devuelve asientos contables agrupados por entry_id,
    con sus líneas (debe / haber)

    Filtros soportados:
    - period (YYYY-MM)
    - origin (COLLECTIONS, ITP, CASH_APP, MANUAL, etc.)
    - account_code (1101, 2101, 5101, etc.)
    """

    _ensure_accounting_professional_schema(conn)
    company = _company_code(company_code, x_company_code)
    cur = conn.cursor(cursor_factory=RealDictCursor)

    conditions = ["e.entry_date <= CURRENT_DATE", "e.company_code = %s"]
    params = [company]

    # -----------------------------
    # VALIDACIONES
    # -----------------------------
    if origin and not period:
        raise HTTPException(
            status_code=400,
            detail="period es obligatorio cuando se filtra por origin"
        )

    # -----------------------------
    # FILTROS
    # -----------------------------
    if period:
        conditions.append("e.period = %s")
        params.append(period)
    if period_from:
        conditions.append("e.period >= %s")
        params.append(period_from)
    if period_to:
        conditions.append("e.period <= %s")
        params.append(period_to)

    if origin:
        conditions.append("e.origin = %s")
        params.append(origin)

    _append_account_filter(conditions, params, account_code)

    where_clause = ""
    if conditions:
        where_clause = "WHERE " + " AND ".join(conditions)

    # -----------------------------
    # QUERY PRINCIPAL
    # -----------------------------
    query = f"""
        SELECT
            e.id AS entry_id,
            e.entry_date,
            e.period,
            e.description AS entry_description,
            e.origin,
            e.origin_id,
            e.workflow_status,
            e.version,

            l.id AS line_id,
            l.account_code,
            l.account_name,
            l.debit,
            l.credit,
            l.line_description

        FROM accounting_entries e
        JOIN accounting_lines l ON l.entry_id = e.id
        {where_clause}
        ORDER BY
            e.entry_date DESC,
            e.id DESC,
            l.id ASC
    """

    cur.execute(query, params)
    rows = cur.fetchall()

    # -----------------------------
    # AGRUPAR POR entry_id
    # -----------------------------
    entries = {}

    for row in rows:
        entry_id = row["entry_id"]

        if entry_id not in entries:
            entries[entry_id] = {
                "entry_id": entry_id,
                "entry_date": row["entry_date"],
                "period": row["period"],
                "description": row["entry_description"],
                "origin": row["origin"],
                "origin_id": row["origin_id"],
                "workflow_status": row["workflow_status"],
                "version": row["version"],
                "lines": []
            }

        entries[entry_id]["lines"].append({
            "line_id": row["line_id"],
            "account_code": row["account_code"],
            "account_name": row["account_name"],
            "debit": float(row["debit"] or 0),
            "credit": float(row["credit"] or 0),
            "line_description": row["line_description"]
        })

    return {
        "data": list(entries.values())
    }




# ============================================================
# IVA (FUENTE ÚNICA: accounting_lines.created_at)
# ============================================================
@router.get("/reports/excel")
def download_accounting_report_excel(
    report: str | None = "ASIENTOS",
    period: str | None = None,
    period_from: str | None = None,
    period_to: str | None = None,
    origin: str | None = None,
    account_code: str | None = None,
    conn=Depends(get_db)
):
    rows = _fetch_accounting_report_lines(conn, period, period_from, period_to, origin, account_code)
    wb = Workbook()
    ws = wb.active
    ws.title = "Accounting"

    title = _report_title(report)
    scope = period or (f"{period_from or 'inicio'} a {period_to or 'fin'}" if period_from or period_to else "Todos")
    ws.merge_cells("A1:J1")
    ws["A1"] = f"{title} - {scope}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = ["Fecha", "Asiento", "Periodo", "Origen", "Origen ID", "Cuenta", "Nombre cuenta", "Detalle", "Debe", "Haber"]
    ws.append([])
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="003A75")
    for cell in ws[3]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    total_debit = 0.0
    total_credit = 0.0
    for row in rows:
        debit = float(row.get("debit") or 0)
        credit = float(row.get("credit") or 0)
        total_debit += debit
        total_credit += credit
        ws.append([
            row.get("entry_date"),
            row.get("entry_id"),
            row.get("period"),
            row.get("origin"),
            row.get("origin_id"),
            row.get("account_code"),
            row.get("account_name"),
            row.get("line_description") or row.get("entry_description"),
            debit,
            credit
        ])

    ws.append(["", "", "", "", "", "", "", "Totales", total_debit, total_credit])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    widths = [14, 10, 12, 16, 14, 14, 28, 48, 14, 14]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width

    for row in ws.iter_rows(min_row=4, min_col=9, max_col=10):
        for cell in row:
            cell.number_format = '#,##0.00'

    tmp_dir = tempfile.mkdtemp(prefix="erp_som_accounting_")
    filename = _report_filename("xlsx", report, period, period_from, period_to)
    path = os.path.join(tmp_dir, filename)
    wb.save(path)

    return FileResponse(
        path,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@router.get("/reports/pdf")
def download_accounting_report_pdf(
    report: str | None = "ASIENTOS",
    period: str | None = None,
    period_from: str | None = None,
    period_to: str | None = None,
    origin: str | None = None,
    account_code: str | None = None,
    conn=Depends(get_db)
):
    rows = _fetch_accounting_report_lines(conn, period, period_from, period_to, origin, account_code)
    tmp_dir = tempfile.mkdtemp(prefix="erp_som_accounting_")
    filename = _report_filename("pdf", report, period, period_from, period_to)
    path = os.path.join(tmp_dir, filename)

    title = _report_title(report)
    scope = period or (f"{period_from or 'inicio'} a {period_to or 'fin'}" if period_from or period_to else "Todos")
    styles = getSampleStyleSheet()
    doc = SimpleDocTemplate(path, pagesize=landscape(letter), rightMargin=24, leftMargin=24, topMargin=24, bottomMargin=24)

    data = [["Fecha", "Asiento", "Periodo", "Origen", "Cuenta", "Detalle", "Debe", "Haber"]]
    total_debit = 0.0
    total_credit = 0.0
    for row in rows:
        debit = float(row.get("debit") or 0)
        credit = float(row.get("credit") or 0)
        total_debit += debit
        total_credit += credit
        data.append([
            str(row.get("entry_date") or ""),
            str(row.get("entry_id") or ""),
            str(row.get("period") or ""),
            str(row.get("origin") or ""),
            str(row.get("account_code") or ""),
            str(row.get("line_description") or row.get("entry_description") or "")[:80],
            f"{debit:,.2f}",
            f"{credit:,.2f}"
        ])
    data.append(["", "", "", "", "", "Totales", f"{total_debit:,.2f}", f"{total_credit:,.2f}"])

    table = Table(data, repeatRows=1, colWidths=[58, 48, 58, 72, 58, 270, 72, 72])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#003A75")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D7DEE8")),
        ("ALIGN", (6, 1), (7, -1), "RIGHT"),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#EEF3F8"))
    ]))

    doc.build([
        Paragraph(f"{title} - {scope}", styles["Title"]),
        Spacer(1, 12),
        table
    ])

    return FileResponse(path, filename=filename, media_type="application/pdf")


@router.get("/iva")
def get_accounting_iva(
    period: str,  # 'YYYY-MM'
    company_code: str | None = None,
    x_company_code: str | None = Header(None, alias="X-Company-Code"),
    conn=Depends(get_db)
):
    """
    IVA ERP-SOM - DEFINITIVO

    - Fuente ÚNICA: accounting_lines
    - Periodo = to_char(created_at,'YYYY-MM')
    - Siempre calcula el mes actual
    - Arrastra saldo a favor SOLO si existe en el mes anterior
    """

    _ensure_accounting_professional_schema(conn)
    company = _company_code(company_code, x_company_code)
    cur = conn.cursor(cursor_factory=RealDictCursor)

    try:
        start = datetime.strptime(period + "-01", "%Y-%m-%d").date()
        end = start.replace(year=start.year + 1, month=1) if start.month == 12 else start.replace(month=start.month + 1)
        if company == "MSL-CR":
            cur.execute("""
                WITH tax_ranked AS (
                    SELECT d.*,
                           ROW_NUMBER() OVER (
                               PARTITION BY d.direction,
                                            COALESCE(NULLIF(d.document_number, ''), NULLIF(d.electronic_key, ''), d.source_table || ':' || d.source_id, d.id::text)
                               ORDER BY
                                   CASE
                                       WHEN d.source_table IN ('hacienda_emitted_excel', 'hacienda_acceptance_excel', 'xml_upload') THEN 0
                                       WHEN d.xml_path IS NOT NULL THEN 1
                                       WHEN d.source_table IN ('invoicing', 'collections') THEN 2
                                       WHEN d.source_table = 'payment_obligations' THEN 3
                                       ELSE 4
                                   END,
                                   CASE WHEN COALESCE(d.tax_amount, 0) <> 0 THEN 0 ELSE 1 END,
                                   d.id DESC
                           ) AS tax_rank
                    FROM tax_electronic_documents d
                    WHERE d.issue_datetime >= %s
                      AND d.issue_datetime < %s
                      AND COALESCE(d.issue_datetime::date, CURRENT_DATE) <= CURRENT_DATE
                      AND (
                          d.source_table IN ('hacienda_emitted_excel', 'hacienda_acceptance_excel')
                          OR (
                              d.direction = 'SALE'
                              AND COALESCE(d.receiver_identification, '') IN ('3101065618', '3101660512')
                          )
                          OR (
                              d.direction = 'PURCHASE'
                              AND (
                                  COALESCE(d.receiver_identification, '') = '3102920372'
                                  OR UPPER(COALESCE(d.receiver_name, '')) LIKE '%%MSL%%'
                                  OR UPPER(COALESCE(d.receiver_name, '')) LIKE '%%MARINE SURVEYORS%%'
                              )
                          )
                      )
                )
                SELECT direction,
                       COALESCE(SUM(
                           CASE
                               WHEN UPPER(COALESCE(currency_code,'CRC')) IN ('CRC','COLON','COLONES')
                               THEN tax_amount
                               ELSE tax_amount * COALESCE(NULLIF(exchange_rate,0),1)
                           END
                       ),0) AS tax_crc
                FROM tax_ranked
                WHERE tax_rank = 1
                GROUP BY direction
            """, (start, end))
            rows = {r["direction"]: r for r in cur.fetchall()}
            iva_por_pagar = float(rows.get("SALE", {}).get("tax_crc") or 0)
            iva_credito = float(rows.get("PURCHASE", {}).get("tax_crc") or 0)
            if iva_por_pagar or iva_credito:
                return {
                    "period": period,
                    "iva_por_pagar": round(iva_por_pagar, 2),
                    "iva_credito": round(iva_credito, 2),
                    "saldo_favor_anterior": 0.0,
                    "iva_total": round(iva_por_pagar - iva_credito, 2),
                    "source": "tax_electronic_documents"
                }

        # -------------------------------------------------
        # Helper: IVA por periodo (100% SQL SAFE)
        # -------------------------------------------------
        def iva_por_periodo(p):
            cur.execute("""
                SELECT
                    SUM(
                        CASE
                            WHEN l.account_code IN ('2108', '2.1.02.03') -- IVA por pagar
                            THEN COALESCE(l.credit,0) - COALESCE(l.debit,0)
                            ELSE 0
                        END
                    ) AS iva_por_pagar,

                    SUM(
                        CASE
                            WHEN l.account_code IN ('1131', '1.1.13.99') -- IVA credito fiscal
                            THEN COALESCE(l.debit,0) - COALESCE(l.credit,0)
                            ELSE 0
                        END
                    ) AS iva_credito
                FROM accounting_lines l
                JOIN accounting_entries e ON e.id = l.entry_id
                WHERE e.period = %s
                  AND e.company_code = %s
            """, (p, company))

            row = cur.fetchone() or {}
            return (
                float(row.get("iva_por_pagar") or 0),
                float(row.get("iva_credito") or 0)
            )

        # -------------------------------------------------
        # 1️⃣ IVA DEL MES ACTUAL (SIEMPRE)
        # -------------------------------------------------
        iva_por_pagar, iva_credito = iva_por_periodo(period)

        # -------------------------------------------------
        # 2️⃣ PERIODO ANTERIOR
        # -------------------------------------------------
        year, month = map(int, period.split("-"))
        if month == 1:
            prev_period = f"{year-1}-12"
        else:
            prev_period = f"{year}-{month-1:02d}"

        prev_pagar, prev_credito = iva_por_periodo(prev_period)

        # -------------------------------------------------
        # 3️⃣ SALDO A FAVOR (SOLO SI EXISTE)
        # -------------------------------------------------
        if prev_credito > prev_pagar:
            saldo_favor_anterior = prev_credito - prev_pagar
        else:
            saldo_favor_anterior = 0.0

        # -------------------------------------------------
        # 4️⃣ IVA FINAL
        # -------------------------------------------------
        iva_total = iva_por_pagar - iva_credito - saldo_favor_anterior

        return {
            "period": period,
            "iva_por_pagar": round(iva_por_pagar, 2),
            "iva_credito": round(iva_credito, 2),
            "saldo_favor_anterior": round(saldo_favor_anterior, 2),
            "iva_total": round(iva_total, 2)
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=repr(e))

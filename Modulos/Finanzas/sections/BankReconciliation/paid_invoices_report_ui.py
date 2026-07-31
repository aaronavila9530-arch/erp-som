import calendar
from datetime import datetime
import tkinter as tk
from tkinter import ttk, messagebox, filedialog

from api_client import get_clientes_finanzas_api, get_paid_invoices_report_api
from Modulos.Finanzas.date_utils import to_long_english_date


class PaidInvoicesReportUI(tk.Frame):
    def __init__(self, parent):
        super().__init__(parent, bg="white")
        self.rows = []
        self.summary = {}
        self.clientes_map = {}

        self._build_header()
        self._build_filters()
        self._build_summary()
        self._build_table()
        self._load_clientes()

    def _build_header(self):
        header = tk.Frame(self, bg="white")
        header.pack(fill="x", padx=10, pady=(5, 0))

        ttk.Label(
            header,
            text="Reporte de Facturas Pagadas",
            font=("Segoe UI", 14, "bold"),
            background="white"
        ).pack(side="left")

        ttk.Label(
            header,
            text="Muestra cada factura y la fecha en que fue pagada.",
            foreground="#555",
            background="white"
        ).pack(side="left", padx=12)

    def _build_filters(self):
        frame = ttk.LabelFrame(self, text="Filtros")
        frame.pack(fill="x", padx=10, pady=8)

        year_values = [""] + [str(y) for y in range(datetime.now().year - 5, datetime.now().year + 2)]
        month_values = [""] + [f"{i:02d} - {calendar.month_name[i]}" for i in range(1, 13)]
        period_values = self._period_values()

        ttk.Label(frame, text="Ano").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.cmb_year = ttk.Combobox(frame, values=year_values, width=10, state="readonly")
        self.cmb_year.set(str(datetime.now().year))
        self.cmb_year.grid(row=0, column=1, padx=5, pady=5, sticky="w")

        ttk.Label(frame, text="Mes").grid(row=0, column=2, padx=5, pady=5, sticky="w")
        self.cmb_month = ttk.Combobox(frame, values=month_values, width=18, state="readonly")
        self.cmb_month.grid(row=0, column=3, padx=5, pady=5, sticky="w")

        ttk.Label(frame, text="Desde").grid(row=0, column=4, padx=5, pady=5, sticky="w")
        self.cmb_from = ttk.Combobox(frame, values=period_values, width=18, state="readonly")
        self.cmb_from.grid(row=0, column=5, padx=5, pady=5, sticky="w")

        ttk.Label(frame, text="Hasta").grid(row=0, column=6, padx=5, pady=5, sticky="w")
        self.cmb_to = ttk.Combobox(frame, values=period_values, width=18, state="readonly")
        self.cmb_to.grid(row=0, column=7, padx=5, pady=5, sticky="w")

        ttk.Label(frame, text="Cliente").grid(row=0, column=8, padx=5, pady=5, sticky="w")
        self.cmb_cliente = ttk.Combobox(frame, width=34, state="readonly")
        self.cmb_cliente.grid(row=0, column=9, padx=5, pady=5, sticky="ew")

        ttk.Button(frame, text="Buscar", command=self._search).grid(row=0, column=10, padx=5, pady=5)
        ttk.Button(frame, text="Limpiar", command=self._clear_filters).grid(row=0, column=11, padx=5, pady=5)

        ttk.Label(frame, text="Exportar").grid(row=0, column=12, padx=(18, 5), pady=5, sticky="w")
        self.cmb_export = ttk.Combobox(frame, values=["Excel", "PDF"], width=8, state="readonly")
        self.cmb_export.set("Excel")
        self.cmb_export.grid(row=0, column=13, padx=5, pady=5)
        ttk.Button(frame, text="Generar", command=self._export).grid(row=0, column=14, padx=5, pady=5)

        frame.columnconfigure(9, weight=1)

    def _period_values(self):
        current_year = datetime.now().year
        values = [""]
        for year in range(current_year - 5, current_year + 2):
            for month in range(1, 13):
                values.append(f"{year}-{month:02d} - {calendar.month_name[month]}")
        return values

    def _load_clientes(self):
        try:
            values = [""]
            for client in get_clientes_finanzas_api():
                codigo = client.get("codigo") or ""
                nombre = client.get("nombre") or ""
                label = f"{codigo} - {nombre}".strip(" -")
                if label:
                    values.append(label)
                    self.clientes_map[label] = codigo or nombre
            self.cmb_cliente["values"] = values
        except Exception:
            self.cmb_cliente["values"] = [""]

    def _build_summary(self):
        frame = ttk.LabelFrame(self, text="Resumen")
        frame.pack(fill="x", padx=10, pady=(0, 8))

        self.lbl_total = ttk.Label(frame, text="Pagos: 0")
        self.lbl_total.pack(side="left", padx=10, pady=6)

        self.lbl_invoices = ttk.Label(frame, text="Facturas: 0")
        self.lbl_invoices.pack(side="left", padx=10, pady=6)

        self.lbl_clients = ttk.Label(frame, text="Clientes: 0")
        self.lbl_clients.pack(side="left", padx=10, pady=6)

        self.lbl_paid = ttk.Label(frame, text="Total pagado: 0.00")
        self.lbl_paid.pack(side="left", padx=10, pady=6)

        self.lbl_commission = ttk.Label(frame, text="Comision: 0.00")
        self.lbl_commission.pack(side="left", padx=10, pady=6)

    def _build_table(self):
        container = ttk.Frame(self)
        container.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        columns = (
            "invoice", "client", "payment_date", "amount", "commission",
            "bank", "reference", "status", "source"
        )
        self.table = ttk.Treeview(container, columns=columns, show="headings")

        headers = {
            "invoice": "Factura",
            "client": "Cliente",
            "payment_date": "Fecha de pago",
            "amount": "Monto pagado",
            "commission": "Comision",
            "bank": "Banco",
            "reference": "Referencia",
            "status": "Estado",
            "source": "Origen",
        }

        widths = {
            "invoice": 130,
            "client": 260,
            "payment_date": 145,
            "amount": 120,
            "commission": 100,
            "bank": 130,
            "reference": 160,
            "status": 130,
            "source": 120,
        }

        for col in columns:
            self.table.heading(col, text=headers[col])
            self.table.column(col, width=widths[col], anchor="center")

        vsb = ttk.Scrollbar(container, orient="vertical", command=self.table.yview)
        hsb = ttk.Scrollbar(container, orient="horizontal", command=self.table.xview)
        self.table.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.table.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        container.rowconfigure(0, weight=1)
        container.columnconfigure(0, weight=1)

    def _search(self):
        year = self.cmb_year.get().strip()
        month = self.cmb_month.get().strip()
        date_from = self._period_start(self.cmb_from.get())
        date_to = self._period_end(self.cmb_to.get())
        cliente_label = self.cmb_cliente.get().strip()
        cliente = self.clientes_map.get(cliente_label, cliente_label)

        month_num = None
        if month:
            month_num = int(month.split(" - ", 1)[0])

        if date_from and date_to and date_from > date_to:
            messagebox.showwarning("Fechas", "El periodo Desde no puede ser mayor que Hasta.")
            return

        resp = get_paid_invoices_report_api(
            year=int(year) if year else None,
            month=month_num,
            date_from=date_from,
            date_to=date_to,
            cliente=cliente or None,
            page=1,
            page_size=1000,
        )

        if resp.get("error"):
            messagebox.showerror("Reporte", f"No se pudo cargar el reporte:\n{resp['error']}")
            return

        self.rows = resp.get("data", [])
        self.summary = resp.get("summary", {})
        self._render_rows()
        self._render_summary(resp.get("total", 0))

    def _render_rows(self):
        self.table.delete(*self.table.get_children())
        for index, row in enumerate(self.rows):
            self.table.insert(
                "",
                "end",
                iid=str(index),
                values=(
                    row.get("numero_documento") or "",
                    row.get("nombre_cliente") or "",
                    to_long_english_date(row.get("fecha_pago")),
                    self._money(row.get("monto_pagado")),
                    self._money(row.get("comision")),
                    row.get("banco") or "",
                    row.get("referencia") or "",
                    row.get("estado_factura") or "",
                    row.get("source") or "",
                )
            )

    def _render_summary(self, total):
        self.lbl_total.config(text=f"Pagos: {total}")
        self.lbl_invoices.config(text=f"Facturas: {self.summary.get('total_facturas', 0)}")
        self.lbl_clients.config(text=f"Clientes: {self.summary.get('total_clientes', 0)}")
        self.lbl_paid.config(text=f"Total pagado: {self._money(self.summary.get('total_pagado'))}")
        self.lbl_commission.config(text=f"Comision: {self._money(self.summary.get('total_comision'))}")

    def _clear_filters(self):
        self.cmb_year.set(str(datetime.now().year))
        self.cmb_month.set("")
        self.cmb_from.set("")
        self.cmb_to.set("")
        self.cmb_cliente.set("")
        self.table.delete(*self.table.get_children())
        self.rows = []
        self.summary = {}
        self._render_summary(0)

    def _export(self):
        if not self.rows:
            messagebox.showwarning("Exportar", "Primero cargue datos con Buscar.")
            return

        export_type = self.cmb_export.get()
        if export_type == "PDF":
            self._export_pdf()
        else:
            self._export_excel()

    def _export_excel(self):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except Exception as exc:
            messagebox.showerror("Excel", f"No se pudo cargar openpyxl:\n{exc}")
            return

        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            title="Guardar reporte de facturas pagadas"
        )
        if not path:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Facturas Pagadas"
        ws.append(["Reporte de Facturas Pagadas"])
        ws.append([])
        ws.append(["Factura", "Cliente", "Fecha pago", "Monto pagado", "Comision", "Banco", "Referencia", "Estado", "Origen"])

        for cell in ws[3]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E79")

        for row in self.rows:
            ws.append([
                row.get("numero_documento") or "",
                row.get("nombre_cliente") or "",
                to_long_english_date(row.get("fecha_pago")),
                float(row.get("monto_pagado") or 0),
                float(row.get("comision") or 0),
                row.get("banco") or "",
                row.get("referencia") or "",
                row.get("estado_factura") or "",
                row.get("source") or "",
            ])

        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

        wb.save(path)
        messagebox.showinfo("Excel", "Reporte exportado correctamente.")

    def _export_pdf(self):
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, landscape
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        except Exception as exc:
            messagebox.showerror("PDF", f"No se pudo cargar reportlab:\n{exc}")
            return

        path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF", "*.pdf")],
            title="Guardar reporte de facturas pagadas"
        )
        if not path:
            return

        doc = SimpleDocTemplate(path, pagesize=landscape(letter), rightMargin=24, leftMargin=24, topMargin=24, bottomMargin=24)
        styles = getSampleStyleSheet()
        story = [
            Paragraph("Reporte de Facturas Pagadas", styles["Title"]),
            Paragraph(
                f"Pagos: {len(self.rows)} | Facturas: {self.summary.get('total_facturas', 0)} | "
                f"Clientes: {self.summary.get('total_clientes', 0)} | Total pagado: {self._money(self.summary.get('total_pagado'))}",
                styles["Normal"]
            ),
            Spacer(1, 12),
        ]

        data = [["Factura", "Cliente", "Fecha pago", "Monto", "Banco", "Referencia", "Estado"]]
        for row in self.rows:
            data.append([
                str(row.get("numero_documento") or ""),
                str(row.get("nombre_cliente") or "")[:32],
                to_long_english_date(row.get("fecha_pago")),
                self._money(row.get("monto_pagado")),
                str(row.get("banco") or "")[:18],
                str(row.get("referencia") or "")[:22],
                str(row.get("estado_factura") or "")[:18],
            ])

        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F4F7FA")]),
        ]))
        story.append(table)

        doc.build(story)
        messagebox.showinfo("PDF", "Reporte exportado correctamente.")

    def _period_start(self, value):
        value = (value or "").strip()
        if not value:
            return None
        return value.split(" - ", 1)[0] + "-01"

    def _period_end(self, value):
        value = (value or "").strip()
        if not value:
            return None
        year_month = value.split(" - ", 1)[0]
        year, month = [int(part) for part in year_month.split("-")]
        last_day = calendar.monthrange(year, month)[1]
        return f"{year}-{month:02d}-{last_day:02d}"

    def _money(self, value):
        try:
            return f"{float(value or 0):,.2f}"
        except Exception:
            return "0.00"

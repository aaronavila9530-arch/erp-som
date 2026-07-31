import csv
import os
import tkinter as tk
from datetime import date
from tkinter import filedialog, messagebox, simpledialog, ttk

from openpyxl import load_workbook

from api_client import (
    auto_match_bank_statement_api,
    close_bank_reconciliation_statement_api,
    get_accounting_bank_accounts_api,
    get_bank_reconciliation_statement_lines_api,
    get_bank_reconciliation_statements_api,
    import_bank_statement_api,
    mark_bank_statement_line_fee_api,
)


class PopupProfessionalBankReconciliation(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Conciliacion bancaria profesional")
        self.geometry("1320x760")
        self.minsize(1100, 620)
        self.configure(bg="#f3f2ee")
        self.accounts = []
        self.account_map = {}
        self.selected_statement_id = None
        self._build_ui()
        self._load_accounts()
        self._load_statements()

    def _build_ui(self):
        top = ttk.LabelFrame(self, text="Extracto bancario", padding=8)
        top.pack(fill="x", padx=10, pady=8)

        self.bank_var = tk.StringVar()
        self.account_var = tk.StringVar()
        self.currency_var = tk.StringVar(value="CRC")
        self.period_var = tk.StringVar(value=date.today().strftime("%Y-%m"))
        self.status_var = tk.StringVar(value="")

        ttk.Label(top, text="Banco").grid(row=0, column=0, sticky="w", padx=4, pady=4)
        ttk.Entry(top, textvariable=self.bank_var, width=24).grid(row=0, column=1, sticky="w", padx=4, pady=4)
        ttk.Label(top, text="Cuenta contable").grid(row=0, column=2, sticky="w", padx=4, pady=4)
        self.cmb_account = ttk.Combobox(top, textvariable=self.account_var, width=42, state="readonly")
        self.cmb_account.grid(row=0, column=3, sticky="w", padx=4, pady=4)
        ttk.Label(top, text="Moneda").grid(row=0, column=4, sticky="w", padx=4, pady=4)
        ttk.Combobox(top, textvariable=self.currency_var, values=("CRC", "USD"), width=8, state="readonly").grid(row=0, column=5, sticky="w", padx=4, pady=4)
        ttk.Label(top, text="Periodo").grid(row=0, column=6, sticky="w", padx=4, pady=4)
        ttk.Entry(top, textvariable=self.period_var, width=10).grid(row=0, column=7, sticky="w", padx=4, pady=4)

        ttk.Button(top, text="Importar CSV/XLSX", command=self._import_statement).grid(row=1, column=0, padx=4, pady=4)
        ttk.Button(top, text="Buscar extractos", command=self._load_statements).grid(row=1, column=1, padx=4, pady=4)
        ttk.Button(top, text="Matching automatico", command=self._auto_match).grid(row=1, column=2, padx=4, pady=4)
        ttk.Button(top, text="Cargo bancario", command=self._mark_fee).grid(row=1, column=3, sticky="w", padx=4, pady=4)
        ttk.Button(top, text="Cerrar conciliacion", command=self._close_statement).grid(row=1, column=4, padx=4, pady=4)
        ttk.Button(top, text="Cerrar", command=self.destroy).grid(row=1, column=7, padx=4, pady=4)

        pane = ttk.Panedwindow(self, orient="vertical")
        pane.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        frame_statements = ttk.LabelFrame(pane, text="Extractos")
        frame_lines = ttk.LabelFrame(pane, text="Lineas del extracto / partidas abiertas / diferencias")
        pane.add(frame_statements, weight=1)
        pane.add(frame_lines, weight=3)

        self.statement_tree = self._tree(frame_statements, (
            "id", "bank", "account", "currency", "period", "status", "lines", "open", "total", "matched", "open_total"
        ), (
            "ID", "Banco", "Cuenta", "Moneda", "Periodo", "Status", "Lineas", "Abiertas", "Total", "Matcheado", "Abierto"
        ))
        self.statement_tree.bind("<<TreeviewSelect>>", lambda _e: self._load_lines())

        self.line_tree = self._tree(frame_lines, (
            "id", "date", "reference", "description", "debit", "credit", "amount",
            "status", "source", "matched", "difference"
        ), (
            "ID", "Fecha", "Referencia", "Descripcion", "Debito", "Credito", "Monto",
            "Status", "Fuente", "Match ID", "Diferencia"
        ), widths={"description": 300, "reference": 170})
        self.line_tree.tag_configure("OPEN", foreground="#b42318")
        self.line_tree.tag_configure("AUTO_MATCHED", foreground="#167c3a")
        self.line_tree.tag_configure("BANK_FEE", foreground="#946200")
        self.line_tree.tag_configure("DIFFERENCE", foreground="#b42318")

    def _tree(self, parent, cols, headers, widths=None):
        frame = ttk.Frame(parent)
        frame.pack(fill="both", expand=True, padx=5, pady=5)
        tree = ttk.Treeview(frame, columns=cols, show="headings")
        yscroll = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        xscroll = ttk.Scrollbar(frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)
        for col, header in zip(cols, headers):
            tree.heading(col, text=header)
            tree.column(col, width=(widths or {}).get(col, 120), stretch=False)
        tree.grid(row=0, column=0, sticky="nsew")
        yscroll.grid(row=0, column=1, sticky="ns")
        xscroll.grid(row=1, column=0, sticky="ew")
        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)
        return tree

    def _load_accounts(self):
        self.accounts = get_accounting_bank_accounts_api()
        labels = []
        for acc in self.accounts:
            label = f"{acc.get('account_code')} - {acc.get('account_name')}"
            self.account_map[label] = acc
            labels.append(label)
        self.cmb_account["values"] = labels
        if labels and not self.account_var.get():
            self.account_var.set(labels[0])

    def _load_statements(self):
        try:
            account = self.account_map.get(self.account_var.get()) or {}
            rows = get_bank_reconciliation_statements_api(
                bank_account_code=account.get("account_code") or None,
                currency_code=self.currency_var.get() or None,
                period=self.period_var.get() or None,
            )
            self.statement_tree.delete(*self.statement_tree.get_children())
            for row in rows:
                self.statement_tree.insert("", "end", iid=str(row["id"]), values=(
                    row["id"], row["bank_name"], row.get("bank_account_code") or "",
                    row["currency_code"], row.get("statement_period") or "", row["status"],
                    row.get("line_count") or 0, row.get("open_count") or 0,
                    self._amount(row.get("statement_total")),
                    self._amount(row.get("matched_total")),
                    self._amount(row.get("open_total")),
                ))
        except Exception as exc:
            messagebox.showerror("Conciliacion bancaria", str(exc))

    def _load_lines(self):
        selected = self.statement_tree.selection()
        if not selected:
            return
        self.selected_statement_id = int(selected[0])
        rows = get_bank_reconciliation_statement_lines_api(self.selected_statement_id)
        self.line_tree.delete(*self.line_tree.get_children())
        for row in rows:
            status = row.get("match_status") or ""
            self.line_tree.insert("", "end", iid=str(row["id"]), values=(
                row["id"], row.get("line_date") or "", row.get("reference") or "",
                row.get("description") or "", self._amount(row.get("debit")),
                self._amount(row.get("credit")), self._amount(row.get("amount")),
                status, row.get("matched_source") or "", row.get("matched_id") or "",
                self._amount(row.get("difference")),
            ), tags=(status,))

    def _import_statement(self):
        path = filedialog.askopenfilename(
            title="Importar extracto bancario",
            filetypes=(("Excel/CSV", "*.xlsx *.csv"), ("Excel", "*.xlsx"), ("CSV", "*.csv"), ("Todos", "*.*")),
        )
        if not path:
            return
        try:
            rows = self._read_statement_file(path)
            account = self.account_map.get(self.account_var.get()) or {}
            payload = {
                "bank_name": self.bank_var.get() or account.get("account_name") or "Banco",
                "bank_account_code": account.get("account_code"),
                "bank_account_name": account.get("account_name"),
                "currency_code": self.currency_var.get(),
                "statement_period": self.period_var.get(),
                "statement_date": date.today().isoformat(),
                "source_filename": os.path.basename(path),
                "rows": rows,
            }
            result = import_bank_statement_api(payload)
            self._load_statements()
            messagebox.showinfo("Importacion", f"Extracto importado.\nLineas: {result.get('inserted')}\nOmitidas: {result.get('skipped')}")
        except Exception as exc:
            messagebox.showerror("Importacion", str(exc))

    def _read_statement_file(self, path):
        if path.lower().endswith(".csv"):
            with open(path, newline="", encoding="utf-8-sig") as fh:
                raw = list(csv.DictReader(fh))
        else:
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            values = list(ws.iter_rows(values_only=True))
            headers = [str(v or "").strip() for v in values[0]]
            raw = [dict(zip(headers, row)) for row in values[1:] if any(row)]
        return [self._normalize_row(row) for row in raw if self._normalize_row(row)]

    def _normalize_row(self, row):
        lowered = {str(k).strip().lower(): v for k, v in row.items()}
        def pick(*names):
            for name in names:
                if name in lowered and lowered[name] not in (None, ""):
                    return lowered[name]
            return None
        line_date = pick("fecha", "date", "line_date", "fecha pago", "fecha_pago")
        if not line_date:
            return None
        debit = pick("debito", "debit", "retiro", "withdrawal") or 0
        credit = pick("credito", "credit", "deposito", "deposit") or 0
        amount = pick("monto", "amount", "importe") or 0
        return {
            "line_date": str(line_date)[:10],
            "description": pick("descripcion", "description", "detalle", "concepto"),
            "reference": pick("referencia", "reference", "comprobante", "numero", "documento"),
            "debit": self._clean_number(debit),
            "credit": self._clean_number(credit),
            "amount": self._clean_number(amount),
            "currency_code": self.currency_var.get(),
        }

    def _auto_match(self):
        if not self.selected_statement_id:
            messagebox.showwarning("Matching", "Seleccione un extracto.")
            return
        tolerance = simpledialog.askfloat("Matching automatico", "Tolerancia:", parent=self, initialvalue=1.0, minvalue=0)
        if tolerance is None:
            return
        result = auto_match_bank_statement_api(self.selected_statement_id, tolerance)
        self._load_statements()
        self._load_lines()
        messagebox.showinfo("Matching", f"Matcheadas: {result.get('matched')}\nDiferencias: {result.get('differences')}")

    def _mark_fee(self):
        selected = self.line_tree.selection()
        if not selected:
            messagebox.showwarning("Cargo bancario", "Seleccione una linea.")
            return
        note = simpledialog.askstring("Cargo bancario", "Nota:", parent=self)
        mark_bank_statement_line_fee_api(int(selected[0]), note)
        self._load_lines()

    def _close_statement(self):
        if not self.selected_statement_id:
            messagebox.showwarning("Cierre", "Seleccione un extracto.")
            return
        note = simpledialog.askstring("Cierre", "Nota de cierre:", parent=self) or ""
        force = messagebox.askyesno("Cierre", "Si quedan partidas abiertas, desea forzar cierre documentado?")
        try:
            close_bank_reconciliation_statement_api(self.selected_statement_id, note=note, force_close=force)
            self._load_statements()
            self._load_lines()
            messagebox.showinfo("Cierre", "Conciliacion cerrada.")
        except Exception as exc:
            messagebox.showerror("Cierre", str(exc))

    @staticmethod
    def _clean_number(value):
        text = str(value or "0").replace(",", "").replace(" ", "")
        try:
            return float(text)
        except Exception:
            return 0

    @staticmethod
    def _amount(value):
        try:
            return f"{float(value or 0):,.2f}"
        except Exception:
            return "0.00"

import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
from datetime import date
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from api_client import (
    get_itp_biweekly_obligations_preview_api,
    post_itp_biweekly_obligations_apply_api,
)


CATEGORIES = [
    "Planilla",
    "CCSS",
    "IVA",
    "Surveyors",
    "Viaticos",
    "Tarjetas de credito",
    "Alquiler",
    "Internet",
    "Telefonia",
    "Otros",
]

BANK_ACCOUNT_OPTIONS = [
    "1.1.02.02.01",
    "1.1.02.02.02",
    "1.1.02.04.01",
]


class PopupObligacionesQuincenales(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Obligaciones quincenales")
        self.geometry("1220x760")
        self.minsize(1020, 620)
        self.rows = []
        today = date.today()
        self.period_var = tk.StringVar(value=f"{today.year:04d}-{today.month:02d}")
        self.fortnight_var = tk.IntVar(value=1 if today.day <= 15 else 2)
        self.total_var = tk.StringVar(value="CRC 0.00")
        self.total_usd_var = tk.StringVar(value="USD 0.00")
        self.count_var = tk.StringVar(value="0 lineas")
        self._build_ui()
        self._load_preview()
        self.transient(parent)
        self.grab_set()

    def _build_ui(self):
        self.columnconfigure(0, weight=1)
        self.rowconfigure(2, weight=1)

        header = ttk.Frame(self, padding=(10, 8))
        header.grid(row=0, column=0, sticky="ew")
        header.columnconfigure(8, weight=1)
        ttk.Label(header, text="Obligaciones quincenales", font=("Segoe UI", 15, "bold")).grid(row=0, column=0, columnspan=2, sticky="w")
        ttk.Label(header, text="Periodo").grid(row=1, column=0, sticky="w", pady=(8, 0))
        ttk.Entry(header, textvariable=self.period_var, width=10).grid(row=1, column=1, sticky="w", padx=(4, 14), pady=(8, 0))
        ttk.Label(header, text="Quincena").grid(row=1, column=2, sticky="w", pady=(8, 0))
        ttk.Combobox(header, textvariable=self.fortnight_var, values=[1, 2], state="readonly", width=5).grid(row=1, column=3, sticky="w", padx=(4, 14), pady=(8, 0))
        ttk.Button(header, text="Generar automatico", command=self._load_preview).grid(row=1, column=4, padx=4, pady=(8, 0))
        ttk.Button(header, text="Exportar Excel", command=self._export_excel).grid(row=1, column=5, padx=4, pady=(8, 0))
        ttk.Button(header, text="Aplicar pagos y crear asientos", command=self._save_and_post).grid(row=1, column=6, padx=4, pady=(8, 0))
        ttk.Button(header, text="Cerrar", command=self.destroy).grid(row=1, column=7, padx=4, pady=(8, 0))
        ttk.Label(header, textvariable=self.total_var, font=("Segoe UI", 12, "bold")).grid(row=0, column=8, sticky="e")
        ttk.Label(header, textvariable=self.total_usd_var, font=("Segoe UI", 11, "bold")).grid(row=1, column=8, sticky="e", padx=(0, 90))
        ttk.Label(header, textvariable=self.count_var).grid(row=1, column=8, sticky="e")

        tools = ttk.LabelFrame(self, text="Agregar / ajustar lineas")
        tools.grid(row=1, column=0, sticky="ew", padx=10, pady=(0, 8))
        for idx, category in enumerate(["Planilla", "CCSS", "Surveyors", "Viaticos", "Telefonia", "Otros"]):
            ttk.Button(tools, text=f"+ {category}", command=lambda c=category: self._add_line(c)).grid(row=0, column=idx, padx=4, pady=6)
        ttk.Button(tools, text="Editar linea", command=self._edit_selected).grid(row=0, column=7, padx=(18, 4), pady=6)
        ttk.Button(tools, text="Quitar linea", command=self._delete_selected).grid(row=0, column=8, padx=4, pady=6)
        ttk.Label(
            tools,
            text="Obligatorio antes de aplicar: comprobante bancario y cuenta contable banco en cada linea con monto.",
            foreground="#7f1d1d",
            font=("Segoe UI", 9, "bold"),
        ).grid(row=1, column=0, columnspan=9, sticky="w", padx=6, pady=(0, 6))

        pane = ttk.Panedwindow(self, orient="vertical")
        pane.grid(row=2, column=0, sticky="nsew", padx=10, pady=(0, 10))

        detail = ttk.Frame(pane)
        detail.columnconfigure(0, weight=1)
        detail.rowconfigure(0, weight=1)
        pane.add(detail, weight=4)

        columns = ("category", "name", "amount", "currency", "bank_account", "bank_accounting_code", "bank_voucher", "due_date", "obligation_id", "reference", "balance", "source", "notes")
        self.tree = ttk.Treeview(detail, columns=columns, show="headings", height=17)
        labels = {
            "category": "Rubro",
            "name": "Nombre / beneficiario",
            "amount": "Monto",
            "currency": "Moneda",
            "bank_account": "Cuenta destino / IBAN",
            "bank_accounting_code": "Cuenta contable banco",
            "bank_voucher": "Comprobante bancario",
            "due_date": "Fecha pago",
            "obligation_id": "ITP ID",
            "reference": "Referencia",
            "balance": "Saldo ITP",
            "source": "Fuente",
            "notes": "Notas",
        }
        widths = {
            "category": 140,
            "name": 260,
            "amount": 120,
            "currency": 70,
            "bank_account": 220,
            "bank_accounting_code": 150,
            "bank_voucher": 150,
            "due_date": 95,
            "obligation_id": 70,
            "reference": 150,
            "balance": 110,
            "source": 95,
            "notes": 290,
        }
        for col in columns:
            self.tree.heading(col, text=labels[col])
            self.tree.column(col, width=widths[col], anchor="e" if col == "amount" else "w")
        self.tree.tag_configure("auto", background="#eef7f2")
        self.tree.tag_configure("manual", background="#fff7ed")
        self.tree.tag_configure("review", background="#fee2e2")
        self.tree.bind("<Double-1>", lambda _e: self._edit_selected())
        yscroll = ttk.Scrollbar(detail, orient="vertical", command=self.tree.yview)
        xscroll = ttk.Scrollbar(detail, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        yscroll.grid(row=0, column=1, sticky="ns")
        xscroll.grid(row=1, column=0, sticky="ew")

        summary = ttk.Frame(pane)
        summary.columnconfigure(0, weight=1)
        summary.columnconfigure(1, weight=1)
        summary.rowconfigure(0, weight=1)
        pane.add(summary, weight=1)

        self.tree_category = self._summary_tree(summary, "Resumen por rubro")
        self.tree_category.grid(row=0, column=0, sticky="nsew", padx=(0, 5))
        self.tree_bank = self._summary_tree(summary, "Resumen por cuenta / destino")
        self.tree_bank.grid(row=0, column=1, sticky="nsew", padx=(5, 0))

    def _summary_tree(self, parent, title):
        frame = ttk.LabelFrame(parent, text=title)
        frame.columnconfigure(0, weight=1)
        frame.rowconfigure(0, weight=1)
        tree = ttk.Treeview(frame, columns=("key", "currency", "amount"), show="headings", height=7)
        for col, label, width in (("key", "Concepto", 280), ("currency", "Moneda", 70), ("amount", "Total", 130)):
            tree.heading(col, text=label)
            tree.column(col, width=width, anchor="e" if col == "amount" else "w")
        tree.grid(row=0, column=0, sticky="nsew")
        return frame

    def _summary_tree_widget(self, frame):
        for child in frame.winfo_children():
            if isinstance(child, ttk.Treeview):
                return child
        return None

    def _load_preview(self):
        try:
            data = get_itp_biweekly_obligations_preview_api(self.period_var.get().strip(), int(self.fortnight_var.get() or 1))
            self.rows = data.get("rows") or []
            self._render()
        except Exception as exc:
            messagebox.showerror("Obligaciones quincenales", f"No se pudo generar preview:\n{exc}", parent=self)

    def _render(self):
        self.tree.delete(*self.tree.get_children())
        for idx, row in enumerate(self.rows):
            source = str(row.get("source") or "").upper()
            tag = "review" if source == "REVISION" else ("manual" if source == "MANUAL" else "auto")
            self.tree.insert("", "end", iid=str(idx), tags=(tag,), values=(
                row.get("category") or "",
                row.get("name") or "",
                self._fmt(row.get("amount")),
                row.get("currency") or "CRC",
                row.get("bank_account") or "",
                row.get("bank_accounting_code") or "",
                row.get("bank_voucher") or "",
                row.get("due_date") or "",
                row.get("obligation_id") or "",
                row.get("reference") or "",
                self._fmt(row.get("balance")) if row.get("obligation_id") else "",
                row.get("source") or "",
                row.get("notes") or "",
            ))
        self._render_summaries()

    def _render_summaries(self):
        by_category = defaultdict(float)
        by_bank = defaultdict(float)
        total_crc = 0.0
        total_usd = 0.0
        for row in self.rows:
            amount = self._money(row.get("amount"))
            currency = row.get("currency") or "CRC"
            by_category[(row.get("category") or "Otros", currency)] += amount
            by_bank[(row.get("bank_account") or "Sin cuenta", currency)] += amount
            if currency == "CRC":
                total_crc += amount
            elif currency == "USD":
                total_usd += amount
        for frame, data in ((self.tree_category, by_category), (self.tree_bank, by_bank)):
            tree = self._summary_tree_widget(frame)
            tree.delete(*tree.get_children())
            for (key, currency), amount in sorted(data.items()):
                tree.insert("", "end", values=(key, currency, self._fmt(amount)))
        self.total_var.set(f"CRC {total_crc:,.2f}")
        self.total_usd_var.set(f"USD {total_usd:,.2f}")
        self.count_var.set(f"{len(self.rows):,} lineas")

    def _add_line(self, category):
        line = self._line_dialog({"category": category, "currency": "CRC", "source": "MANUAL"})
        if line:
            self.rows.append(line)
            self._render()

    def _edit_selected(self):
        selected = self.tree.focus()
        if not selected:
            messagebox.showwarning("Obligaciones quincenales", "Selecciona una linea.", parent=self)
            return
        idx = int(selected)
        line = self._line_dialog(dict(self.rows[idx]))
        if line:
            self.rows[idx] = line
            self._render()

    def _delete_selected(self):
        selected = self.tree.focus()
        if not selected:
            return
        idx = int(selected)
        del self.rows[idx]
        self._render()

    def _line_dialog(self, initial):
        win = tk.Toplevel(self)
        win.title("Linea de obligacion")
        win.geometry("620x430")
        win.transient(self)
        win.grab_set()
        vars_ = {
            "category": tk.StringVar(value=initial.get("category") or "Otros"),
            "name": tk.StringVar(value=initial.get("name") or ""),
            "amount": tk.StringVar(value=str(initial.get("amount") or "")),
            "currency": tk.StringVar(value=initial.get("currency") or "CRC"),
            "bank_account": tk.StringVar(value=initial.get("bank_account") or ""),
            "bank_accounting_code": tk.StringVar(value=initial.get("bank_accounting_code") or "1.1.02.02.01"),
            "bank_voucher": tk.StringVar(value=initial.get("bank_voucher") or ""),
            "due_date": tk.StringVar(value=initial.get("due_date") or ""),
            "notes": tk.StringVar(value=initial.get("notes") or ""),
            "reference": tk.StringVar(value=initial.get("reference") or ""),
            "balance": tk.StringVar(value=str(initial.get("balance") or initial.get("amount") or "")),
        }
        result = {}
        labels = [
            ("Rubro", "category"),
            ("Nombre / beneficiario", "name"),
            ("Monto", "amount"),
            ("Moneda", "currency"),
            ("Cuenta destino / IBAN", "bank_account"),
            ("Cuenta contable banco", "bank_accounting_code"),
            ("Comprobante bancario", "bank_voucher"),
            ("Fecha pago", "due_date"),
            ("Referencia ITP", "reference"),
            ("Saldo ITP", "balance"),
            ("Notas", "notes"),
        ]
        for idx, (label, key) in enumerate(labels):
            ttk.Label(win, text=label).grid(row=idx, column=0, sticky="w", padx=12, pady=6)
            if key == "category":
                widget = ttk.Combobox(win, textvariable=vars_[key], values=CATEGORIES, state="readonly")
            elif key == "currency":
                widget = ttk.Combobox(win, textvariable=vars_[key], values=["CRC", "USD"], state="readonly")
            elif key == "bank_accounting_code":
                widget = ttk.Combobox(win, textvariable=vars_[key], values=BANK_ACCOUNT_OPTIONS)
            else:
                widget = ttk.Entry(win, textvariable=vars_[key])
            widget.grid(row=idx, column=1, sticky="ew", padx=12, pady=6)
        win.columnconfigure(1, weight=1)

        def ok():
            try:
                amount = self._money(vars_["amount"].get())
            except Exception:
                messagebox.showwarning("Monto", "Monto invalido.", parent=win)
                return
            result.update({
                "category": vars_["category"].get(),
                "name": vars_["name"].get().strip(),
                "amount": amount,
                "currency": vars_["currency"].get(),
                "bank_account": vars_["bank_account"].get().strip(),
                "bank_accounting_code": vars_["bank_accounting_code"].get().strip(),
                "bank_accounting_name": initial.get("bank_accounting_name") or "",
                "bank_voucher": vars_["bank_voucher"].get().strip(),
                "due_date": vars_["due_date"].get().strip(),
                "source": initial.get("source") or "MANUAL",
                "notes": vars_["notes"].get().strip(),
                "obligation_id": initial.get("obligation_id"),
                "reference": vars_["reference"].get().strip(),
                "balance": self._money(vars_["balance"].get()),
            })
            win.destroy()

        footer = ttk.Frame(win)
        footer.grid(row=len(labels), column=0, columnspan=2, sticky="e", padx=12, pady=12)
        ttk.Button(footer, text="Guardar", command=ok).pack(side="left", padx=4)
        ttk.Button(footer, text="Cancelar", command=win.destroy).pack(side="left", padx=4)
        self.wait_window(win)
        return result or None

    def _export_excel(self):
        if not self.rows:
            messagebox.showwarning("Exportar", "No hay lineas para exportar.", parent=self)
            return
        path = filedialog.asksaveasfilename(
            parent=self,
            defaultextension=".xlsx",
            initialfile=f"obligaciones_quincenales_{self.period_var.get()}_Q{self.fortnight_var.get()}.xlsx",
            filetypes=[("Excel", "*.xlsx")],
        )
        if not path:
            return
        wb = Workbook()
        ws = wb.active
        ws.title = "Obligaciones"
        ws.append(["Periodo", self.period_var.get(), "Quincena", self.fortnight_var.get()])
        ws.append([])
        headers = ["Rubro", "Nombre / beneficiario", "Monto a pagar", "Moneda", "Cuenta destino / IBAN", "Cuenta contable banco", "Comprobante bancario", "Fecha pago", "ITP ID", "Referencia", "Saldo ITP", "Pago parcial", "Fuente", "Notas"]
        ws.append(headers)
        for row in self.rows:
            amount = self._money(row.get("amount"))
            balance = self._money(row.get("balance"))
            ws.append([
                row.get("category"),
                row.get("name"),
                amount,
                row.get("currency"),
                row.get("bank_account"),
                row.get("bank_accounting_code"),
                row.get("bank_voucher"),
                row.get("due_date"),
                row.get("obligation_id") or "",
                row.get("reference") or "",
                balance if row.get("obligation_id") else "",
                "SI" if row.get("obligation_id") and amount < balance else "",
                row.get("source"),
                row.get("notes"),
            ])
        ws2 = wb.create_sheet("Resumen")
        ws2.append(["Resumen por rubro"])
        ws2.append(["Rubro", "Moneda", "Total"])
        for (key, currency), amount in sorted(self._summary_data("category").items()):
            ws2.append([key, currency, amount])
        ws2.append([])
        ws2.append(["Resumen por cuenta / destino"])
        ws2.append(["Cuenta / destino", "Moneda", "Total"])
        for (key, currency), amount in sorted(self._summary_data("bank_account").items()):
            ws2.append([key, currency, amount])
        ws3 = wb.create_sheet("Aplicacion ITP")
        ws3.append(["Lineas vinculadas a obligaciones pendientes"])
        ws3.append(["ITP ID", "Referencia", "Beneficiario", "Saldo ITP", "Monto a pagar", "Moneda", "Pago parcial", "Fecha pago", "Cuenta destino", "Cuenta contable banco", "Comprobante"])
        for row in self.rows:
            if not row.get("obligation_id"):
                continue
            amount = self._money(row.get("amount"))
            balance = self._money(row.get("balance"))
            ws3.append([
                row.get("obligation_id"),
                row.get("reference"),
                row.get("name"),
                balance,
                amount,
                row.get("currency"),
                "SI" if amount < balance else "NO",
                row.get("due_date"),
                row.get("bank_account"),
                row.get("bank_accounting_code"),
                row.get("bank_voucher"),
            ])
        self._style_workbook(wb)
        wb.save(path)
        messagebox.showinfo("Exportar", f"Excel generado correctamente:\n{path}", parent=self)

    def _missing_required(self):
        missing = []
        for idx, row in enumerate(self.rows, start=1):
            if self._money(row.get("amount")) <= 0:
                continue
            if not str(row.get("bank_accounting_code") or "").strip():
                missing.append(f"Linea {idx}: falta cuenta contable banco")
            if not str(row.get("bank_voucher") or "").strip():
                missing.append(f"Linea {idx}: falta comprobante bancario")
        return missing

    def _save_and_post(self):
        missing = self._missing_required()
        if missing:
            messagebox.showwarning(
                "Aplicar pagos y crear asientos",
                "No se puede aplicar ni contabilizar hasta completar estos campos obligatorios:\n"
                + "\n".join(missing[:12]),
                parent=self,
            )
            return
        total_crc = sum(self._money(row.get("amount")) for row in self.rows if (row.get("currency") or "CRC") == "CRC")
        ok = messagebox.askyesno(
            "Aplicar pagos y crear asientos",
            f"Se guardaran las lineas, se aplicaran pagos ITP vinculados y se generaran asientos contables de pago.\n\nTotal CRC visible: {total_crc:,.2f}\n\nContinuar?",
            parent=self,
        )
        if not ok:
            return
        result = post_itp_biweekly_obligations_apply_api({
            "period": self.period_var.get().strip(),
            "fortnight": int(self.fortnight_var.get() or 1),
            "rows": self.rows,
        })
        if result.get("status") == "error":
            messagebox.showerror("Aplicar pagos y crear asientos", result.get("error") or "No se pudo guardar.", parent=self)
        else:
            messagebox.showinfo(
                "Aplicar pagos y crear asientos",
                f"Proceso aplicado.\nBatch: {result.get('batch_id')}\nLineas guardadas: {result.get('saved')}\nAsientos: {result.get('posted')}\nPagos ITP: {result.get('applied')}",
                parent=self,
            )
        self._load_preview()

    def _summary_data(self, key):
        data = defaultdict(float)
        for row in self.rows:
            data[(row.get(key) or ("Sin cuenta" if key == "bank_account" else "Otros"), row.get("currency") or "CRC")] += self._money(row.get("amount"))
        return data

    def _style_workbook(self, wb):
        header_fill = PatternFill("solid", fgColor="0F4C81")
        header_font = Font(color="FFFFFF", bold=True)
        thin = Side(style="thin", color="D0D5DD")
        for ws in wb.worksheets:
            ws.freeze_panes = "A4" if ws.title == "Obligaciones" else "A3"
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                    cell.alignment = Alignment(vertical="center")
                    if isinstance(cell.value, (int, float)) and cell.column >= 3:
                        cell.number_format = '#,##0.00'
            for row in (1, 2, 3):
                for cell in ws[row]:
                    if cell.value:
                        cell.fill = header_fill
                        cell.font = header_font
            if ws.max_row > 3 and ws.max_column:
                ws.auto_filter.ref = ws.dimensions
            for col in ws.columns:
                width = max(len(str(cell.value or "")) for cell in col) + 2
                ws.column_dimensions[col[0].column_letter].width = min(max(width, 12), 42)

    @staticmethod
    def _money(value):
        try:
            return float(str(value or 0).replace(",", ""))
        except Exception:
            return 0.0

    @staticmethod
    def _fmt(value):
        return f"{PopupObligacionesQuincenales._money(value):,.2f}"

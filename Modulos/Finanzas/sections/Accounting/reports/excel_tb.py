from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from tkinter import filedialog
from typing import Any

from Modulos.Finanzas.sections.Accounting.reports.build_tb_from_lines import (
    build_tb_from_lines
)

THIN = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def export_tb_excel_from_lines(rows_or_built: Any):
    """
    Exporta Balance de Comprobación (TB) a Excel

    ✔ Acepta TB ya construido (dict)
    ✔ Acepta accounting_lines (list)
    ✔ Save As interno
    ✔ SIN company_code
    ✔ Style blindado
    """

    # ==================================================
    # NORMALIZACIÓN
    # ==================================================
    if isinstance(rows_or_built, list):
        tb = build_tb_from_lines(rows_or_built)
    elif isinstance(rows_or_built, dict):
        tb = rows_or_built
    else:
        raise ValueError("export_tb_excel_from_lines esperaba dict o list")

    # ==================================================
    # VALIDACIÓN
    # ==================================================
    if "rows" not in tb:
        raise ValueError("Balance de Comprobación inválido")

    # ==================================================
    # SAVE AS
    # ==================================================
    file_path = filedialog.asksaveasfilename(
        title="Guardar Balance de Comprobación",
        defaultextension=".xlsx",
        filetypes=[("Excel", "*.xlsx")]
    )
    if not file_path:
        return None

    # ==================================================
    # WORKBOOK
    # ==================================================
    wb = Workbook()
    ws = wb.active
    ws.title = "Balance de Comprobación"

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18

    row = 1

    # ==================================================
    # ENCABEZADO
    # ==================================================
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row=row, column=1, value="BALANCE DE COMPROBACIÓN")
    ws.cell(row=row, column=1).font = Font(bold=True, size=14)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(
        row=row,
        column=1,
        value=f"Periodo {tb['period_label']}"
    )
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
    row += 2

    # ==================================================
    # HEADER TABLA
    # ==================================================
    headers = [
        "Cuenta",
        "Debe",
        "Haber",
        "Saldo Deudor",
        "Saldo Acreedor"
    ]
    ws.append(headers)

    for col in range(1, 6):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    row += 1

    # ==================================================
    # DATA
    # ==================================================
    for r in tb["rows"]:
        ws.append([
            r["account"],
            r["debit"],
            r["credit"],
            r["saldo_deudor"],
            r["saldo_acreedor"],
        ])

        ws.cell(row=row, column=2).number_format = "#,##0.00"
        ws.cell(row=row, column=3).number_format = "#,##0.00"
        ws.cell(row=row, column=4).number_format = "#,##0.00"
        ws.cell(row=row, column=5).number_format = "#,##0.00"

        ws.cell(row=row, column=2).alignment = Alignment(horizontal="right")
        ws.cell(row=row, column=3).alignment = Alignment(horizontal="right")
        ws.cell(row=row, column=4).alignment = Alignment(horizontal="right")
        ws.cell(row=row, column=5).alignment = Alignment(horizontal="right")

        row += 1

    # ==================================================
    # TOTALES
    # ==================================================
    ws.append([
        "TOTAL",
        tb["total_debit"],
        tb["total_credit"],
        tb["total_saldo_deudor"],
        tb["total_saldo_acreedor"],
    ])

    for col in range(1, 6):
        ws.cell(row=row, column=col).font = Font(bold=True)

    # ==================================================
    # BORDES (BLINDADO)
    # ==================================================
    for r in ws.iter_rows(min_row=1, max_row=row, min_col=1, max_col=5):
        for cell in r:
            if cell.value is not None:
                try:
                    cell.border = BORDER
                except Exception:
                    pass

    wb.save(file_path)
    return file_path


# 🔑 Alias compatible con popup
export_tb_excel = export_tb_excel_from_lines

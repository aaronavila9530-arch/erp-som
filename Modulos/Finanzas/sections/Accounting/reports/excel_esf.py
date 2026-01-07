from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from tkinter import filedialog
from typing import Dict, Any, List

from Modulos.Finanzas.sections.Accounting.reports.build_esf_from_trial_balance import (
    build_esf_from_trial_balance
)


THIN = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def export_esf_excel_from_esf(esf_or_rows: Any):
    """
    Exporta Estado de Situación Financiera a Excel

    ✔ Acepta ESF ya construido (dict)
    ✔ Acepta accounting_lines (list) y construye ESF automáticamente
    ✔ Save As interno
    ✔ Blindado total
    """

    # ==================================================
    # NORMALIZACIÓN DE ENTRADA
    # ==================================================
    if isinstance(esf_or_rows, list):
        esf = build_esf_from_trial_balance(esf_or_rows)
    elif isinstance(esf_or_rows, dict):
        esf = esf_or_rows
    else:
        raise ValueError(
            "export_esf_excel_from_esf esperaba un ESF (dict) "
            "o accounting_lines (list)"
        )

    # ==================================================
    # VALIDACIÓN FUERTE DE ESF
    # ==================================================
    required_keys = [
        "activo_corriente",
        "total_activo_corriente",
        "activo_no_corriente",
        "total_activo_no_corriente",
        "total_activo",
        "pasivo_corriente",
        "total_pasivo_corriente",
        "pasivo_no_corriente",
        "total_pasivo_no_corriente",
        "total_pasivo",
        "patrimonio",
        "total_patrimonio",
        "total_pasivo_patrimonio",
    ]

    for k in required_keys:
        if k not in esf:
            raise ValueError(f"ESF inválido. Falta clave '{k}'")

    # ==================================================
    # NORMALIZADORES
    # ==================================================
    def safe_list(val) -> List[dict]:
        return val if isinstance(val, list) else []

    def safe_amount(val) -> float:
        try:
            return float(val)
        except Exception:
            return 0.0

    activo_corriente = safe_list(esf["activo_corriente"])
    activo_no_corriente = safe_list(esf["activo_no_corriente"])
    pasivo_corriente = safe_list(esf["pasivo_corriente"])
    pasivo_no_corriente = safe_list(esf["pasivo_no_corriente"])
    patrimonio = safe_list(esf["patrimonio"])

    # ==================================================
    # SAVE AS
    # ==================================================
    file_path = filedialog.asksaveasfilename(
        title="Guardar Estado de Situación Financiera",
        defaultextension=".xlsx",
        filetypes=[("Excel", "*.xlsx")]
    )
    if not file_path:
        return None

    # ==================================================
    # WORKBOOK
    # ==================================================
    wb = Workbook()
    ws = wb.active
    ws.title = "Estado Situación Financiera"

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 20

    row = 1

    # ==================================================
    # ENCABEZADO (🔥 CORREGIDO AQUÍ 🔥)
    # ==================================================
    ws.merge_cells(
        start_row=row,
        start_column=1,
        end_row=row,
        end_column=2
    )
    ws.cell(row=row, column=1, value="ESTADO DE SITUACIÓN FINANCIERA")
    ws.cell(row=row, column=1).font = Font(bold=True, size=14)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
    row += 2

    # ==================================================
    # HELPERS
    # ==================================================
    def section(title: str):
        nonlocal row
        ws.cell(row=row, column=1, value=title).font = Font(bold=True)
        row += 1

    def line(label: str, amount: Any, bold: bool = False):
        nonlocal row
        amount = safe_amount(amount)

        ws.cell(row=row, column=1, value=str(label))
        ws.cell(row=row, column=2, value=round(amount, 2))
        ws.cell(row=row, column=2).number_format = "#,##0.00"
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="right")

        if bold:
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=2).font = Font(bold=True)

        row += 1

    def spacer():
        nonlocal row
        row += 1

    # ==================================================
    # ACTIVO
    # ==================================================
    section("ACTIVO")

    section("Activo Corriente")
    for item in activo_corriente:
        line(item.get("label", ""), item.get("amount", 0))

    line("Total Activo Corriente", esf["total_activo_corriente"], bold=True)
    spacer()

    section("Activo No Corriente")
    for item in activo_no_corriente:
        line(item.get("label", ""), item.get("amount", 0))

    line("Total Activo No Corriente", esf["total_activo_no_corriente"], bold=True)
    spacer()

    line("TOTAL ACTIVO", esf["total_activo"], bold=True)
    spacer()
    spacer()

    # ==================================================
    # PASIVO
    # ==================================================
    section("PASIVO")

    section("Pasivo Corriente")
    for item in pasivo_corriente:
        line(item.get("label", ""), item.get("amount", 0))

    line("Total Pasivo Corriente", esf["total_pasivo_corriente"], bold=True)
    spacer()

    section("Pasivo No Corriente")
    for item in pasivo_no_corriente:
        line(item.get("label", ""), item.get("amount", 0))

    line("Total Pasivo No Corriente", esf["total_pasivo_no_corriente"], bold=True)
    spacer()

    line("TOTAL PASIVO", esf["total_pasivo"], bold=True)
    spacer()
    spacer()

    # ==================================================
    # PATRIMONIO
    # ==================================================
    section("PATRIMONIO")

    for item in patrimonio:
        line(item.get("label", ""), item.get("amount", 0))

    line("TOTAL PATRIMONIO", esf["total_patrimonio"], bold=True)
    spacer()

    line(
        "TOTAL PASIVO Y PATRIMONIO",
        esf["total_pasivo_patrimonio"],
        bold=True
    )

    # ==================================================
    # BORDES
    # ==================================================
    for r in ws.iter_rows(min_row=1, max_row=row, min_col=1, max_col=2):
        for cell in r:
            if cell.value is not None:
                cell.border = BORDER

    wb.save(file_path)
    return file_path


# 🔑 Alias para imports existentes
export_esf_excel = export_esf_excel_from_esf

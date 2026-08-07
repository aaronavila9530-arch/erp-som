from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, PatternFill, Side
from tkinter import filedialog
from Modulos.Finanzas.date_utils import to_long_english_date

THIN = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="003A75")
ENTRY_FILL = PatternFill("solid", fgColor="D9EAF7")
ALT_FILL = PatternFill("solid", fgColor="F7FBFF")


def export_diario_excel(rows: list[dict], fiscal_year: int, period: int):
    """
    Exporta Libro Diario DIRECTAMENTE desde accounting_lines.

    Reglas:
    - NO usa company_code / company_name
    - Fiscal year y period vienen del POPUP (no se calculan aquí)
    - created_at es la fecha real por línea
    """

    if not rows:
        raise ValueError("No hay datos para exportar")

    # =============================
    # SAVE AS
    # =============================
    file_path = filedialog.asksaveasfilename(
        title="Guardar Libro Diario",
        defaultextension=".xlsx",
        filetypes=[("Excel", "*.xlsx")]
    )
    if not file_path:
        return None

    period_label = f"{int(fiscal_year)}-{int(period):02d}"

    wb = Workbook()
    ws = wb.active
    ws.title = "Libro Diario"

    # =============================
    # ANCHOS
    # =============================
    widths = {
        "A": 14,  # Fecha
        "B": 12,  # Asiento
        "C": 14,  # Cuenta
        "D": 35,  # Nombre de la Cuenta
        "E": 16,  # Debe
        "F": 16,  # Haber
        "G": 40,  # Detalle
        "H": 18,  # Origen
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # =============================
    # ENCABEZADO (SIN EMPRESA)
    # =============================
    ws.merge_cells("A1:H1")
    ws["A1"] = "LIBRO DIARIO"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:H2")
    ws["A2"] = f"Período fiscal: {period_label}"
    ws["A2"].font = Font(bold=True, size=12)
    ws["A2"].alignment = Alignment(horizontal="center")

    row = 4

    # =============================
    # CABECERAS
    # =============================
    headers = [
        "Fecha",
        "Asiento",
        "Cuenta",
        "Nombre de la Cuenta",
        "Debe",
        "Haber",
        "Detalle",
        "Origen"
    ]

    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
        c.border = BORDER

    row += 1

    total_debe = 0.0
    total_haber = 0.0
    last_entry_id = None
    first_entry = True

    rows = sorted(
        rows,
        key=lambda item: (
            str(item.get("period") or ""),
            str(item.get("entry_date") or item.get("created_at") or ""),
            int(item.get("entry_id") or 0),
            int(item.get("line_id") or 0),
        ),
    )

    # =============================
    # DATA (SQL PURO)
    # =============================
    for r in rows:
        entry_id = r.get("entry_id")
        if entry_id != last_entry_id:
            if not first_entry:
                row += 1
            marker = ws.cell(row=row, column=1, value=f"Asiento {entry_id} | {r.get('period') or period_label}")
            marker.font = Font(bold=True, color="003A75")
            marker.fill = ENTRY_FILL
            marker.alignment = Alignment(horizontal="left")
            for c in range(1, 9):
                cell = ws.cell(row=row, column=c)
                cell.fill = ENTRY_FILL
                cell.border = BORDER
            row += 1
            last_entry_id = entry_id
            first_entry = False

        # created_at real
        fecha = to_long_english_date(r.get("entry_date") or r.get("created_at"))

        ws.cell(row=row, column=1, value=fecha)
        ws.cell(row=row, column=2, value=entry_id)
        ws.cell(row=row, column=3, value=r.get("account_code"))
        ws.cell(row=row, column=4, value=r.get("account_name"))

        debe = float(r.get("debit") or 0)
        haber = float(r.get("credit") or 0)

        ws.cell(row=row, column=5, value=debe)
        ws.cell(row=row, column=6, value=haber)
        ws.cell(row=row, column=7, value=r.get("line_description", ""))
        ws.cell(row=row, column=8, value=r.get("origin", ""))

        ws.cell(row=row, column=5).number_format = "#,##0.00"
        ws.cell(row=row, column=6).number_format = "#,##0.00"

        total_debe += debe
        total_haber += haber

        for c in range(1, 9):
            cell = ws.cell(row=row, column=c)
            cell.border = BORDER
            if row % 2 == 0:
                cell.fill = ALT_FILL

        row += 1

    # =============================
    # TOTALES
    # =============================
    ws.cell(row=row, column=4, value="TOTALES").font = Font(bold=True)
    ws.cell(row=row, column=5, value=total_debe).font = Font(bold=True)
    ws.cell(row=row, column=6, value=total_haber).font = Font(bold=True)

    ws.cell(row=row, column=5).number_format = "#,##0.00"
    ws.cell(row=row, column=6).number_format = "#,##0.00"

    for c in range(1, 9):
        ws.cell(row=row, column=c).border = BORDER

    wb.save(file_path)
    return file_path

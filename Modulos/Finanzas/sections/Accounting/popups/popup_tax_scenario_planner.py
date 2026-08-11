import threading
import tkinter as tk
from datetime import date
from tkinter import ttk, messagebox, filedialog

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from api_client import post_accounting_tax_scenario_analysis_api


def _money(value):
    try:
        return f"CRC {float(value or 0):,.2f}"
    except Exception:
        return "CRC 0.00"


def _pct(value):
    try:
        return f"{float(value or 0):,.2f}%"
    except Exception:
        return "0.00%"


class PopupTaxScenarioPlanner(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Simulador fiscal multiempresa")
        self.geometry("1340x820")
        self.minsize(1120, 680)
        self.resizable(True, True)
        try:
            self.wm_attributes("-toolwindow", False)
        except Exception:
            pass
        self.grab_set()
        self.result = None
        self.client_moves = []
        self.expense_moves = []
        self.fixed_clients = []
        self.fixed_expenses = []
        self.client_rows_by_item = {}
        self.expense_rows_by_item = {}
        self._build_ui()
        self.after(150, self._try_zoom)

    def _try_zoom(self):
        try:
            self.state("zoomed")
        except Exception:
            pass

    def _build_ui(self):
        header = ttk.Frame(self, padding=10)
        header.pack(fill="x")
        ttk.Label(header, text="Simulador fiscal multiempresa con IA", font=("Segoe UI", 14, "bold")).pack(side="left")
        ttk.Button(header, text="Cerrar", command=self.destroy).pack(side="right")
        ttk.Button(header, text="Exportar Excel", command=self._export_excel).pack(side="right", padx=6)
        ttk.Button(header, text="Copiar analisis", command=self._copy_analysis).pack(side="right", padx=6)

        controls = ttk.LabelFrame(self, text="Parametros de simulacion", padding=10)
        controls.pack(fill="x", padx=10, pady=(0, 8))

        self.year_var = tk.IntVar(value=date.today().year)
        self.month_var = tk.IntVar(value=date.today().month)
        self.source_company_var = tk.StringVar(value="MSL-CR")
        self.target_company_var = tk.StringVar(value="MMS-CR")
        self.msl_pyme_year_var = tk.IntVar(value=4)
        self.mms_pyme_year_var = tk.IntVar(value=1)
        self.save_var = tk.BooleanVar(value=False)

        fields = [
            ("Ano", self.year_var, 7),
            ("Mes corte", self.month_var, 7),
            ("Empresa actual", self.source_company_var, 12),
            ("Empresa alterna", self.target_company_var, 12),
            ("Ano PYME actual", self.msl_pyme_year_var, 7),
            ("Ano PYME alterna", self.mms_pyme_year_var, 7),
        ]
        for col, (label, var, width) in enumerate(fields):
            ttk.Label(controls, text=label).grid(row=0, column=col * 2, sticky="w", padx=(0, 4), pady=3)
            ttk.Entry(controls, textvariable=var, width=width).grid(row=0, column=col * 2 + 1, padx=(0, 10), pady=3)

        ttk.Checkbutton(controls, text="Guardar escenario", variable=self.save_var).grid(row=1, column=0, columnspan=2, sticky="w", pady=(8, 0))
        ttk.Button(controls, text="Analizar actual", command=self._run).grid(row=1, column=8, sticky="e", padx=5, pady=(8, 0))
        ttk.Button(controls, text="Aplicar IA recomendada", command=self._apply_auto_moves).grid(row=1, column=9, sticky="e", padx=5, pady=(8, 0))
        ttk.Button(controls, text="Recalcular escenario", command=self._run).grid(row=1, column=10, columnspan=2, sticky="e", padx=5, pady=(8, 0))

        kpi = ttk.Frame(self, padding=(10, 0))
        kpi.pack(fill="x")
        self.kpi_vars = {
            "gross": tk.StringVar(value="Real 2026: CRC 0.00"),
            "baseline": tk.StringVar(value="Actual: CRC 0.00"),
            "optimized": tk.StringVar(value="Escenario: CRC 0.00"),
            "saving": tk.StringVar(value="Impacto: CRC 0.00"),
            "threshold": tk.StringVar(value="Umbral PYME: 0.00%"),
        }
        for idx, (title, key) in enumerate((
            ("Ventas reales 2026", "gross"),
            ("Renta actual", "baseline"),
            ("Renta escenario", "optimized"),
            ("Ahorro / impacto", "saving"),
            ("Uso umbral PYME", "threshold"),
        )):
            box = ttk.LabelFrame(kpi, text=title, padding=8)
            box.grid(row=0, column=idx, sticky="ew", padx=(0, 7))
            kpi.columnconfigure(idx, weight=1)
            ttk.Label(box, textvariable=self.kpi_vars[key], font=("Segoe UI", 11, "bold")).pack(anchor="w")

        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill="both", expand=True, padx=10, pady=10)
        self._build_sales_tab()
        self._build_expenses_tab()
        self._build_statement_tab()
        self._build_analysis_tab()

        self.status_var = tk.StringVar(value="Listo. Pulsa Analizar actual.")
        ttk.Label(self, textvariable=self.status_var).pack(anchor="w", padx=12, pady=(0, 8))

    def _build_sales_tab(self):
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="Ventas por cliente")
        actions = ttk.LabelFrame(tab, text="Mover venta entre sociedades", padding=8)
        actions.pack(fill="x")
        self.sales_hint_var = tk.StringVar(value="Selecciona un cliente. La columna Decision muestra si IA o usuario lo mueve.")
        ttk.Label(actions, textvariable=self.sales_hint_var, font=("Segoe UI", 9, "bold")).pack(side="left")
        ttk.Button(actions, text=f"Mover venta a {self.target_company_var.get()}", command=self._move_selected_client_to_target).pack(side="right", padx=4)
        ttk.Button(actions, text=f"Dejar venta en {self.source_company_var.get()}", command=self._move_selected_client_to_source).pack(side="right", padx=4)
        ttk.Button(actions, text="Mantener monto", command=self._lock_selected_client_projection).pack(side="right", padx=4)
        ttk.Button(actions, text="Proyectar normal", command=self._unlock_selected_client_projection).pack(side="right", padx=4)

        cols = ("company", "client", "invoices", "ytd", "future", "projected", "decision")
        self.clients_tree = ttk.Treeview(tab, columns=cols, show="headings")
        headers = {
            "company": "Empresa",
            "client": "Cliente",
            "invoices": "Facturas",
            "ytd": "YTD",
            "future": "Futuro movible",
            "projected": "Proyeccion anual",
            "decision": "Decision",
        }
        widths = {"company": 95, "client": 310, "invoices": 75, "ytd": 145, "future": 145, "projected": 150, "decision": 190}
        self._setup_tree(self.clients_tree, headers, widths)
        self._pack_tree(tab, self.clients_tree)
        self.clients_tree.bind("<<TreeviewSelect>>", self._on_client_select)
        self.clients_tree.bind("<Double-1>", lambda _e: self._move_selected_client_to_target())

    def _build_expenses_tab(self):
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="Gastos deducibles")
        actions = ttk.LabelFrame(tab, text="Mover gasto entre sociedades", padding=8)
        actions.pack(fill="x")
        self.expense_hint_var = tk.StringVar(value="Incluye gastos POSTED, facturas ITP pendientes sin asiento y planilla sin asiento.")
        ttk.Label(actions, textvariable=self.expense_hint_var, font=("Segoe UI", 9, "bold")).pack(side="left")
        ttk.Button(actions, text=f"Mover gasto a {self.target_company_var.get()}", command=self._move_selected_expense_to_target).pack(side="right", padx=4)
        ttk.Button(actions, text=f"Dejar gasto en {self.source_company_var.get()}", command=self._move_selected_expense_to_source).pack(side="right", padx=4)
        ttk.Button(actions, text="Mantener monto", command=self._lock_selected_expense_projection).pack(side="right", padx=4)
        ttk.Button(actions, text="Proyectar normal", command=self._unlock_selected_expense_projection).pack(side="right", padx=4)

        cols = ("company", "source", "account", "name", "status", "entries", "ytd", "projected", "decision")
        self.expenses_tree = ttk.Treeview(tab, columns=cols, show="headings")
        headers = {
            "company": "Empresa",
            "source": "Fuente",
            "account": "Cuenta",
            "name": "Gasto",
            "status": "Estado",
            "entries": "Asientos",
            "ytd": "YTD",
            "projected": "Proyeccion anual",
            "decision": "Decision",
        }
        widths = {"company": 90, "source": 120, "account": 120, "name": 340, "status": 90, "entries": 70, "ytd": 140, "projected": 150, "decision": 170}
        self._setup_tree(self.expenses_tree, headers, widths)
        self._pack_tree(tab, self.expenses_tree)
        self.expenses_tree.bind("<<TreeviewSelect>>", self._on_expense_select)
        self.expenses_tree.bind("<Double-1>", lambda _e: self._move_selected_expense_to_target())

    def _build_statement_tab(self):
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="Estado de resultados y renta")
        intro = ttk.LabelFrame(tab, text="Lectura tipo D-102", padding=8)
        intro.pack(fill="x", pady=(0, 8))
        ttk.Label(
            intro,
            text=(
                "Primero lee los 4 cuadros de resumen. La tabla inferior es el soporte tipo D-102 para ver de donde sale cada monto."
            ),
            font=("Segoe UI", 9, "bold"),
        ).pack(anchor="w")

        self.statement_summary_vars = {
            "today": tk.StringVar(value="Hoy: CRC 0.00"),
            "without_moves": tk.StringVar(value="Sin mover: CRC 0.00"),
            "with_moves": tk.StringVar(value="Con movimientos: CRC 0.00"),
            "decision": tk.StringVar(value="Ejecuta Analizar actual para ver recomendacion."),
        }
        summary = ttk.Frame(tab)
        summary.pack(fill="x", pady=(0, 8))
        for idx, (title, key) in enumerate((
            ("1. Venta real MSL hoy", "today"),
            ("2. Cierre estimado sin mover", "without_moves"),
            ("3. Cierre estimado con movimientos", "with_moves"),
            ("4. Lectura ejecutiva", "decision"),
        )):
            box = ttk.LabelFrame(summary, text=title, padding=8)
            box.grid(row=0, column=idx, sticky="ew", padx=(0, 7))
            summary.columnconfigure(idx, weight=1)
            ttk.Label(box, textvariable=self.statement_summary_vars[key], font=("Segoe UI", 10, "bold"), wraplength=280).pack(anchor="w")

        self.pyme_badge_vars = {
            "today": tk.StringVar(value="PYME hoy: pendiente"),
            "dec_no_move": tk.StringVar(value="PYME diciembre sin mover: pendiente"),
            "dec_moves": tk.StringVar(value="PYME diciembre con movimientos: pendiente"),
        }
        badges = ttk.Frame(tab)
        badges.pack(fill="x", pady=(0, 8))
        for idx, (key, title) in enumerate((
            ("today", "Estado real a hoy"),
            ("dec_no_move", "Cierre sin mover"),
            ("dec_moves", "Cierre con movimientos"),
        )):
            box = ttk.LabelFrame(badges, text=title, padding=8)
            box.grid(row=0, column=idx, sticky="ew", padx=(0, 7))
            badges.columnconfigure(idx, weight=1)
            ttk.Label(box, textvariable=self.pyme_badge_vars[key], font=("Segoe UI", 10, "bold"), wraplength=390).pack(anchor="w")

        self.d102_cards = {}
        cards = ttk.LabelFrame(tab, text="Formulario D-102 por escenario", padding=8)
        cards.pack(fill="both", expand=True, pady=(0, 8))
        for idx, (key, title) in enumerate((
            ("baseline_msl", "1. MSL sin movimientos"),
            ("scenario_msl", "2. MSL con movimientos"),
            ("scenario_mms", "3. MMS con movimientos"),
        )):
            card = tk.Frame(cards, bg="#ffffff", bd=1, relief="solid")
            card.grid(row=0, column=idx, sticky="nsew", padx=(0, 8))
            cards.columnconfigure(idx, weight=1)
            cards.rowconfigure(0, weight=1)
            self.d102_cards[key] = card
            self._render_d102_empty_card(card, title)

        d102_frame = ttk.LabelFrame(tab, text="D-102 comparativo principal", padding=6)
        d102_frame.pack(fill="x", pady=(0, 8))
        d102_cols = ("line", "msl_no_move", "msl_scenario", "mms_scenario", "meaning")
        self.d102_tree = ttk.Treeview(d102_frame, columns=d102_cols, show="headings", height=7)
        d102_headers = {
            "line": "Linea D-102",
            "msl_no_move": "MSL dic sin mover",
            "msl_scenario": "MSL dic con cambios",
            "mms_scenario": "MMS dic con cambios",
            "meaning": "Lectura",
        }
        d102_widths = {"line": 285, "msl_no_move": 170, "msl_scenario": 170, "mms_scenario": 170, "meaning": 500}
        self._setup_tree(self.d102_tree, d102_headers, d102_widths)
        self._pack_tree(d102_frame, self.d102_tree)

        cols = ("company", "line", "actual", "scenario", "diff", "note")
        self.statement_tree = ttk.Treeview(tab, columns=cols, show="headings", height=6)
        headers = {
            "company": "Empresa",
            "line": "Linea D-102 / Estado de resultados",
            "actual": "Sin mover",
            "scenario": "Con movimientos",
            "diff": "Diferencia",
            "note": "Lectura",
        }
        widths = {"company": 95, "line": 280, "actual": 155, "scenario": 155, "diff": 155, "note": 430}
        self._setup_tree(self.statement_tree, headers, widths)
        self._pack_tree(tab, self.statement_tree)

        detail_frame = ttk.LabelFrame(tab, text="Desglose de tramos de renta")
        detail_frame.pack(fill="both", expand=True, pady=(8, 0))
        cols2 = ("scenario", "company", "from", "to", "rate", "taxable", "tax")
        self.tax_detail_tree = ttk.Treeview(detail_frame, columns=cols2, show="headings", height=8)
        headers2 = {"scenario": "Escenario", "company": "Empresa", "from": "Desde", "to": "Hasta", "rate": "Tarifa", "taxable": "Base", "tax": "Impuesto"}
        widths2 = {"scenario": 130, "company": 90, "from": 130, "to": 130, "rate": 90, "taxable": 150, "tax": 150}
        self._setup_tree(self.tax_detail_tree, headers2, widths2)
        self._pack_tree(detail_frame, self.tax_detail_tree)

    def _build_analysis_tab(self):
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="Analisis inteligente")
        self.analysis_text = tk.Text(tab, wrap="word", font=("Segoe UI", 10), padx=8, pady=8)
        scroll = ttk.Scrollbar(tab, orient="vertical", command=self.analysis_text.yview)
        self.analysis_text.configure(yscrollcommand=scroll.set)
        self.analysis_text.pack(side="left", fill="both", expand=True)
        scroll.pack(side="right", fill="y")

    def _setup_tree(self, tree, headers, widths):
        for col, text in headers.items():
            tree.heading(col, text=text)
            tree.column(col, width=widths.get(col, 120), anchor="w", stretch=True)

    def _pack_tree(self, parent, tree):
        yscroll = ttk.Scrollbar(parent, orient="vertical", command=tree.yview)
        xscroll = ttk.Scrollbar(parent, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)
        tree.pack(side="left", fill="both", expand=True)
        yscroll.pack(side="right", fill="y")
        xscroll.pack(side="bottom", fill="x")

    def _payload(self):
        return {
            "year": int(self.year_var.get() or date.today().year),
            "through_month": int(self.month_var.get() or date.today().month),
            "source_company": self.source_company_var.get().strip() or "MSL-CR",
            "target_company": self.target_company_var.get().strip() or "MMS-CR",
            "client_moves": list(self.client_moves),
            "expense_moves": list(self.expense_moves),
            "fixed_clients": list(self.fixed_clients),
            "fixed_expenses": list(self.fixed_expenses),
            "save": bool(self.save_var.get()),
            "label": f"Escenario fiscal {self.year_var.get()}-{int(self.month_var.get() or 1):02d}",
            "company_options": [
                {"company_code": self.source_company_var.get().strip() or "MSL-CR", "is_pyme": True, "pyme_year": int(self.msl_pyme_year_var.get() or 0)},
                {"company_code": self.target_company_var.get().strip() or "MMS-CR", "is_pyme": True, "pyme_year": int(self.mms_pyme_year_var.get() or 0)},
            ],
        }

    def _run(self):
        self.status_var.set("Analizando ventas, gastos, renta y umbral PYME...")
        self._clear()
        threading.Thread(target=self._worker, args=(self._payload(),), daemon=True).start()

    def _worker(self, payload):
        try:
            data = post_accounting_tax_scenario_analysis_api(payload)
            self.after(0, lambda: self._render(data))
        except Exception as exc:
            self.after(0, lambda: self._error(exc))

    def _clear(self):
        for tree in (self.clients_tree, self.expenses_tree, self.d102_tree, self.statement_tree, self.tax_detail_tree):
            for item in tree.get_children():
                tree.delete(item)
        for frame in getattr(self, "d102_cards", {}).values():
            self._clear_frame(frame)
        self.client_rows_by_item.clear()
        self.expense_rows_by_item.clear()
        self.analysis_text.delete("1.0", "end")

    def _render(self, data):
        self.result = data
        baseline = data.get("baseline") or {}
        optimized = data.get("optimized") or {}
        analysis = data.get("analysis") or {}
        auto_moves = {(m.get("from_company"), m.get("client_name")): m for m in data.get("auto_moves") or []}
        manual_client_moves = {(m.get("from_company"), m.get("client_name")): m for m in self.client_moves}
        manual_expense_moves = {
            (m.get("from_company"), m.get("source_type"), m.get("account_code"), m.get("account_name")): m
            for m in self.expense_moves
        }

        total_gross = sum(float((r or {}).get("gross_ytd_crc") or 0) for r in baseline.get("companies") or [])
        threshold_usage = max((float((r or {}).get("pyme_threshold_usage_pct") or 0) for r in baseline.get("companies") or []), default=0)
        self.kpi_vars["gross"].set(_money(total_gross))
        self.kpi_vars["baseline"].set(_money(baseline.get("total_projected_tax_crc")))
        self.kpi_vars["optimized"].set(_money(optimized.get("total_projected_tax_crc")))
        self.kpi_vars["saving"].set(_money(analysis.get("tax_saving_crc")))
        self.kpi_vars["threshold"].set(_pct(threshold_usage))

        for row in data.get("clients") or []:
            key = (row.get("company_code"), row.get("client_name"))
            decision = ""
            if key in manual_client_moves:
                target_company = manual_client_moves[key].get("to_company")
                if target_company == row.get("company_code"):
                    decision = f"Dejar en {target_company}"
                else:
                    decision = f"Manual -> {target_company}"
            elif key in auto_moves:
                decision = f"IA -> {auto_moves[key].get('to_company')}"
            if row.get("projection_mode") == "FIXED":
                decision = f"{decision} | Monto fijo" if decision else "Monto fijo"
            tags = ("moved",) if decision else ()
            item = self.clients_tree.insert("", "end", values=(
                row.get("company_code"),
                row.get("client_name"),
                row.get("invoice_count"),
                _money(row.get("ytd_amount_crc")),
                _money(row.get("future_projected_crc")),
                _money(row.get("projected_annual_crc")),
                decision,
            ), tags=tags)
            self.client_rows_by_item[item] = row
        self.clients_tree.tag_configure("moved", background="#e9f7ef")

        for row in data.get("expense_rows") or []:
            key = (row.get("company_code"), row.get("source_type"), row.get("account_code"), row.get("account_name"))
            decision = f"Manual -> {manual_expense_moves[key].get('to_company')}" if key in manual_expense_moves else ""
            if row.get("projection_mode") == "FIXED":
                decision = f"{decision} | Monto fijo" if decision else "Monto fijo"
            source_label = {
                "POSTED_GL": "Contable POSTED",
                "ITP_PENDING": "ITP pendiente",
                "PAYROLL_PENDING": "Planilla pendiente",
            }.get(row.get("source_type"), row.get("source_type"))
            tags = ("moved",) if decision else (("pending",) if row.get("source_type") != "POSTED_GL" else ())
            item = self.expenses_tree.insert("", "end", values=(
                row.get("company_code"),
                source_label,
                row.get("account_code"),
                row.get("account_name"),
                row.get("status"),
                row.get("entry_count"),
                _money(row.get("ytd_amount_crc")),
                _money(row.get("projected_annual_crc")),
                decision,
            ), tags=tags)
            self.expense_rows_by_item[item] = row
        self.expenses_tree.tag_configure("moved", background="#e9f7ef")
        self.expenses_tree.tag_configure("pending", background="#fff8e1")

        self._render_d102_cards(baseline, optimized)
        self._render_d102_main(baseline, optimized)
        self._render_statement_comparison(baseline, optimized)
        self._render_statement_summary(baseline, optimized, analysis)
        for section, title in ((baseline, "Actual"), (optimized, "Escenario")):
            for row in section.get("companies") or []:
                for bracket in row.get("tax_detail") or []:
                    self.tax_detail_tree.insert("", "end", values=(
                        title,
                        row.get("company_code"),
                        _money(bracket.get("from")),
                        "En adelante" if bracket.get("to") is None else _money(bracket.get("to")),
                        _pct(float(bracket.get("rate") or 0) * 100),
                        _money(bracket.get("taxable")),
                        _money(bracket.get("tax")),
                    ))

        self._render_analysis(data)
        saved = f" Guardado #{data.get('saved_id')}." if data.get("saved_id") else ""
        self.status_var.set(f"Analisis completo.{saved}")

    def _company_row(self, section, code):
        for row in section.get("companies") or []:
            if row.get("company_code") == code:
                return row
        return {}

    def _pyme_status(self, row, projected=False):
        if projected:
            usage = float(row.get("pyme_threshold_usage_pct") or 0)
            margin = float(row.get("pyme_threshold_remaining_crc") or 0)
        else:
            usage = float(row.get("pyme_threshold_ytd_usage_pct") or 0)
            margin = float(row.get("pyme_threshold_ytd_remaining_crc") or 0)
        status = "EXCEDIDO" if usage > 100 else "NO excedido"
        margin_text = f"exceso {_money(abs(margin))}" if margin < 0 else f"margen {_money(margin)}"
        return f"{status} | {_pct(usage)} | {margin_text}"

    def _clear_frame(self, frame):
        for child in frame.winfo_children():
            child.destroy()

    def _render_d102_empty_card(self, frame, title):
        self._clear_frame(frame)
        tk.Label(frame, text=title, bg="#f4f6f8", fg="#0b3f75", font=("Segoe UI", 10, "bold"), anchor="w", padx=8, pady=6).pack(fill="x")
        tk.Label(frame, text="Ejecuta Analizar actual para cargar el formulario.", bg="#ffffff", fg="#444444", anchor="w", padx=8, pady=12).pack(fill="x")

    def _render_d102_cards(self, baseline, optimized):
        source = self.source_company_var.get().strip() or "MSL-CR"
        target = self.target_company_var.get().strip() or "MMS-CR"
        cards = (
            ("baseline_msl", f"{source} sin movimientos", self._company_row(baseline, source)),
            ("scenario_msl", f"{source} con movimientos", self._company_row(optimized, source)),
            ("scenario_mms", f"{target} con movimientos", self._company_row(optimized, target)),
        )
        for key, title, row in cards:
            frame = self.d102_cards.get(key)
            if not frame:
                continue
            self._clear_frame(frame)
            exceeded = bool(row.get("pyme_gross_limit_exceeded"))
            status_bg = "#fdecea" if exceeded else "#e9f7ef"
            status_fg = "#8a1f11" if exceeded else "#0d5b2a"
            tk.Label(frame, text="D-102 | Impuesto sobre utilidades", bg="#0b3f75", fg="#ffffff", font=("Segoe UI", 10, "bold"), anchor="w", padx=8, pady=6).pack(fill="x")
            tk.Label(frame, text=title, bg="#ffffff", fg="#111111", font=("Segoe UI", 11, "bold"), anchor="w", padx=8, pady=(8, 2)).pack(fill="x")
            tk.Label(frame, text=self._pyme_status(row, projected=True), bg=status_bg, fg=status_fg, font=("Segoe UI", 9, "bold"), anchor="w", padx=8, pady=5, wraplength=360).pack(fill="x", padx=8, pady=(4, 8))
            lines = [
                ("1", "Renta bruta / ingresos", row.get("gross_projected_crc"), True),
                ("1.1", "Ventas facturadas 2026", row.get("gross_ytd_crc"), False),
                ("1.2", "Venta futura estimada", row.get("gross_future_projected_crc"), False),
                ("2", "Costos, gastos y deducciones", row.get("deductible_expenses_projected_crc"), True),
                ("3", "Renta neta gravable", row.get("net_taxable_projected_crc"), True),
                ("4", "Impuesto base", row.get("base_income_tax_crc"), False),
                ("5", "Exoneracion PYME", self._exemption_amount(row), False),
                ("6", "Total impuesto determinado", row.get("income_tax_projected_crc"), True),
            ]
            for line, label, value, strong in lines:
                self._d102_card_line(frame, line, label, _money(value), strong=strong)
            tk.Label(frame, text=self._tax_rule_text(row), bg="#ffffff", fg="#555555", font=("Segoe UI", 8), anchor="w", padx=8, pady=6, wraplength=360).pack(fill="x")

    def _d102_card_line(self, frame, line, label, value, strong=False):
        bg = "#f7f9fb" if strong else "#ffffff"
        row_frame = tk.Frame(frame, bg=bg)
        row_frame.pack(fill="x", padx=8)
        font = ("Segoe UI", 9, "bold") if strong else ("Segoe UI", 9)
        tk.Label(row_frame, text=line, width=5, bg=bg, fg="#0b3f75", font=font, anchor="w").pack(side="left")
        tk.Label(row_frame, text=label, bg=bg, fg="#222222", font=font, anchor="w").pack(side="left", fill="x", expand=True)
        tk.Label(row_frame, text=value, bg=bg, fg="#111111", font=font, anchor="e").pack(side="right")

    def _render_d102_main(self, baseline, optimized):
        source = self.source_company_var.get().strip() or "MSL-CR"
        target = self.target_company_var.get().strip() or "MMS-CR"
        base_source = self._company_row(baseline, source)
        opt_source = self._company_row(optimized, source)
        opt_target = self._company_row(optimized, target)
        baseline_total = float(baseline.get("total_projected_tax_crc") or 0)
        optimized_total = float(optimized.get("total_projected_tax_crc") or 0)
        self.pyme_badge_vars["today"].set(
            f"{source}: {self._pyme_status(base_source, projected=False)}. Es la foto real facturada a hoy."
        )
        self.pyme_badge_vars["dec_no_move"].set(
            f"{source}: {self._pyme_status(base_source, projected=True)}. Esta es la estimacion si sigue el mismo ritmo."
        )
        self.pyme_badge_vars["dec_moves"].set(
            f"{source}: {self._pyme_status(opt_source, projected=True)} | {target}: {self._pyme_status(opt_target, projected=True)}."
        )
        rows = [
            (
                "1. Ventas reales ya facturadas 2026",
                _money(base_source.get("gross_ytd_crc")),
                _money(opt_source.get("gross_ytd_crc")),
                _money(opt_target.get("gross_ytd_crc")),
                "Lo ya facturado no se mueve. Sirve para auditar que el escenario no eche ventas hacia atras.",
                "",
            ),
            (
                "1.1 Venta futura estimada",
                _money(base_source.get("gross_future_projected_crc")),
                _money(opt_source.get("gross_future_projected_crc")),
                _money(opt_target.get("gross_future_projected_crc")),
                "Solo esta parte se reasigna al mover clientes; es la diferencia entre diciembre proyectado y hoy.",
                "",
            ),
            (
                "1.2 Total ingresos / renta bruta",
                _money(base_source.get("gross_projected_crc")),
                _money(opt_source.get("gross_projected_crc")),
                _money(opt_target.get("gross_projected_crc")),
                "Ventas reales mas venta futura estimada por sociedad.",
                "",
            ),
            (
                "2. Costos, gastos y deducciones",
                _money(base_source.get("deductible_expenses_projected_crc")),
                _money(opt_source.get("deductible_expenses_projected_crc")),
                _money(opt_target.get("deductible_expenses_projected_crc")),
                "Gastos proyectados; al mover gastos cambian aqui y luego impactan renta neta.",
                "",
            ),
            (
                "3. Renta neta",
                _money(base_source.get("net_taxable_projected_crc")),
                _money(opt_source.get("net_taxable_projected_crc")),
                _money(opt_target.get("net_taxable_projected_crc")),
                "Renta bruta menos costos, gastos y deducciones.",
                "",
            ),
            (
                "4. Regla impuesto SA/PYME",
                self._tax_rule_text(base_source),
                self._tax_rule_text(opt_source),
                self._tax_rule_text(opt_target),
                "Se decide con la renta bruta proyectada contra el umbral PYME.",
                "",
            ),
            (
                "5. Impuesto sobre utilidades base",
                _money(base_source.get("base_income_tax_crc")),
                _money(opt_source.get("base_income_tax_crc")),
                _money(opt_target.get("base_income_tax_crc")),
                "Impuesto antes de aplicar exoneracion PYME.",
                "",
            ),
            (
                "6. Exoneracion PYME",
                _money(self._exemption_amount(base_source)),
                _money(self._exemption_amount(opt_source)),
                _money(self._exemption_amount(opt_target)),
                "Beneficio por año PYME, si no se excede el umbral.",
                "",
            ),
            (
                "7. Total impuesto determinado",
                _money(base_source.get("income_tax_projected_crc")),
                _money(opt_source.get("income_tax_projected_crc")),
                _money(opt_target.get("income_tax_projected_crc")),
                "Base menos exoneracion PYME. Si ano PYME 1-3 y no excede, puede quedar en cero.",
                "good",
            ),
            (
                "8. Total impuesto grupo",
                _money(baseline_total),
                _money(optimized_total),
                "",
                "Total a pagar sumando las sociedades del escenario.",
                "good" if optimized_total <= baseline_total else "warn",
            ),
        ]
        for values in rows:
            tag = values[-1]
            self.d102_tree.insert("", "end", values=values[:-1], tags=(tag,) if tag else ())
        self.d102_tree.tag_configure("good", background="#e9f7ef")
        self.d102_tree.tag_configure("warn", background="#fff3cd")
        self.d102_tree.tag_configure("bad", background="#ffd6d6")

    def _tax_rule_text(self, row):
        if not row:
            return ""
        regime = row.get("regime") or ""
        exemption = float(row.get("pyme_exemption_rate") or 0) * 100
        pyme_year = row.get("pyme_year")
        if regime == "GENERAL_30":
            return "SA general 30% (sin PYME por umbral)"
        if exemption:
            return f"PYME tramos D-102; ano {pyme_year}; exonera {_pct(exemption)}"
        return f"PYME tramos D-102; ano {pyme_year}; sin exoneracion"

    def _render_statement_summary(self, baseline, optimized, analysis):
        source = self.source_company_var.get().strip() or "MSL-CR"
        target = self.target_company_var.get().strip() or "MMS-CR"
        base_source = self._company_row(baseline, source)
        opt_source = self._company_row(optimized, source)
        opt_target = self._company_row(optimized, target)

        ytd = float(base_source.get("gross_ytd_crc") or 0)
        ytd_margin = float(base_source.get("pyme_threshold_ytd_remaining_crc") or 0)
        base_projection = float(base_source.get("gross_projected_crc") or 0)
        base_projection_margin = float(base_source.get("pyme_threshold_remaining_crc") or 0)
        opt_projection = float(opt_source.get("gross_projected_crc") or 0)
        target_projection = float(opt_target.get("gross_projected_crc") or 0)
        opt_margin = float(opt_source.get("pyme_threshold_remaining_crc") or 0)
        target_margin = float(opt_target.get("pyme_threshold_remaining_crc") or 0)

        self.statement_summary_vars["today"].set(
            f"{_money(ytd)} | margen real {_money(ytd_margin)} antes de 119.174M."
        )
        self.statement_summary_vars["without_moves"].set(
            f"{_money(base_projection)} | {'exceso' if base_projection_margin < 0 else 'margen'} {_money(abs(base_projection_margin))}."
        )
        self.statement_summary_vars["with_moves"].set(
            f"{source}: {_money(opt_projection)} | {target}: {_money(target_projection)}."
        )
        if opt_margin >= 0 and target_margin >= 0:
            decision = "Con estos movimientos, ambas sociedades quedan debajo del umbral proyectado."
        elif opt_margin < 0:
            decision = f"{source} aun excede por {_money(abs(opt_margin))}; mueva mas venta o revise proyeccion."
        elif target_margin < 0:
            decision = f"{target} excede por {_money(abs(target_margin))}; no conviene mover tanto ahi."
        else:
            decision = analysis.get("recommendation") or "Revise escenario."
        self.statement_summary_vars["decision"].set(decision)

    def _render_statement_comparison(self, baseline, optimized):
        baseline_map = {row.get("company_code"): row for row in baseline.get("companies") or []}
        optimized_map = {row.get("company_code"): row for row in optimized.get("companies") or []}
        for company in sorted(set(baseline_map) | set(optimized_map)):
            actual = baseline_map.get(company) or {}
            scenario = optimized_map.get(company) or {}
            rows = [
                (
                    "1. Total ingresos / renta bruta real 2026",
                    actual.get("gross_ytd_crc"),
                    scenario.get("gross_ytd_crc"),
                    "Ventas reales posteadas a hoy desde ingresos contables.",
                    "money",
                ),
                (
                    "1.1 Proyeccion renta bruta a diciembre",
                    actual.get("gross_projected_crc"),
                    scenario.get("gross_projected_crc"),
                    "Proyeccion gerencial para anticipar si se excede PYME al cierre.",
                    "money",
                ),
                (
                    "2. Costos, gastos y deducciones",
                    actual.get("deductible_expenses_projected_crc"),
                    scenario.get("deductible_expenses_projected_crc"),
                    "Incluye gastos POSTED y pendientes visibles en la pestaña Gastos.",
                    "money",
                ),
                (
                    "3. Renta neta antes de exoneraciones",
                    actual.get("net_taxable_projected_crc"),
                    scenario.get("net_taxable_projected_crc"),
                    "Base fiscal antes de beneficio PYME.",
                    "money",
                ),
                (
                    "4. Impuesto sobre utilidades base",
                    actual.get("base_income_tax_crc"),
                    scenario.get("base_income_tax_crc"),
                    "Impuesto calculado por tarifa general o tramos D-102.",
                    "money",
                ),
                (
                    "5. Exoneracion PYME aplicada",
                    self._exemption_amount(actual),
                    self._exemption_amount(scenario),
                    "Monto rebajado por año PYME, si no se excede el umbral.",
                    "money",
                ),
                (
                    "6. Impuesto final proyectado",
                    actual.get("income_tax_projected_crc"),
                    scenario.get("income_tax_projected_crc"),
                    "Monto estimado a pagar por impuesto sobre utilidades.",
                    "money",
                ),
                (
                    "7. Uso real del umbral PYME a hoy",
                    actual.get("pyme_threshold_ytd_usage_pct"),
                    scenario.get("pyme_threshold_ytd_usage_pct"),
                    "Foto real 2026: sobre 100% ya estaria excedido a hoy.",
                    "pct",
                ),
                (
                    "7.1 Uso proyectado del umbral PYME",
                    actual.get("pyme_threshold_usage_pct"),
                    scenario.get("pyme_threshold_usage_pct"),
                    "Proyeccion a diciembre: sobre 100% se perderia PYME al cierre.",
                    "pct",
                ),
                (
                    "8. Margen real contra 119.174M",
                    actual.get("pyme_threshold_ytd_remaining_crc"),
                    scenario.get("pyme_threshold_ytd_remaining_crc"),
                    "Margen disponible real a hoy antes de exceder el umbral.",
                    "money_signed",
                ),
                (
                    "8.1 Margen proyectado contra 119.174M",
                    actual.get("pyme_threshold_remaining_crc"),
                    scenario.get("pyme_threshold_remaining_crc"),
                    "Margen o exceso estimado a diciembre.",
                    "money_signed",
                ),
            ]
            for label, actual_value, scenario_value, note, kind in rows:
                diff = float(scenario_value or 0) - float(actual_value or 0)
                tags = ("good",) if label.startswith("6.") and diff < 0 else (("warn",) if diff > 0 and (label.startswith("6.") or label.startswith("7.1")) else ())
                self.statement_tree.insert("", "end", values=(
                    company,
                    label,
                    self._format_statement_value(actual_value, kind),
                    self._format_statement_value(scenario_value, kind),
                    self._format_statement_value(diff, kind),
                    note,
                ), tags=tags)
            self.statement_tree.insert("", "end", values=("", "", "", "", "", ""), tags=("separator",))
        self.statement_tree.tag_configure("good", background="#e9f7ef")
        self.statement_tree.tag_configure("warn", background="#fff3cd")
        self.statement_tree.tag_configure("separator", background="#f2f2f2")

    def _exemption_amount(self, row):
        return float(row.get("base_income_tax_crc") or 0) * float(row.get("pyme_exemption_rate") or 0)

    def _format_statement_value(self, value, kind):
        if kind == "pct":
            return _pct(value)
        if kind == "money_signed":
            amount = float(value or 0)
            return _money(amount) if amount >= 0 else f"({ _money(abs(amount)) })"
        return _money(value)

    def _statement_export_rows(self):
        baseline = (self.result or {}).get("baseline") or {}
        optimized = (self.result or {}).get("optimized") or {}
        baseline_map = {row.get("company_code"): row for row in baseline.get("companies") or []}
        optimized_map = {row.get("company_code"): row for row in optimized.get("companies") or []}
        rows = []
        for company in sorted(set(baseline_map) | set(optimized_map)):
            actual = baseline_map.get(company) or {}
            scenario = optimized_map.get(company) or {}
            concepts = [
                ("1. Total ingresos / renta bruta real 2026", actual.get("gross_ytd_crc"), scenario.get("gross_ytd_crc"), "Ventas reales posteadas a hoy desde ingresos contables.", "money"),
                ("1.1 Proyeccion renta bruta a diciembre", actual.get("gross_projected_crc"), scenario.get("gross_projected_crc"), "Proyeccion gerencial para anticipar PYME al cierre.", "money"),
                ("2. Costos, gastos y deducciones", actual.get("deductible_expenses_projected_crc"), scenario.get("deductible_expenses_projected_crc"), "Gastos POSTED y pendientes visibles en la pestaña Gastos.", "money"),
                ("3. Renta neta antes de exoneraciones", actual.get("net_taxable_projected_crc"), scenario.get("net_taxable_projected_crc"), "Base fiscal antes de beneficio PYME.", "money"),
                ("4. Impuesto sobre utilidades base", actual.get("base_income_tax_crc"), scenario.get("base_income_tax_crc"), "Tarifa general o tramos D-102.", "money"),
                ("5. Exoneracion PYME aplicada", self._exemption_amount(actual), self._exemption_amount(scenario), "Monto rebajado por año PYME.", "money"),
                ("6. Impuesto final proyectado", actual.get("income_tax_projected_crc"), scenario.get("income_tax_projected_crc"), "Estimado final a pagar.", "money"),
                ("7. Uso real del umbral PYME a hoy", actual.get("pyme_threshold_ytd_usage_pct"), scenario.get("pyme_threshold_ytd_usage_pct"), "Foto real 2026: sobre 100% ya estaria excedido a hoy.", "pct"),
                ("7.1 Uso proyectado del umbral PYME", actual.get("pyme_threshold_usage_pct"), scenario.get("pyme_threshold_usage_pct"), "Proyeccion a diciembre: sobre 100% se perderia PYME al cierre.", "pct"),
                ("8. Margen real contra 119.174M", actual.get("pyme_threshold_ytd_remaining_crc"), scenario.get("pyme_threshold_ytd_remaining_crc"), "Margen disponible real a hoy antes de exceder el umbral.", "money_signed"),
                ("8.1 Margen proyectado contra 119.174M", actual.get("pyme_threshold_remaining_crc"), scenario.get("pyme_threshold_remaining_crc"), "Margen o exceso estimado a diciembre.", "money_signed"),
            ]
            for label, actual_value, scenario_value, note, kind in concepts:
                diff = float(scenario_value or 0) - float(actual_value or 0)
                rows.append([
                    company,
                    label,
                    self._format_statement_value(actual_value, kind),
                    self._format_statement_value(scenario_value, kind),
                    self._format_statement_value(diff, kind),
                    note,
                ])
        return rows

    def _render_analysis(self, data):
        baseline = data.get("baseline") or {}
        optimized = data.get("optimized") or {}
        analysis = data.get("analysis") or {}
        warnings = []
        warnings.extend(baseline.get("warnings") or [])
        warnings.extend(optimized.get("warnings") or [])
        lines = [
            "Resumen ejecutivo",
            analysis.get("recommendation") or "Sin recomendacion disponible.",
            "",
            f"Renta actual proyectada: {_money(baseline.get('total_projected_tax_crc'))}",
            f"Renta escenario proyectada: {_money(optimized.get('total_projected_tax_crc'))}",
            f"Impacto fiscal: {_money(analysis.get('tax_saving_crc'))}",
            "",
            "Lectura por sociedad",
        ]
        for row in optimized.get("companies") or []:
            remaining = float(row.get("pyme_threshold_remaining_crc") or 0)
            room = _money(remaining) if remaining >= 0 else f"Exceso {_money(abs(remaining))}"
            ytd_remaining = float(row.get("pyme_threshold_ytd_remaining_crc") or 0)
            ytd_room = _money(ytd_remaining) if ytd_remaining >= 0 else f"Exceso {_money(abs(ytd_remaining))}"
            lines.append(
                f"- {row.get('company_code')}: venta real 2026 {_money(row.get('gross_ytd_crc'))} "
                f"(uso real {_pct(row.get('pyme_threshold_ytd_usage_pct'))}, margen real {ytd_room}); "
                f"proyeccion diciembre {_money(row.get('gross_projected_crc'))}, "
                f"gastos {_money(row.get('deductible_expenses_projected_crc'))}, renta neta {_money(row.get('net_taxable_projected_crc'))}, "
                f"uso PYME proyectado {_pct(row.get('pyme_threshold_usage_pct'))}, margen/exceso proyectado {room}."
            )
        lines.extend(["", "Pros:", *[f"- {item}" for item in analysis.get("pros") or []]])
        lines.extend(["", "Contras / cuidados:", *[f"- {item}" for item in analysis.get("cons") or []]])
        if optimized.get("moved_clients"):
            lines.append("")
            lines.append("Clientes movidos en el escenario:")
            for item in optimized.get("moved_clients") or []:
                lines.append(
                    f"- {item.get('client_name')}: {item.get('from_company')} -> {item.get('to_company')} "
                    f"solo venta futura estimada por {_money(item.get('projected_amount_crc'))}; "
                    f"venta real ya facturada se mantiene por {_money(item.get('ytd_kept_crc'))}."
                )
        if optimized.get("moved_expenses"):
            lines.append("")
            lines.append("Gastos movidos en el escenario:")
            for item in optimized.get("moved_expenses") or []:
                lines.append(f"- {item.get('account_code')} {item.get('account_name') or ''}: {item.get('from_company')} -> {item.get('to_company')} por {_money(item.get('projected_amount_crc'))}.")
        if warnings:
            lines.extend(["", "Alertas:", *[f"- {item}" for item in warnings]])
        lines.extend(["", data.get("disclaimer") or ""])
        self.analysis_text.insert("1.0", "\n".join(lines))

    def _selected_client_row(self):
        selected = self.clients_tree.selection()
        if not selected:
            messagebox.showwarning("Ventas", "Selecciona un cliente.", parent=self)
            return None
        return self.client_rows_by_item.get(selected[0])

    def _on_client_select(self, _event=None):
        row = self._selected_client_row_silent()
        if not row:
            return
        mode = "monto fijo" if row.get("projection_mode") == "FIXED" else "proyeccion normal"
        self.sales_hint_var.set(
            f"Seleccionado: {row.get('client_name')} | real {_money(row.get('ytd_amount_crc'))} | "
            f"futuro movible {_money(row.get('future_projected_crc'))} | {mode}. "
            f"Mover venta solo reasigna lo futuro."
        )

    def _selected_client_row_silent(self):
        selected = self.clients_tree.selection()
        return self.client_rows_by_item.get(selected[0]) if selected else None

    def _selected_expense_row(self):
        selected = self.expenses_tree.selection()
        if not selected:
            messagebox.showwarning("Gastos", "Selecciona un gasto.", parent=self)
            return None
        return self.expense_rows_by_item.get(selected[0])

    def _on_expense_select(self, _event=None):
        row = self._selected_expense_row_silent()
        if not row:
            return
        mode = "monto fijo" if row.get("projection_mode") == "FIXED" else "proyeccion normal"
        self.expense_hint_var.set(
            f"Seleccionado: {row.get('account_code')} {row.get('account_name')} | {row.get('company_code')} | "
            f"Fuente {row.get('source_type')} | {mode}: {_money(row.get('projected_annual_crc'))}."
        )

    def _selected_expense_row_silent(self):
        selected = self.expenses_tree.selection()
        return self.expense_rows_by_item.get(selected[0]) if selected else None

    def _move_selected_client_to_target(self):
        row = self._selected_client_row()
        if row:
            self._upsert_client_move(row, self.target_company_var.get().strip() or "MMS-CR")
            self._run()

    def _move_selected_client_to_source(self):
        row = self._selected_client_row()
        if row:
            self._upsert_client_move(row, self.source_company_var.get().strip() or "MSL-CR")
            self._run()

    def _lock_selected_client_projection(self):
        row = self._selected_client_row()
        if row:
            self._upsert_client_lock(row)
            self._run()

    def _unlock_selected_client_projection(self):
        row = self._selected_client_row()
        if row:
            self._remove_client_lock(row)
            self._run()

    def _upsert_client_lock(self, row):
        key = (row.get("company_code"), row.get("client_name"))
        self.fixed_clients = [
            item for item in self.fixed_clients
            if not (item.get("company_code") == key[0] and item.get("client_name") == key[1])
        ]
        self.fixed_clients.append({"company_code": key[0], "client_name": key[1]})

    def _remove_client_lock(self, row):
        key = (row.get("company_code"), row.get("client_name"))
        self.fixed_clients = [
            item for item in self.fixed_clients
            if not (item.get("company_code") == key[0] and item.get("client_name") == key[1])
        ]

    def _upsert_client_move(self, row, to_company):
        from_company = row.get("company_code")
        self.client_moves = [
            item for item in self.client_moves
            if not (item.get("from_company") == from_company and item.get("client_name") == row.get("client_name"))
        ]
        future_amount = max(float(row.get("future_projected_crc") or 0), 0)
        self.client_moves.append({
            "client_name": row.get("client_name"),
            "from_company": from_company,
            "to_company": to_company,
            "projected_amount_crc": 0 if to_company == from_company else future_amount,
        })

    def _move_selected_expense_to_target(self):
        row = self._selected_expense_row()
        if row:
            self._upsert_expense_move(row, self.target_company_var.get().strip() or "MMS-CR")
            self._run()

    def _move_selected_expense_to_source(self):
        row = self._selected_expense_row()
        if row:
            self._upsert_expense_move(row, self.source_company_var.get().strip() or "MSL-CR")
            self._run()

    def _upsert_expense_move(self, row, to_company):
        from_company = row.get("company_code")
        self.expense_moves = [
            item for item in self.expense_moves
            if not (
                item.get("from_company") == from_company
                and item.get("account_code") == row.get("account_code")
                and item.get("source_type") == row.get("source_type")
                and item.get("account_name") == row.get("account_name")
            )
        ]
        if to_company != from_company:
            self.expense_moves.append({
                "account_code": row.get("account_code"),
                "account_name": row.get("account_name"),
                "source_type": row.get("source_type"),
                "from_company": from_company,
                "to_company": to_company,
                "projected_amount_crc": row.get("projected_annual_crc"),
            })

    def _lock_selected_expense_projection(self):
        row = self._selected_expense_row()
        if row:
            self._upsert_expense_lock(row)
            self._run()

    def _unlock_selected_expense_projection(self):
        row = self._selected_expense_row()
        if row:
            self._remove_expense_lock(row)
            self._run()

    def _upsert_expense_lock(self, row):
        key = (row.get("company_code"), row.get("source_type"), row.get("account_code"), row.get("account_name"))
        self.fixed_expenses = [
            item for item in self.fixed_expenses
            if not (
                item.get("company_code") == key[0]
                and item.get("source_type") == key[1]
                and item.get("account_code") == key[2]
                and item.get("account_name") == key[3]
            )
        ]
        self.fixed_expenses.append({
            "company_code": key[0],
            "source_type": key[1],
            "account_code": key[2],
            "account_name": key[3],
        })

    def _remove_expense_lock(self, row):
        key = (row.get("company_code"), row.get("source_type"), row.get("account_code"), row.get("account_name"))
        self.fixed_expenses = [
            item for item in self.fixed_expenses
            if not (
                item.get("company_code") == key[0]
                and item.get("source_type") == key[1]
                and item.get("account_code") == key[2]
                and item.get("account_name") == key[3]
            )
        ]

    def _apply_auto_moves(self):
        if not self.result:
            self._run()
            return
        self.client_moves = list(self.result.get("auto_moves") or [])
        self.status_var.set("Recomendacion IA aplicada. Recalculando...")
        self._run()

    def _export_excel(self):
        if not self.result:
            messagebox.showwarning("Exportar", "Primero ejecuta el analisis.", parent=self)
            return
        path = filedialog.asksaveasfilename(
            parent=self,
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=f"simulador_fiscal_{self.year_var.get()}_{int(self.month_var.get() or 1):02d}.xlsx",
        )
        if not path:
            return
        try:
            self._write_excel(path)
            self.status_var.set(f"Excel exportado: {path}")
        except Exception as exc:
            messagebox.showerror("Exportar", f"No se pudo exportar:\n{exc}", parent=self)

    def _write_excel(self, path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Resumen"
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(color="FFFFFF", bold=True)

        def write_table(sheet, title, headers, rows, start_row=1):
            sheet.cell(start_row, 1, title).font = Font(bold=True, size=13)
            row_idx = start_row + 1
            for col_idx, header in enumerate(headers, 1):
                cell = sheet.cell(row_idx, col_idx, header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            for data_row in rows:
                row_idx += 1
                for col_idx, value in enumerate(data_row, 1):
                    sheet.cell(row_idx, col_idx, value)
            for column_cells in sheet.columns:
                sheet.column_dimensions[column_cells[0].column_letter].width = min(max(len(str(c.value or "")) for c in column_cells) + 2, 45)
            return row_idx + 3

        summary_rows = self._statement_export_rows()
        write_table(
            ws,
            "Estado de resultados y renta tipo D-102",
            ["Empresa", "Linea", "Actual", "Escenario", "Diferencia", "Lectura"],
            summary_rows,
        )

        ws2 = wb.create_sheet("Ventas")
        sales_rows = [[r.get("company_code"), r.get("client_name"), r.get("invoice_count"), r.get("ytd_amount_crc"), r.get("projected_annual_crc"), r.get("projection_mode")] for r in self.result.get("clients") or []]
        write_table(ws2, "Ventas por cliente", ["Empresa", "Cliente", "Facturas", "YTD", "Proyeccion anual", "Modo"], sales_rows)

        ws3 = wb.create_sheet("Gastos")
        expense_rows = [
            [
                r.get("company_code"),
                r.get("source_type"),
                r.get("account_code"),
                r.get("account_name"),
                r.get("status"),
                r.get("entry_count"),
                r.get("ytd_amount_crc"),
                r.get("projected_annual_crc"),
                r.get("projection_mode"),
            ]
            for r in self.result.get("expense_rows") or []
        ]
        write_table(ws3, "Gastos deducibles por cuenta", ["Empresa", "Fuente", "Cuenta", "Gasto", "Estado", "Asientos", "YTD", "Proyeccion anual", "Modo"], expense_rows)

        ws4 = wb.create_sheet("Movimientos")
        move_rows = []
        for item in (self.result.get("optimized") or {}).get("moved_clients") or []:
            move_rows.append(["Venta futura", item.get("client_name"), item.get("from_company"), item.get("to_company"), item.get("projected_amount_crc")])
        for item in (self.result.get("optimized") or {}).get("moved_expenses") or []:
            move_rows.append(["Gasto", f"{item.get('account_code')} {item.get('account_name') or ''}", item.get("from_company"), item.get("to_company"), item.get("projected_amount_crc")])
        write_table(ws4, "Reasignaciones del escenario", ["Tipo", "Detalle", "Desde", "Hacia", "Monto anual"], move_rows)

        ws5 = wb.create_sheet("Analisis")
        for idx, line in enumerate(self.analysis_text.get("1.0", "end").splitlines(), 1):
            ws5.cell(idx, 1, line)
        ws5.column_dimensions["A"].width = 120
        wb.save(path)

    def _error(self, exc):
        self.status_var.set("No se pudo completar el simulador.")
        messagebox.showerror("Simulador fiscal", f"No se pudo analizar:\n{exc}", parent=self)

    def _copy_analysis(self):
        text = self.analysis_text.get("1.0", "end").strip()
        if not text:
            messagebox.showwarning("Simulador fiscal", "No hay analisis para copiar.", parent=self)
            return
        self.clipboard_clear()
        self.clipboard_append(text)
        self.status_var.set("Analisis copiado al portapapeles.")

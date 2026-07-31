# -*- coding: utf-8 -*-
import tkinter as tk
from datetime import date
from tkinter import ttk, filedialog, messagebox

from api_client import get_accounting_complete_financial_statements_api


PURPLE = "#312f91"
TEXT = "#3f3f46"
MUTED = "#6f6f76"
BG = "#f7f7f8"
CARD = "#ffffff"
BORDER = "#e6e6ea"
PANEL = "#f4f6fb"
INPUT_BG = "#ffffff"


class PopupD102(tk.Toplevel):
    """Formulario visual 102 - Impuesto sobre las utilidades ISU - PJ."""

    COMPANY_ID = "3102920372"
    COMPANY_NAME = "MSL MARINE SURVEYORS AND LOGISTICS GROUP\nSOCIEDAD DE RESPONSABILIDAD LIMITADA"

    INCOME_FIELDS = [
        "Venta de bienes y servicios (excepto servicios profesionales)",
        "Servicios profesionales y honorarios",
        "Comisiones",
        "Intereses y rendimientos financieros",
        "Dividendos y participaciones",
        "Alquileres",
        "Otros ingresos diferentes a los anteriores",
        "Exclusiones de la renta bruta",
    ]
    COST_FIELDS = [
        "Inventario inicial",
        "Compras",
        "Inventario final",
        "Costo de ventas",
        "Intereses y gastos financieros",
        "Gastos de ventas y administrativos",
        "Depreciaciones, amortización y agotamiento",
        "Donaciones a entidades autorizadas",
        "Deducción de pérdidas diferidas",
        "IVA soportado no acreditable",
        "Otros costos, gastos y deducciones permitidos por ley",
        "Gastos No deducibles",
    ]
    TAX_BRACKETS = [
        ("Tramo 1\n0,00 - 5.687.000,00", 0.0, 5687000.0, 5),
        ("Tramo 2\n5.687.001,00 - 8.532.000,00", 5687000.0, 8532000.0, 10),
        ("Tramo 3\n8.532.001,00 - 11.376.000,00", 8532000.0, 11376000.0, 15),
        ("Tramo 4\n11.376.001,00 en adelante", 11376000.0, None, 20),
    ]

    def __init__(self, parent, period="202401"):
        super().__init__(parent)
        self.source_period = period or date.today().strftime("%Y-%m")
        self.fiscal_year = self._year_from_period(self.source_period)
        self.period = str(self.fiscal_year)
        self.is_preliminary = self.fiscal_year >= date.today().year
        self.step = 0
        self.title("102 - Impuesto sobre las utilidades ISU - PJ")
        self.geometry("1320x720")
        self.minsize(1120, 640)
        self.configure(bg=BG)

        self.values = {
            "total_activos": 0.0,
            "total_pasivos": 0.0,
            "total_exoneraciones": 0.0,
            "reduccion_impuesto": 0.0,
            "creditos_zona_franca": 0.0,
        }
        self.income_values = {name: 0.0 for name in self.INCOME_FIELDS}
        self.cost_values = {name: 0.0 for name in self.COST_FIELDS}
        self.automation_notes = []
        self.financial_payload = {}

        self.summary_vars = {}
        self.step_labels = []
        self.main_card = None

        self._load_accounting_data()
        self._build_shell()
        self._render_step()
        self._refresh_summary()

    def _load_accounting_data(self):
        """Autollenado con asientos POSTED del año fiscal seleccionado."""
        period_from = f"{self.fiscal_year}-01"
        period_to = date.today().strftime("%Y-%m") if self.is_preliminary else f"{self.fiscal_year}-12"
        try:
            payload = get_accounting_complete_financial_statements_api(
                period_from=period_from,
                period_to=period_to,
                limit=5000,
            )
            self.financial_payload = payload or {}
            balance = self.financial_payload.get("balance_sheet") or {}
            income = self.financial_payload.get("income_statement") or {}
            balance_totals = balance.get("totals") or {}
            income_totals = income.get("totals") or {}

            self.values["total_activos"] = float(balance_totals.get("assets") or 0)
            self.values["total_pasivos"] = float(balance_totals.get("liabilities") or 0)
            self._classify_revenue_lines(income.get("revenue") or [])
            self._classify_expense_lines(income.get("expenses") or [])

            revenue = float(income_totals.get("revenue") or 0)
            expenses = float(income_totals.get("expenses") or 0)
            if revenue and not any(self.income_values.values()):
                self.income_values["Venta de bienes y servicios (excepto servicios profesionales)"] = revenue
            if expenses and not any(self.cost_values.values()):
                self.cost_values["Otros costos, gastos y deducciones permitidos por ley"] = expenses
            self.automation_notes.append(f"Autollenado desde Accounting POSTED: {period_from} a {period_to}.")
        except Exception as exc:
            self.automation_notes.append(f"Autollenado no disponible: {exc}")

    def _classify_revenue_lines(self, rows):
        for row in rows:
            name = (row.get("account_name") or "").lower()
            amount = float(row.get("balance") or 0)
            if not amount:
                continue
            if "honor" in name or "profes" in name:
                key = "Servicios profesionales y honorarios"
            elif "comision" in name or "comisi" in name:
                key = "Comisiones"
            elif "inter" in name or "rendimiento" in name:
                key = "Intereses y rendimientos financieros"
            elif "dividend" in name or "particip" in name:
                key = "Dividendos y participaciones"
            elif "alquiler" in name or "renta" in name:
                key = "Alquileres"
            else:
                key = "Venta de bienes y servicios (excepto servicios profesionales)"
            self.income_values[key] += amount

    def _classify_expense_lines(self, rows):
        for row in rows:
            name = (row.get("account_name") or "").lower()
            amount = float(row.get("balance") or 0)
            if not amount:
                continue
            if "compra" in name or "costo" in name:
                key = "Compras"
            elif "inter" in name or "financ" in name:
                key = "Intereses y gastos financieros"
            elif "deprec" in name or "amort" in name or "agot" in name:
                key = "Depreciaciones, amortización y agotamiento"
            elif "donaci" in name:
                key = "Donaciones a entidades autorizadas"
            elif "perdida" in name or "pérdida" in name:
                key = "Deducción de pérdidas diferidas"
            elif "iva" in name and ("gasto" in name or "no acredit" in name):
                key = "IVA soportado no acreditable"
            elif "multa" in name or "sancion" in name or "sanción" in name or "no deduc" in name:
                key = "Gastos No deducibles"
            elif "sueldo" in name or "salario" in name or "admin" in name or "venta" in name:
                key = "Gastos de ventas y administrativos"
            else:
                key = "Otros costos, gastos y deducciones permitidos por ley"
            self.cost_values[key] += amount

    # ---------------- UI shell ----------------
    def _build_shell(self):
        header = tk.Frame(self, bg=BG)
        header.pack(fill="x", padx=14, pady=(12, 18))
        tk.Button(
            header,
            text="<",
            font=("Segoe UI", 18, "bold"),
            width=2,
            bd=0,
            bg="#ededee",
            fg=TEXT,
            activebackground="#e5e5e7",
            command=self.destroy,
        ).pack(side="left")
        tk.Label(
            header,
            text="102 - Impuesto sobre las utilidades ISU - PJ",
            font=("Segoe UI", 18, "bold"),
            bg=BG,
            fg=TEXT,
        ).pack(side="left", padx=14)
        tk.Label(
            header,
            text="Renta anual automatizada con revision fiscal",
            font=("Segoe UI", 10),
            bg=BG,
            fg=MUTED,
        ).pack(side="left", padx=(0, 14))
        ttk.Button(header, text="Descargar Excel", command=self._export_excel).pack(side="right", padx=(8, 0))
        tk.Label(
            header,
            text=self._period_status_text(),
            font=("Segoe UI", 10, "bold"),
            bg=BG,
            fg="#d9435f" if self.is_preliminary else MUTED,
        ).pack(side="right", padx=(0, 14))

        info = tk.Frame(self, bg="#fff8e8" if self.is_preliminary else "#eef4ff", padx=12, pady=7)
        info.pack(fill="x", padx=14, pady=(0, 8))
        tk.Label(
            info,
            text=(
                "Autollenado Renta: activos, pasivos, ingresos y gastos desde estados financieros POSTED. "
                "Exoneraciones, reducciones y creditos quedan editables por requerir soporte fiscal."
            ),
            bg="#fff8e8" if self.is_preliminary else "#eef4ff",
            fg="#704a00" if self.is_preliminary else "#254064",
            font=("Segoe UI", 9, "bold"),
        ).pack(anchor="w")

        body = tk.Frame(self, bg=BG)
        body.pack(fill="both", expand=True)

        self.sidebar = tk.Frame(body, bg=BG, width=190)
        self.sidebar.pack(side="left", fill="y", padx=(0, 10))
        self.sidebar.pack_propagate(False)
        self._build_stepper()

        self.content = tk.Frame(body, bg=BG)
        self.content.pack(side="left", fill="both", expand=True)

        self.summary = tk.Frame(body, bg=BG, width=320)
        self.summary.pack(side="right", fill="y", padx=(14, 18))
        self.summary.pack_propagate(False)
        self._build_summary()

    def _build_stepper(self):
        steps = [
            "Activos y pasivos",
            "Cálculo de la renta\nneta: Ingresos y\ncostos",
            "Impuesto sobre las\nutilidades",
            "Determinación del\nimpuesto",
        ]
        self.step_labels = []
        for index, text in enumerate(steps):
            row = tk.Frame(self.sidebar, bg=BG)
            row.pack(fill="x", pady=(0, 10), padx=(6, 0))
            dot = tk.Label(
                row,
                text=str(index + 1),
                bg="#e8e8ee",
                fg=MUTED,
                font=("Segoe UI", 9, "bold"),
                width=3,
                pady=4,
            )
            dot.pack(side="left", anchor="n")
            label = tk.Label(
                row,
                text=text,
                bg=BG,
                fg=MUTED,
                justify="left",
                font=("Segoe UI", 10, "bold" if index == 0 else "normal"),
            )
            label.pack(side="left", padx=(8, 0), anchor="w")
            self.step_labels.append((dot, label))

    def _build_summary(self):
        card = self._card(self.summary, padx=18, pady=18)
        card.pack(fill="both", expand=True, pady=(14, 30))
        tk.Label(card, text="Resumen", font=("Segoe UI", 14, "bold"), bg=CARD, fg=TEXT).pack(anchor="w")

        info = tk.Frame(card, bg=CARD)
        info.pack(fill="x", pady=(20, 8))
        self._summary_text(info, "Identificación", self.COMPANY_ID)
        self._summary_text(info, "Nombre", self.COMPANY_NAME)
        self._summary_text(info, "Periodo", self.period)
        self._summary_text(info, "Declaración", "102 - Impuesto sobre las utilidades ISU - PJ")

        dates = tk.Frame(info, bg=CARD)
        dates.pack(fill="x", pady=(10, 0))
        self._summary_text(dates, "Fecha inicio", self._year_start(), side="left")
        self._summary_text(dates, "Fecha fin", self._year_end(), side="left")
        tk.Label(dates, text=self._period_status_text(), bg=CARD, fg="#d9435f" if self.is_preliminary else MUTED, font=("Segoe UI", 9)).pack(
            side="left", padx=(14, 0), anchor="n"
        )

        tk.Label(card, text=". " * 34, bg=CARD, fg="#dddddf").pack(fill="x", pady=(10, 8))

        for label in [
            "Total patrimonio",
            "Renta neta",
            "Renta neta después de\nexoneraciones",
            "Total impuesto\ndeterminado",
        ]:
            var = tk.StringVar(value="0,00")
            self.summary_vars[label] = var
            row = tk.Frame(card, bg=CARD)
            row.pack(fill="x", pady=7)
            tk.Label(row, text=label, bg=CARD, fg=MUTED, justify="left", font=("Segoe UI", 10)).pack(side="left")
            tk.Label(row, textvariable=var, bg=CARD, fg=TEXT, font=("Segoe UI", 10)).pack(side="right")

    def _summary_text(self, parent, title, value, side=None):
        box = tk.Frame(parent, bg=CARD)
        pack_args = {"anchor": "w", "pady": (0, 9)}
        if side:
            pack_args = {"side": side, "anchor": "n", "padx": (0, 6)}
        box.pack(**pack_args)
        tk.Label(box, text=title, bg=CARD, fg="#8a8a8f", font=("Segoe UI", 8)).pack(anchor="w")
        tk.Label(box, text=value, bg=CARD, fg=MUTED, justify="left", font=("Segoe UI", 9)).pack(anchor="w")

    # ---------------- steps ----------------
    def _render_step(self):
        for widget in self.content.winfo_children():
            widget.destroy()
        for index, (dot, label) in enumerate(self.step_labels):
            active = index == self.step
            dot.configure(bg=PURPLE if active else "#e8e8ee", fg="white" if active else MUTED)
            label.configure(fg=PURPLE if active else MUTED, font=("Segoe UI", 10, "bold" if active else "normal"))

        self.main_card = self._card(self.content, padx=34, pady=26)
        self.main_card.pack(fill="both", expand=True, pady=(10, 30))
        if self.step == 0:
            self._step_assets()
        elif self.step == 1:
            self._step_income_costs()
        elif self.step == 2:
            self._step_tax()
        else:
            self._step_determination()
        self._refresh_summary()

    def _step_assets(self):
        tk.Label(self.main_card, text="Activos y pasivos", font=("Segoe UI", 18, "bold"), bg=CARD, fg=TEXT).pack(anchor="w")
        for title, key, editable in [
            ("Total activos", "total_activos", True),
            ("Total pasivos", "total_pasivos", True),
            ("Total patrimonio", "patrimonio", False),
        ]:
            self._amount_line(title, key, editable, pady=(30 if key == "total_activos" else 18, 0))
        self._nav_buttons(show_back=False)

    def _step_income_costs(self):
        tk.Label(
            self.main_card,
            text="Cálculo de la renta neta: Ingresos y costos",
            font=("Segoe UI", 18, "bold"),
            bg=CARD,
            fg=TEXT,
        ).pack(anchor="w")
        self._divider(self.main_card)
        self._amount_line("Total ingresos o renta bruta", "income_total", True, self._open_income_modal, pady=(30, 0))
        self._amount_line("Total costos, gastos y deducciones", "cost_total", True, self._open_cost_modal, pady=(22, 0))
        self._amount_line("Renta neta", "renta_neta", False, pady=(32, 0))
        self._amount_line("Pérdida neta", "perdida_neta", False, pady=(22, 0))
        self._nav_buttons()

    def _step_tax(self):
        tk.Label(self.main_card, text="Impuesto sobre las utilidades", font=("Segoe UI", 18, "bold"), bg=CARD, fg=TEXT).pack(anchor="w")
        self._divider(self.main_card)
        self._amount_line("Total exoneraciones", "total_exoneraciones", True, pady=(24, 0))
        self._amount_line("Renta neta después de exoneraciones", "renta_after_exonerations", False, pady=(18, 0))
        tk.Label(
            self.main_card,
            text="El formulario de forma automática con la renta neta calcula los importes de cada uno de los tramos.",
            bg=CARD,
            fg=MUTED,
            font=("Segoe UI", 11),
        ).pack(anchor="w", pady=(18, 18))
        for text, base, rate, tax in self._tax_brackets():
            row = tk.Frame(self.main_card, bg=CARD)
            row.pack(fill="x", padx=12, pady=2)
            tk.Label(row, text=text, bg=CARD, fg=MUTED, font=("Segoe UI", 10), width=36, anchor="w", justify="left").pack(side="left")
            tk.Label(row, text=self._fmt(base), bg=CARD, fg=MUTED, font=("Segoe UI", 10), width=18, anchor="e").pack(side="left")
            tk.Label(row, text=f"{rate}%", bg=CARD, fg=MUTED, font=("Segoe UI", 10), width=12, anchor="e").pack(side="left")
            tk.Label(row, text=self._fmt(tax), bg=CARD, fg=MUTED, font=("Segoe UI", 10), width=18, anchor="e").pack(side="left")
        self._amount_line("Total impuesto sobre las utilidades", "tax_total", False, pady=(24, 0))
        self._amount_line("Reducción del impuesto para las micro y pequeñas empresas", "reduccion_impuesto", True, pady=(18, 0))
        self._amount_line("Impuesto sobre las utilidades después de\nreducciones", "tax_after_reduction", False, pady=(18, 0))
        self._nav_buttons()

    def _step_determination(self):
        tk.Label(self.main_card, text="Determinación del impuesto", font=("Segoe UI", 18, "bold"), bg=CARD, fg=TEXT).pack(anchor="w")
        self._divider(self.main_card)
        self._amount_line("Créditos Régimen de Zonas Francas y otros permitidos por ley", "creditos_zona_franca", True, pady=(30, 0))
        self._amount_line("Impuesto determinado", "impuesto_determinado", False, pady=(26, 0))
        self._nav_buttons(next_text="Presentar", next_command=self._present)

    # ---------------- reusable widgets ----------------
    def _amount_line(self, label, key, editable=False, command=None, pady=(12, 0)):
        row = tk.Frame(self.main_card, bg="#fbfbfd", padx=12, pady=10, highlightbackground=BORDER, highlightthickness=1)
        row.pack(fill="x", padx=12, pady=pady)
        tk.Label(row, text=label, bg="#fbfbfd", fg=TEXT, font=("Segoe UI", 10), justify="left", wraplength=540).pack(side="left")
        value = self._value_for_key(key)
        if editable:
            tk.Button(
                row,
                text="Editar",
                bd=0,
                bg=PANEL,
                fg=MUTED,
                font=("Segoe UI", 9),
                activebackground="#e9edf7",
                command=command or (lambda k=key: self._open_single_amount_modal(k)),
                padx=10,
                pady=4,
            ).pack(side="right", padx=(8, 0))
            self._value_box(row, value, key).pack(side="right")
        else:
            tk.Label(row, text=self._fmt(value), bg="#fbfbfd", fg=TEXT, font=("Segoe UI", 13, "bold")).pack(side="right")

    def _value_box(self, parent, value, key):
        frame = tk.Frame(parent, bg=INPUT_BG, highlightbackground=BORDER, highlightthickness=1, bd=0)
        tk.Label(frame, text=self._short_label(key), bg=INPUT_BG, fg=MUTED, font=("Segoe UI", 7)).pack(anchor="w", padx=12, pady=(6, 0))
        tk.Label(frame, text=self._fmt(value), bg=INPUT_BG, fg=TEXT, font=("Segoe UI", 11), width=11, anchor="e").pack(padx=12, pady=(0, 6))
        return frame

    def _nav_buttons(self, show_back=True, next_text="Siguiente  >", next_command=None):
        bar = tk.Frame(self.main_card, bg=CARD)
        bar.pack(side="bottom", fill="x", pady=(20, 6))
        if show_back:
            tk.Button(
                bar,
                text="<  Anterior",
                command=self._back,
                bg=CARD,
                fg=TEXT,
                relief="solid",
                bd=1,
                font=("Segoe UI", 11, "bold"),
                height=2,
            ).pack(side="left", fill="x", expand=True, padx=(0, 14))
        spacer = tk.Frame(bar, bg=CARD)
        spacer.pack(side="left", fill="x", expand=True)
        tk.Button(
            bar,
            text=next_text,
            command=next_command or self._next,
            bg=PURPLE,
            fg="white",
            activebackground="#282575",
            activeforeground="white",
            relief="flat",
            font=("Segoe UI", 11, "bold"),
            height=2,
        ).pack(side="right", fill="x", expand=True, padx=(14, 0))

    def _card(self, parent, padx=10, pady=10):
        return tk.Frame(parent, bg=CARD, padx=padx, pady=pady, highlightthickness=0)

    def _divider(self, parent):
        tk.Frame(parent, bg=BORDER, height=1).pack(fill="x", pady=(24, 10))

    # ---------------- modals ----------------
    def _open_single_amount_modal(self, key):
        self._open_amount_modal(self._short_label(key), [(key, self._short_label(key))], self.values)

    def _open_income_modal(self):
        self._open_amount_modal("Total ingresos o renta bruta", [(name, name) for name in self.INCOME_FIELDS], self.income_values)

    def _open_cost_modal(self):
        self._open_amount_modal("Total costos, gastos y deducciones", [(name, name) for name in self.COST_FIELDS], self.cost_values, scroll=True)

    def _open_amount_modal(self, title, rows, target, scroll=False):
        popup = tk.Toplevel(self)
        popup.title(title)
        popup.geometry("710x690" if scroll else "710x610")
        popup.configure(bg=CARD)
        popup.transient(self)
        popup.grab_set()

        header = tk.Frame(popup, bg=CARD)
        header.pack(fill="x", padx=20, pady=(18, 8))
        tk.Label(header, text=title, bg=CARD, fg=TEXT, font=("Segoe UI", 20, "bold")).pack(side="left")
        tk.Button(header, text="X", bd=0, bg=CARD, fg=TEXT, font=("Segoe UI", 18, "bold"), command=popup.destroy).pack(side="right")

        entries = {}
        form = self._scrollable_form(popup) if scroll else tk.Frame(popup, bg=CARD)
        if not scroll:
            form.pack(fill="both", expand=True, padx=20, pady=(8, 0))

        for key, label in rows:
            row = tk.Frame(form, bg=CARD)
            row.pack(fill="x", pady=10)
            tk.Label(row, text=label, bg=CARD, fg=TEXT, font=("Segoe UI", 11), wraplength=380, justify="left").pack(side="left", anchor="w")
            tk.Label(row, text="?", bg=CARD, fg=MUTED, font=("Segoe UI", 9)).pack(side="left", padx=6)
            var = tk.StringVar(value=self._fmt(target.get(key, 0.0)))
            tk.Entry(row, textvariable=var, width=24, relief="solid", bd=1, font=("Segoe UI", 10)).pack(side="right", padx=(14, 24), ipady=8)
            entries[key] = var

        footer = tk.Frame(popup, bg=CARD)
        footer.pack(fill="x", padx=28, pady=18)
        tk.Button(
            footer,
            text="Actualizar",
            bg=PURPLE,
            fg="white",
            activebackground="#282575",
            activeforeground="white",
            relief="flat",
            font=("Segoe UI", 11, "bold"),
            command=lambda: self._save_modal_values(popup, entries, target),
            width=18,
            height=2,
        ).pack(side="left")
        tk.Button(
            footer,
            text="Cancelar",
            bg=CARD,
            fg=TEXT,
            relief="solid",
            bd=1,
            font=("Segoe UI", 11, "bold"),
            command=popup.destroy,
            width=18,
            height=2,
        ).pack(side="left", padx=16)

    def _scrollable_form(self, popup):
        canvas = tk.Canvas(popup, bg=CARD, highlightthickness=0)
        sb = ttk.Scrollbar(popup, orient="vertical", command=canvas.yview)
        form = tk.Frame(canvas, bg=CARD)
        form.bind("<Configure>", lambda _e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=form, anchor="nw")
        canvas.configure(yscrollcommand=sb.set)
        canvas.pack(side="left", fill="both", expand=True, padx=(20, 0), pady=(8, 0))
        sb.pack(side="right", fill="y")
        return form

    def _save_modal_values(self, popup, entries, target):
        for key, var in entries.items():
            target[key] = self._parse(var.get())
        popup.destroy()
        self._render_step()

    # ---------------- calculations ----------------
    def _value_for_key(self, key):
        if key == "patrimonio":
            return self.values["total_activos"] - self.values["total_pasivos"]
        if key == "income_total":
            return self._income_total()
        if key == "cost_total":
            return self._cost_total()
        if key == "renta_neta":
            return max(self._income_total() - self._cost_total(), 0.0)
        if key == "perdida_neta":
            return max(self._cost_total() - self._income_total(), 0.0)
        if key == "renta_after_exonerations":
            return max(self._value_for_key("renta_neta") - self.values["total_exoneraciones"], 0.0)
        if key == "tax_total":
            return sum(item[3] for item in self._tax_brackets())
        if key == "tax_after_reduction":
            return max(self._value_for_key("tax_total") - self.values["reduccion_impuesto"], 0.0)
        if key == "impuesto_determinado":
            return max(self._value_for_key("tax_after_reduction") - self.values["creditos_zona_franca"], 0.0)
        return self.values.get(key, 0.0)

    def _income_total(self):
        total = sum(v for k, v in self.income_values.items() if k != "Exclusiones de la renta bruta")
        return max(total - self.income_values.get("Exclusiones de la renta bruta", 0.0), 0.0)

    def _cost_total(self):
        total = sum(v for k, v in self.cost_values.items() if k != "Gastos No deducibles")
        return max(total - self.cost_values.get("Gastos No deducibles", 0.0), 0.0)

    def _tax_brackets(self):
        taxable = self._value_for_key("renta_after_exonerations")
        rows = []
        for text, start, end, rate in self.TAX_BRACKETS:
            if taxable <= start:
                base = 0.0
            elif end is None:
                base = taxable - start
            else:
                base = min(taxable, end) - start
            base = max(base, 0.0)
            rows.append((text, base, rate, base * rate / 100.0))
        return rows

    def _refresh_summary(self):
        if not self.summary_vars:
            return
        self.summary_vars["Total patrimonio"].set(self._fmt(self._value_for_key("patrimonio")))
        self.summary_vars["Renta neta"].set(self._fmt(self._value_for_key("renta_neta")))
        self.summary_vars["Renta neta después de\nexoneraciones"].set(self._fmt(self._value_for_key("renta_after_exonerations")))
        self.summary_vars["Total impuesto\ndeterminado"].set(self._fmt(self._value_for_key("impuesto_determinado")))

    # ---------------- Excel ----------------
    def _export_excel(self):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
        except Exception as exc:
            messagebox.showerror("Declaración 102", f"No se pudo cargar openpyxl:\n{exc}", parent=self)
            return

        path = filedialog.asksaveasfilename(
            parent=self,
            title="Descargar D102 en Excel",
            defaultextension=".xlsx",
            initialfile=f"D102_ISU_PJ_{self.period.replace('-', '')}.xlsx",
            filetypes=[("Excel", "*.xlsx")],
        )
        if not path:
            return

        wb = Workbook()
        styles = {
            "title_fill": PatternFill("solid", fgColor="312F91"),
            "header_fill": PatternFill("solid", fgColor="E9E9F6"),
            "bold": Font(bold=True, color="3F3F46"),
            "white_bold": Font(bold=True, color="FFFFFF"),
            "money": '#,##0.00',
        }
        self._build_excel_summary(wb.active, styles)
        self._build_excel_dashboard(wb.create_sheet("Dashboard"), styles)
        self._build_excel_assets(wb.create_sheet("Activos y pasivos"), styles)
        self._build_excel_income(wb.create_sheet("Ingresos"), styles)
        self._build_excel_costs(wb.create_sheet("Costos y deducciones"), styles)
        self._build_excel_tax(wb.create_sheet("Impuesto"), styles)
        self._build_excel_source(wb.create_sheet("Fuente automatica"), styles)
        self._build_excel_review(wb.create_sheet("Revision"), styles)

        for sheet in wb.worksheets:
            self._autosize_excel(sheet, get_column_letter)
            sheet.freeze_panes = "A3"
            for row in sheet.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

        try:
            wb.save(path)
        except PermissionError:
            messagebox.showerror(
                "Declaración 102",
                "No se pudo guardar el Excel. Cierre el archivo si está abierto y vuelva a intentar.",
                parent=self,
            )
            return
        except Exception as exc:
            messagebox.showerror("Declaración 102", f"No se pudo guardar el Excel:\n{exc}", parent=self)
            return

        messagebox.showinfo("Declaración 102", "Excel D102 generado correctamente.", parent=self)

    def _build_excel_dashboard(self, ws, styles):
        self._excel_title(ws, 1, "Dashboard ejecutivo D102", styles, 4)
        self._excel_headers(ws, ["Indicador", "Monto", "Estado", "Comentario"], styles)
        rows = [
            ("Total patrimonio", self._value_for_key("patrimonio"), "Calculado", "Activos menos pasivos"),
            ("Ingresos renta bruta", self._income_total(), "Calculado", "Ingresos menos exclusiones"),
            ("Costos y deducciones", self._cost_total(), "Calculado", "Deducibles menos no deducibles"),
            ("Renta neta", self._value_for_key("renta_neta"), "Calculado", "Base antes de exoneraciones"),
            ("Impuesto determinado", self._value_for_key("impuesto_determinado"), "Preliminar" if self.is_preliminary else "Final", self._period_status_text()),
        ]
        for label, value, status, comment in rows:
            ws.append([label, value, status, comment])
            self._money_if_number(ws.cell(ws.max_row, 2), styles)

    def _build_excel_summary(self, ws, styles):
        ws.title = "D102 Resumen"
        self._excel_title(ws, 1, "Formulario 102 - Impuesto sobre las utilidades ISU - PJ", styles, 4)
        rows = [
            ("Identificación", self.COMPANY_ID),
            ("Nombre", self.COMPANY_NAME.replace("\n", " ")),
            ("Periodo fiscal", self.period),
            ("Periodo seleccionado en ERP", self.source_period),
            ("Declaración", "102 - Impuesto sobre las utilidades ISU - PJ"),
            ("Fecha inicio", self._year_start()),
            ("Fecha fin", self._year_end()),
            ("Estado", self._period_status_text()),
            ("Total patrimonio", self._value_for_key("patrimonio")),
            ("Renta neta", self._value_for_key("renta_neta")),
            ("Renta neta después de exoneraciones", self._value_for_key("renta_after_exonerations")),
            ("Total impuesto determinado", self._value_for_key("impuesto_determinado")),
        ]
        ws.append([])
        for label, value in rows:
            ws.append([label, value])
            ws.cell(ws.max_row, 1).font = styles["bold"]
            self._money_if_number(ws.cell(ws.max_row, 2), styles)

    def _build_excel_assets(self, ws, styles):
        self._excel_title(ws, 1, "Activos y pasivos", styles, 3)
        self._excel_headers(ws, ["Campo", "Monto", "Comentario"], styles)
        for label, key in [
            ("Total activos", "total_activos"),
            ("Total pasivos", "total_pasivos"),
            ("Total patrimonio", "patrimonio"),
        ]:
            ws.append([label, self._value_for_key(key), "Calculado" if key == "patrimonio" else "Editable"])
            self._money_if_number(ws.cell(ws.max_row, 2), styles)

    def _build_excel_income(self, ws, styles):
        self._excel_title(ws, 1, "Total ingresos o renta bruta", styles, 3)
        self._excel_headers(ws, ["Rubro", "Monto", "Tratamiento"], styles)
        for label in self.INCOME_FIELDS:
            treatment = "Resta de renta bruta" if label == "Exclusiones de la renta bruta" else "Suma"
            ws.append([label, self.income_values.get(label, 0.0), treatment])
            self._money_if_number(ws.cell(ws.max_row, 2), styles)
        self._excel_total_row(ws, "Total ingresos o renta bruta", self._income_total(), styles)

    def _build_excel_costs(self, ws, styles):
        self._excel_title(ws, 1, "Total costos, gastos y deducciones", styles, 3)
        self._excel_headers(ws, ["Rubro", "Monto", "Tratamiento"], styles)
        for label in self.COST_FIELDS:
            treatment = "No deducible / resta del total deducible" if label == "Gastos No deducibles" else "Suma deducible"
            ws.append([label, self.cost_values.get(label, 0.0), treatment])
            self._money_if_number(ws.cell(ws.max_row, 2), styles)
        self._excel_total_row(ws, "Total costos, gastos y deducciones", self._cost_total(), styles)

    def _build_excel_tax(self, ws, styles):
        self._excel_title(ws, 1, "Cálculo del impuesto sobre las utilidades", styles, 5)
        self._excel_headers(ws, ["Concepto", "Base", "Tarifa", "Impuesto", "Nota"], styles)
        for label, base, rate, amount in self._tax_brackets():
            ws.append([label.replace("\n", " "), base, rate / 100.0, amount, "Tramo automático"])
            ws.cell(ws.max_row, 2).number_format = styles["money"]
            ws.cell(ws.max_row, 3).number_format = "0%"
            ws.cell(ws.max_row, 4).number_format = styles["money"]
        for label, key in [
            ("Renta neta", "renta_neta"),
            ("Total exoneraciones", "total_exoneraciones"),
            ("Renta neta después de exoneraciones", "renta_after_exonerations"),
            ("Total impuesto sobre las utilidades", "tax_total"),
            ("Reducción del impuesto para micro y pequeñas empresas", "reduccion_impuesto"),
            ("Impuesto después de reducciones", "tax_after_reduction"),
            ("Créditos Régimen de Zonas Francas y otros permitidos por ley", "creditos_zona_franca"),
            ("Impuesto determinado", "impuesto_determinado"),
        ]:
            ws.append([label, self._value_for_key(key), None, None, "Resumen"])
            ws.cell(ws.max_row, 1).font = styles["bold"] if key == "impuesto_determinado" else Font(color="3F3F46")
            ws.cell(ws.max_row, 2).number_format = styles["money"]

    def _build_excel_review(self, ws, styles):
        self._excel_title(ws, 1, "Checklist de revisión antes de presentar", styles, 4)
        self._excel_headers(ws, ["Estado", "Punto de revisión", "Responsable", "Comentario"], styles)
        for item in [
            "Validar que los activos y pasivos coincidan con el balance final.",
            "Validar que ingresos y exclusiones coincidan con el mayor contable.",
            "Validar costos deducibles y gastos no deducibles con soporte.",
            "Confirmar exoneraciones, reducciones y créditos con documentación.",
            "Comparar impuesto determinado contra TRIBU-CR antes de presentar.",
        ]:
            ws.append(["Pendiente", item, "", ""])

    def _build_excel_source(self, ws, styles):
        self._excel_title(ws, 1, "Fuente automatica y reglas aplicadas", styles, 4)
        self._excel_headers(ws, ["Tipo", "Detalle", "Regla", "Comentario"], styles)
        for note in self.automation_notes or ["Sin notas de autollenado."]:
            ws.append(["Autollenado", note, "POSTED accounting entries", "Editable antes de presentar"])
        ws.append(["Renta", "Ingresos", "Cuentas familia 4", "Clasificadas por nombre de cuenta cuando aplica"])
        ws.append(["Renta", "Costos y gastos", "Cuentas familia 5-9", "Clasificadas por nombre de cuenta cuando aplica"])
        ws.append(["Renta", "Exoneraciones, reducciones y créditos", "Manual", "Requiere soporte fiscal específico"])
        ws.append(["Advertencia", self._period_status_text(), "Periodo fiscal", "Si es preliminar, el año fiscal no está cerrado"])

    @staticmethod
    def _excel_title(ws, row, text, styles, columns):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=columns)
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = styles["white_bold"]
        cell.fill = styles["title_fill"]

    @staticmethod
    def _excel_headers(ws, headers, styles):
        ws.append(headers)
        row = ws.max_row
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = styles["bold"]
            cell.fill = styles["header_fill"]

    def _excel_total_row(self, ws, label, value, styles):
        ws.append([label, value, "Calculado"])
        ws.cell(ws.max_row, 1).font = styles["bold"]
        ws.cell(ws.max_row, 2).font = styles["bold"]
        self._money_if_number(ws.cell(ws.max_row, 2), styles)

    @staticmethod
    def _money_if_number(cell, styles):
        if isinstance(cell.value, (int, float)):
            cell.number_format = styles["money"]

    @staticmethod
    def _autosize_excel(sheet, get_column_letter):
        for column_cells in sheet.columns:
            width = 12
            col = get_column_letter(column_cells[0].column)
            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                width = max(width, min(len(value) + 2, 55))
            sheet.column_dimensions[col].width = width

    # ---------------- navigation/helpers ----------------
    def _next(self):
        self.step = min(self.step + 1, 3)
        self._render_step()

    def _back(self):
        self.step = max(self.step - 1, 0)
        self._render_step()

    def _present(self):
        messagebox.showinfo(
            "Declaración 102",
            "Formulario 102 preparado visualmente. Si el año fiscal no ha terminado, este resultado es preliminar. Use Descargar Excel para revisar y manipular el detalle antes de presentar en TRIBU-CR.",
            parent=self,
        )

    def _short_label(self, key):
        labels = {
            "total_activos": "Total activos",
            "total_pasivos": "Total pasivos",
            "total_exoneraciones": "Total exoneraciones",
            "reduccion_impuesto": "Reducción del impuesto",
            "creditos_zona_franca": "Créditos Régimen de Zo...",
            "income_total": "Total ingresos o r...",
            "cost_total": "Total costos, gast...",
        }
        return labels.get(key, key)

    @staticmethod
    def _year_from_period(period):
        try:
            return int(str(period).split("-")[0])
        except Exception:
            return date.today().year

    def _year_start(self):
        return f"01/01/{self.fiscal_year}"

    def _year_end(self):
        today = date.today()
        if self.is_preliminary:
            return f"{today.day:02d}/{today.month:02d}/{today.year}"
        return f"31/12/{self.fiscal_year}"

    def _period_status_text(self):
        if self.is_preliminary:
            return "Preliminar: año fiscal no terminado"
        return "Año fiscal cerrado"

    @staticmethod
    def _parse(text):
        if text is None:
            return 0.0
        clean = str(text).strip().replace("₡", "").replace(" ", "")
        if not clean:
            return 0.0
        if "," in clean and "." in clean:
            clean = clean.replace(".", "").replace(",", ".")
        elif "," in clean:
            clean = clean.replace(",", ".")
        try:
            return float(clean)
        except ValueError:
            return 0.0

    @staticmethod
    def _fmt(value):
        text = f"{float(value or 0):,.2f}"
        return text.replace(",", "X").replace(".", ",").replace("X", ".")
